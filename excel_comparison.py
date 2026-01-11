#!/usr/bin/env python3
"""
Excel File Comparison Tool
Compare Patient IDs between two Excel files and add insurance columns
"""

from flask import Flask, render_template_string, request, jsonify, send_file, redirect
import pandas as pd
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import re

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max file size

# Global variables to store session data
raw_data = None
previous_data = None
raw_filename = None
previous_filename = None
comparison_result = None

# Global variables for conversion report
conversion_data = None
conversion_filename = None
conversion_result = None

# Global variables for insurance name formatting
insurance_formatting_data = None
insurance_formatting_filename = None
insurance_formatting_result = None
insurance_formatting_output = ""

# Global variables for remarks update
remarks_appointments_data = None
remarks_excel_data = None
remarks_appointments_filename = None
remarks_remarks_filename = None
remarks_result = None
remarks_updated_count = 0

# Global variables for appointment report formatting
appointment_report_data = None
appointment_report_filename = None
appointment_report_result = None
appointment_report_output = ""

# Global variables for smart assist report formatting
smart_assist_data = None
smart_assist_filename = None
smart_assist_result = None
smart_assist_output = ""

# Global variables for consolidate report
consolidate_master_data = None
consolidate_daily_data = None
consolidate_master_filename = None
consolidate_daily_filename = None
consolidate_result = None
consolidate_output = ""

# Global variables for reallocation data generation
reallocation_consolidate_data = None
reallocation_blank_data = None
reallocation_consolidate_filename = None
reallocation_blank_filename = None
reallocation_result = None
reallocation_output = ""
reallocation_merged_data = None
reallocation_available_remarks = []
reallocation_available_agents = []
reallocation_selected_remarks = []
reallocation_selected_agents = []

# Global variables for general comparison
general_primary_data = None
general_main_data = None
general_primary_filename = None
general_main_filename = None
general_comparison_result = None
general_comparison_output = ""
general_comparison_updated_data = None


@app.route("/load_reallocation_consolidate", methods=["POST"])
def load_reallocation_consolidate():
    """Load the Current Consolidate File, align remarks, and return available remarks/agents as JSON."""
    global reallocation_consolidate_data, reallocation_consolidate_filename, reallocation_available_remarks, reallocation_available_agents

    try:
        if "consolidate_file" not in request.files:
            return jsonify({"ok": False, "error": "No consolidate file provided."}), 400

        consolidate_file = request.files["consolidate_file"]
        if not consolidate_file or consolidate_file.filename == "":
            return jsonify({"ok": False, "error": "Empty consolidate file."}), 400

        consolidate_filename = secure_filename(consolidate_file.filename)
        consolidate_filepath = os.path.join("/tmp", consolidate_filename)
        consolidate_file.save(consolidate_filepath)

        # Load and align remarks on all sheets
        consolidate_xls = pd.ExcelFile(consolidate_filepath)
        reallocation_consolidate_data = {}
        for sheet in consolidate_xls.sheet_names:
            df = pd.read_excel(consolidate_filepath, sheet_name=sheet)
            if "Remark" in df.columns:
                df["Remark"] = df["Remark"].apply(align_remark)
            reallocation_consolidate_data[sheet] = df

        reallocation_consolidate_filename = consolidate_filename

        # Build available remark list from All Agent Data if present
        remarks = []
        agents = []
        if "All Agent Data" in reallocation_consolidate_data:
            df = reallocation_consolidate_data["All Agent Data"]
            if "Remark" in df.columns:
                remarks = sorted(
                    pd.Series(
                        df["Remark"].dropna().astype(str).str.strip().unique()
                    ).tolist()
                )
            # Only include agents who have Workable remark
            if "Agent Name" in df.columns and "Remark" in df.columns:
                workable_df = df[df["Remark"].astype(str).str.strip() == "Workable"]
                agents = sorted(
                    pd.Series(
                        workable_df["Agent Name"]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .unique()
                    ).tolist()
                )

        reallocation_available_remarks = remarks
        reallocation_available_agents = agents

        return jsonify(
            {
                "ok": True,
                "filename": reallocation_consolidate_filename,
                "remarks": reallocation_available_remarks,
                "agents": reallocation_available_agents,
            }
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def align_remark(remark_value):
    """
    Align remark values to standardized forms.
    Maps various remark variations to their standardized equivalents.
    """
    if pd.isna(remark_value) or remark_value == "":
        return remark_value

    # Convert to string and strip whitespace
    remark_str = str(remark_value).strip()

    # Define mapping rules (case-insensitive)
    remark_lower = remark_str.lower()

    # Direct exact matches first
    exact_mappings = {
        "ast": "ATS",
        "ntpb": "NTBP",
    }

    if remark_lower in exact_mappings:
        return exact_mappings[remark_lower]

    # Pattern-based mappings (handle variations with slashes and spaces)
    if "ats" in remark_lower and "qcp" in remark_lower:
        return "ATS"

    if "qcp" in remark_lower and "ats" in remark_lower:
        return "ATS"

    if "qcp" in remark_lower and "hold" in remark_lower:
        return "QCP"

    # QCP variations (with or without slashes/spaces)
    if remark_lower.startswith("qcp") and (
        remark_lower.endswith("qcp") or "/" in remark_lower
    ):
        return "QCP"

    # Keep original if no match found
    return remark_value


# HTML Template for Excel Comparison
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel File Comparison Tool</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: #f5f7fa;
            min-height: 100vh;
            overflow-x: hidden;
        }
        
        /* Layout Structure */
        .app-container {
            display: flex;
            min-height: 100vh;
        }
        
        /* Sidebar */
        .sidebar {
            width: 280px;
            background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
            color: white;
            position: fixed;
            height: 100vh;
            overflow-y: auto;
            box-shadow: 4px 0 10px rgba(0,0,0,0.1);
            z-index: 1000;
            transition: transform 0.3s ease;
        }
        
        .sidebar-header {
            padding: 30px 20px;
            border-bottom: 1px solid rgba(255,255,255,0.1);
            text-align: center;
        }
        
        .sidebar-header h1 {
            font-size: 1.5em;
            margin-bottom: 5px;
            font-weight: 700;
        }
        
        .sidebar-header p {
            font-size: 0.9em;
            opacity: 0.8;
        }
        
        .sidebar-menu {
            padding: 20px 0;
        }
        
        .menu-item {
            padding: 15px 25px;
            margin: 5px 15px;
            border-radius: 10px;
            cursor: pointer;
            transition: all 0.3s ease;
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 1em;
            font-weight: 500;
            color: rgba(255,255,255,0.8);
            border: 2px solid transparent;
        }
        
        .menu-item:hover {
            background: rgba(255,255,255,0.1);
            color: white;
            transform: translateX(5px);
        }
        
        .menu-item.active {
            background: rgba(255,255,255,0.2);
            color: white;
            border-color: rgba(255,255,255,0.3);
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        }
        
        .menu-item-icon {
            font-size: 1.3em;
            width: 25px;
            text-align: center;
        }
        
        /* Main Content */
        .main-content {
            margin-left: 280px;
            flex: 1;
            display: flex;
            flex-direction: column;
            min-height: 100vh;
        }
        
        /* Header */
        .header {
            background: white;
            padding: 25px 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
            position: sticky;
            top: 0;
            z-index: 100;
        }
        
        .header h2 {
            color: #333;
            font-size: 1.8em;
            margin-bottom: 5px;
        }
        
        .header p {
            color: #666;
            font-size: 1em;
        }
        
        /* Content Area */
        .content {
            flex: 1;
            padding: 30px 40px;
            max-width: 1400px;
            width: 100%;
        }
        
        /* Footer */
        .footer {
            background: white;
            padding: 20px 40px;
            text-align: center;
            color: #666;
            border-top: 1px solid #e0e0e0;
            margin-top: auto;
        }
        
        .footer p {
            font-size: 0.9em;
        }
        
        /* Mobile Toggle */
        .mobile-toggle {
            display: none;
            position: fixed;
            top: 20px;
            left: 20px;
            z-index: 1001;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            width: 50px;
            height: 50px;
            border-radius: 50%;
            cursor: pointer;
            box-shadow: 0 4px 15px rgba(0,0,0,0.3);
            font-size: 1.5em;
        }
        
        /* Tab Content */
        .tab-content {
            display: none;
        }
        
        .tab-content.active {
            display: block;
            animation: fadeIn 0.3s ease;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        /* Sections */
        .section { 
            margin: 25px 0; 
            padding: 25px; 
            border: 1px solid #e0e0e0; 
            border-radius: 10px; 
            background: white;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }
        
        .section h3 { 
            color: #333; 
            margin-bottom: 20px; 
            font-size: 1.4em;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
        }
        
        .form-group { margin: 15px 0; }
        
        label { 
            display: block; 
            margin-bottom: 8px; 
            font-weight: 600; 
            color: #555;
        }
        
        input[type="file"], input[type="text"], select { 
            width: 100%; 
            padding: 12px; 
            border: 2px solid #ddd; 
            border-radius: 8px; 
            font-size: 16px;
            transition: border-color 0.3s;
        }
        
        input[type="file"]:focus, input[type="text"]:focus, select:focus { 
            outline: none; 
            border-color: #667eea; 
        }
        
        button { 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            padding: 12px 25px; 
            border: none; 
            border-radius: 8px; 
            cursor: pointer; 
            margin: 5px; 
            font-size: 16px;
            font-weight: 600;
            transition: transform 0.2s;
        }
        
        button:hover { 
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        
        .file-status { 
            background: #f8f9fa; 
            padding: 15px; 
            border-radius: 8px; 
            margin: 15px 0; 
            border-left: 4px solid #667eea;
        }
        
        .status-success { 
            background: #d4edda; 
            color: #155724; 
            border-color: #c3e6cb; 
        }
        
        .status-info { 
            background: #d1ecf1; 
            color: #0c5460; 
            border-color: #bee5eb; 
        }
        
        .status-message {
            background: #f3e5f5;
            border: 2px solid #9c27b0;
            color: #4a148c;
            padding: 20px;
            border-radius: 10px;
            margin: 15px 0;
            font-size: 16px;
            line-height: 1.6;
            white-space: pre-line;
            box-shadow: 0 2px 8px rgba(156, 39, 176, 0.2);
        }
        
        .reset-btn {
            background: linear-gradient(135deg, #ff6b6b, #ee5a52);
            color: white;
            border: none;
            padding: 12px 24px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            transition: all 0.3s ease;
            box-shadow: 0 4px 15px rgba(255, 107, 107, 0.3);
        }
        
        .reset-btn:hover {
            background: linear-gradient(135deg, #ff5252, #d32f2f);
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(255, 107, 107, 0.4);
        }
        
        .comparison-results { 
            background: #1e1e1e; 
            color: #f8f8f2; 
            padding: 20px; 
            border-radius: 8px; 
            white-space: pre-wrap; 
            font-family: 'Courier New', monospace; 
            max-height: 400px; 
            overflow-y: auto;
            border: 1px solid #333;
        }
        
        .two-column { 
            display: grid; 
            grid-template-columns: 1fr 1fr; 
            gap: 20px; 
        }
        
        .loading { 
            display: none; 
            text-align: center; 
            padding: 20px; 
        }
        
        .spinner { 
            border: 4px solid #f3f3f3; 
            border-top: 4px solid #667eea; 
            border-radius: 50%; 
            width: 40px; 
            height: 40px; 
            animation: spin 1s linear infinite; 
            margin: 0 auto;
        }
        
        @keyframes spin { 
            0% { transform: rotate(0deg); } 
            100% { transform: rotate(360deg); } 
        }
        
        /* Processing Modal Overlay */
        .processing-modal-overlay {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.7);
            backdrop-filter: blur(5px);
            z-index: 9999;
            justify-content: center;
            align-items: center;
        }
        
        .processing-modal-overlay.show {
            display: flex;
        }
        
        .processing-modal {
            background: white;
            border-radius: 20px;
            padding: 40px;
            max-width: 400px;
            width: 90%;
            text-align: center;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            animation: slideUp 0.3s ease;
        }
        
        @keyframes slideUp {
            from {
                transform: translateY(30px);
                opacity: 0;
            }
            to {
                transform: translateY(0);
                opacity: 1;
            }
        }
        
        .processing-modal .modal-spinner {
            border: 5px solid #f3f3f3;
            border-top: 5px solid #667eea;
            border-radius: 50%;
            width: 60px;
            height: 60px;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }
        
        .processing-modal h3 {
            color: #333;
            margin-bottom: 15px;
            font-size: 1.5em;
        }
        
        .processing-modal p {
            color: #666;
            font-size: 1.1em;
            margin: 0;
        }
        
        
        .processing-modal .processing-dots {
            display: inline-block;
            animation: dots 1.5s steps(4, end) infinite;
        }
        
        .processing-modal .processing-dots::after {
            content: '...';
            animation: dots 1.5s steps(4, end) infinite;
        }
        
        @keyframes dots {
            0%, 20% { content: '.'; }
            40% { content: '..'; }
            60%, 100% { content: '...'; }
        }
        
        /* Toast Notification */
        .toast {
            position: fixed;
            top: 20px;
            right: 20px;
            background: #4caf50;
            color: white;
            padding: 16px 24px;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            z-index: 10000;
            display: none;
            max-width: 400px;
        }
        .toast.error {
            background: #f44336;
        }
        .toast.show {
            display: block;
            animation: slideInRight 0.3s ease;
        }
        @keyframes slideInRight {
            from {
                transform: translateX(400px);
                opacity: 0;
            }
            to {
                transform: translateX(0);
                opacity: 1;
            }
        }
        .toast-close {
            float: right;
            font-weight: bold;
            cursor: pointer;
            margin-left: 15px;
        }
        
        /* Responsive Design */
        @media (max-width: 968px) {
            .sidebar {
                transform: translateX(-100%);
            }
            
            .sidebar.active {
                transform: translateX(0);
            }
            
            .main-content {
                margin-left: 0;
            }
            
            .mobile-toggle {
                display: block;
            }
            
            .two-column {
                grid-template-columns: 1fr;
            }
        }
    </style>
</head>
<body>
    <!-- Toast Notification -->
    <div id="toast" class="toast">
        <span id="toast-message"></span>
        <span class="toast-close" onclick="hideToast()">&times;</span>
    </div>

    <!-- Processing Modal Overlay -->
    <div id="processing-modal-overlay" class="processing-modal-overlay">
        <div class="processing-modal">
            <div class="modal-spinner"></div>
            <h3 id="processing-title">Processing</h3>
            <p id="processing-message">Please wait while we process your request<span class="processing-dots">...</span></p>
        </div>
    </div>

    <!-- Mobile Toggle Button -->
    <button class="mobile-toggle" onclick="toggleSidebar()">☰</button>

    <div class="app-container">
        <!-- Sidebar -->
        <div class="sidebar" id="sidebar">
            <div class="sidebar-header">
                <h1>📊 Excel Tools</h1>
                <p>Automation Suite</p>
            </div>
            <nav class="sidebar-menu">
                <div class="menu-item {% if active_tab == 'comparison' %}active{% endif %}" onclick="switchTab('comparison')">
                    <span class="menu-item-icon">🔄</span>
                    <span>Comparison Tool</span>
                </div>
                <div class="menu-item {% if active_tab == 'conversion' %}active{% endif %}" onclick="switchTab('conversion')">
                    <span class="menu-item-icon">📋</span>
                    <span>Conversion Report</span>
                </div>
                <div class="menu-item {% if active_tab == 'insurance' %}active{% endif %}" onclick="switchTab('insurance')">
                    <span class="menu-item-icon">🦷</span>
                    <span>Insurance Formatting</span>
                </div>
                <div class="menu-item {% if active_tab == 'remarks' %}active{% endif %}" onclick="switchTab('remarks')">
                    <span class="menu-item-icon">📝</span>
                    <span>Update Remarks</span>
                </div>
                <div class="menu-item {% if active_tab == 'appointment' %}active{% endif %}" onclick="switchTab('appointment')">
                    <span class="menu-item-icon">📅</span>
                    <span>Appointment Report</span>
                </div>
                <div class="menu-item {% if active_tab == 'smartassist' %}active{% endif %}" onclick="switchTab('smartassist')">
                    <span class="menu-item-icon">🤖</span>
                    <span>Smart Assist Report</span>
                </div>
                <!-- <div class="menu-item {% if active_tab == 'consolidate' %}active{% endif %}" onclick="switchTab('consolidate')">
                    <span class="menu-item-icon">📊</span>
                    <span>Consolidate Report</span>
                </div> -->
                <div class="menu-item {% if active_tab == 'reallocation' %}active{% endif %}" onclick="switchTab('reallocation')">
                    <span class="menu-item-icon">♻️</span>
                    <span>Generate Reallocation Data</span>
                </div>
                <div class="menu-item {% if active_tab == 'general' %}active{% endif %}" onclick="switchTab('general')">
                    <span class="menu-item-icon">🧭</span>
                    <span>General Comparison</span>
                </div>
            </nav>
        </div>

        <!-- Main Content -->
        <div class="main-content">
            <div class="header">
                <h2 id="page-title">
                    {% if active_tab == 'comparison' %}🔄 Comparison Tool
                    {% elif active_tab == 'conversion' %}📋 Conversion Report Formatting
                    {% elif active_tab == 'insurance' %}🦷 Insurance Name Formatting
                    {% elif active_tab == 'remarks' %}📝 Update Remarks
                    {% elif active_tab == 'appointment' %}📅 Appointment Report Formatting
                    {% elif active_tab == 'smartassist' %}🤖 Smart Assist Report Formatting
                    {% elif active_tab == 'consolidate' %}📊 Consolidate Report
                    {% elif active_tab == 'reallocation' %}♻️ Generate Reallocation Data
                    {% elif active_tab == 'general' %}🧭 General Comparison
                    {% else %}🔄 Comparison Tool
                    {% endif %}
                </h2>
                <p id="page-description">
                    {% if active_tab == 'comparison' %}Compare Patient IDs and add insurance columns
                    {% elif active_tab == 'conversion' %}Format and process conversion reports
                    {% elif active_tab == 'insurance' %}Standardize insurance carrier names
                    {% elif active_tab == 'remarks' %}Add remarks to appointments
                    {% elif active_tab == 'appointment' %}Format appointment report insurance columns
                    {% elif active_tab == 'smartassist' %}Format smart assist report insurance columns
                    {% elif active_tab == 'consolidate' %}Consolidate master and daily report files
                    {% elif active_tab == 'reallocation' %}Generate reallocation data from consolidate file
                    {% elif active_tab == 'general' %}Compare two files and update primary rows on match
                    {% else %}Compare Patient IDs and add insurance columns
                    {% endif %}
                </p>
            </div>

            <div class="content">
            <!-- Tab 1: Comparison Tool -->
            <div id="comparison-tab" class="tab-content {% if active_tab == 'comparison' %}active{% endif %}">
                <div class="section">
                    <h3>🔄 Comparison Tool</h3>
                    <p>Compare Patient ID from Appointment Report with PATID from Smart Assist file. When matched, insurance columns will be added to the Smart Assist file.</p>
                </div>

            <!-- File Upload Section -->
            <div class="section">
                <h3>📁 Upload Excel Files</h3>
                <div class="two-column">
                    <div>
                        <h4>Appointment Report File</h4>
                        <form action="/upload_raw" method="post" enctype="multipart/form-data" id="raw-form">
                            <div class="form-group">
                                <label for="raw_file">Select Appointment Report Excel File:</label>
                                <input type="file" id="raw_file" name="file" accept=".xlsx,.xls" required>
                            </div>
                            <button type="submit" id="raw-btn">📤 Upload Appointment Report</button>
                        </form>
                        <div class="loading" id="raw-loading">
                            <div class="spinner"></div>
                            <p>Processing Appointment Report...</p>
                        </div>
                    </div>
                    
                    <div>
                        <h4>Smart Assist File</h4>
                        <form action="/upload_previous" method="post" enctype="multipart/form-data" id="previous-form">
                            <div class="form-group">
                                <label for="previous_file">Select Smart Assist Excel File:</label>
                                <input type="file" id="previous_file" name="file" accept=".xlsx,.xls" required>
                            </div>
                            <button type="submit" id="previous-btn">📤 Upload Smart Assist File</button>
                        </form>
                        <div class="loading" id="previous-loading">
                            <div class="spinner"></div>
                            <p>Processing Smart Assist file...</p>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Status Messages -->
            {% if comparison_result %}
            <div class="section">
                <h3>📢 Status Messages</h3>
                <div class="status-message">
                    {{ comparison_result | safe }}
                </div>
            </div>
            {% endif %}

            <!-- File Status -->
            <div class="section">
                <h3>📊 File Status</h3>
                <div class="file-status">
                    {% if raw_filename %}
                        <div class="status-success">
                            ✅ Appointment Report: {{ raw_filename }}<br>
                            📋 Sheets: {{ raw_data.keys() | list | length if raw_data else 0 }}
                        </div>
                    {% else %}
                        <div class="status-info">
                            ℹ️ No Appointment Report file uploaded yet.
                        </div>
                    {% endif %}
                    
                    {% if previous_filename %}
                        <div class="status-success">
                            ✅ Smart Assist File: {{ previous_filename }}<br>
                            📋 Sheets: {{ previous_data.keys() | list | length if previous_data else 0 }}
                        </div>
                    {% else %}
                        <div class="status-info">
                            ℹ️ No Smart Assist file uploaded yet.
                        </div>
                    {% endif %}
                </div>
            </div>

            <!-- Comparison Section -->
            {% if raw_data and previous_data %}
            <div class="section">
                <h3>🔄 Compare Files</h3>
                <form action="/compare" method="post" id="compare-form">
                    <div class="form-group">
                        <label for="raw_sheet">Select Raw File Sheet:</label>
                        <select id="raw_sheet" name="raw_sheet" required style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                            {% for sheet_name in raw_data.keys() %}
                                <option value="{{ sheet_name }}">{{ sheet_name }}</option>
                            {% endfor %}
                        </select>
                    </div>
                    <div class="form-group">
                        <label for="previous_sheet">Select Previous File Sheet:</label>
                        <select id="previous_sheet" name="previous_sheet" required style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                            {% for sheet_name in previous_data.keys() %}
                                <option value="{{ sheet_name }}">{{ sheet_name }}</option>
                            {% endfor %}
                        </select>
                    </div>
                    <button type="submit" id="compare-btn">🔄 Compare and Add Insurance Columns</button>
                </form>
                <div class="loading" id="compare-loading">
                    <div class="spinner"></div>
                    <p>Comparing files...</p>
                </div>
            </div>
            {% endif %}

            <!-- Results Section -->
            <div class="section">
                <h3>📝 Comparison Results</h3>
                <div class="comparison-results" id="results">
                    {% if comparison_result %}
                        {{ comparison_result }}
                    {% else %}
                        Upload both files and click "Compare and Add Insurance Columns" to see results...
                    {% endif %}
                </div>
            </div>

            <!-- Download Section -->
            {% if comparison_result and 'Comparison completed successfully' in comparison_result %}
            <div class="section">
                <h3>💾 Download Results</h3>
                <form action="/download_result" method="post">
                    <div class="form-group">
                        <label for="output_filename">Output filename (optional):</label>
                        <input type="text" id="output_filename" name="filename" 
                               placeholder="comparison_result.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                    </div>
                    <button type="submit">💾 Download Result File</button>
                </form>
            </div>
            {% endif %}

            <!-- Reset Section -->
            <div class="section">
                <h3>🔄 Reset Comparison Tool</h3>
                <p>Clear all uploaded files and reset the comparison tool to start fresh.</p>
                <form action="/reset_comparison" method="post" onsubmit="return confirm('Are you sure you want to reset the comparison tool? This will clear all uploaded files and data.')">
                    <button type="submit" class="reset-btn">🗑️ Reset Comparison Tool</button>
                </form>
            </div>
            </div>

            <!-- Tab 2: Conversion Report Formatting -->
            <div id="conversion-tab" class="tab-content {% if active_tab == 'conversion' %}active{% endif %}">
                <div class="section">
                    <h3>📋 Conversion Report Formatting</h3>
                    <p>Upload an Excel file to process and format conversion reports. The file must contain an "Insurance Note" column.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel File</h3>
                    <form action="/upload_conversion" method="post" enctype="multipart/form-data" id="conversion-form">
                        <div class="form-group">
                            <label for="conversion_file">Select Conversion Report Excel File:</label>
                            <input type="file" id="conversion_file" name="file" accept=".xlsx,.xls" required>
                        </div>
                        <button type="submit" id="conversion-btn">📤 Upload & Validate File</button>
                    </form>
                    <div class="loading" id="conversion-loading">
                        <div class="spinner"></div>
                        <p>Processing conversion report...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if conversion_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ conversion_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if conversion_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ Conversion Report: {{ conversion_filename }}<br>
                            📋 Sheets: {{ conversion_data.keys() | list | length if conversion_data else 0 }}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if conversion_data and conversion_result and 'processing completed successfully' in conversion_result.lower() %}
                <div class="section">
                    <h3>💾 Download Processed File</h3>
                    <form action="/download_conversion" method="post">
                        <div class="form-group">
                            <label for="conversion_output_filename">Output filename (optional):</label>
                            <input type="text" id="conversion_output_filename" name="filename" 
                                   placeholder="conversion_report_formatted.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Processed File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Conversion Tool</h3>
                    <p>Clear all uploaded files and reset the conversion tool to start fresh.</p>
                    <form action="/reset_conversion" method="post" onsubmit="return confirm('Are you sure you want to reset the conversion tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Conversion Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 3: Insurance Name Formatting -->
            <div id="insurance-tab" class="tab-content {% if active_tab == 'insurance' %}active{% endif %}">
                <div class="section">
                    <h3>🦷 Insurance Name Formatting</h3>
                    <p>Upload an Excel file to automatically format "Dental Primary Ins Carr" column in all sheets. A new formatted column will be created.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel File</h3>
                    <form action="/upload_insurance_formatting" method="post" enctype="multipart/form-data" id="insurance-form">
                        <div class="form-group">
                            <label for="insurance_file">Select Excel File (.xlsx, .xls):</label>
                            <input type="file" id="insurance_file" name="file" accept=".xlsx,.xls" required>
                        </div>
                        <button type="submit" id="insurance-btn">📤 Upload & Process File</button>
                    </form>
                    <div class="loading" id="insurance-loading">
                        <div class="spinner"></div>
                        <p>Processing file and formatting columns...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if insurance_formatting_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ insurance_formatting_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if insurance_formatting_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ File loaded: {{ insurance_formatting_filename }}<br>
                            📋 Sheets processed: {{ insurance_formatting_data.keys() | list | length if insurance_formatting_data else 0 }}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if insurance_formatting_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ insurance_formatting_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if insurance_formatting_data and insurance_formatting_result and 'processing complete' in insurance_formatting_result.lower() %}
                <div class="section">
                    <h3>💾 Download Formatted File</h3>
                    <form action="/download_insurance_formatting" method="post">
                        <div class="form-group">
                            <label for="insurance_output_filename">Output filename (optional):</label>
                            <input type="text" id="insurance_output_filename" name="filename" 
                                   placeholder="formatted_insurance_names.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Formatted File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Insurance Formatting Tool</h3>
                    <p>Clear all uploaded files and reset the insurance formatting tool to start fresh.</p>
                    <form action="/reset_insurance_formatting" method="post" onsubmit="return confirm('Are you sure you want to reset the insurance formatting tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Insurance Formatting Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 4: Update Remarks -->
            <div id="remarks-tab" class="tab-content {% if active_tab == 'remarks' %}active{% endif %}">
                <div class="section">
                    <h3>📝 Update Remarks</h3>
                    <p>Upload two Excel files: Appointments file (with Pat ID) and Remarks file (with Patient ID, Remark, and Agent Name). The tool will match Patient IDs and update appointments with remarks.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel Files</h3>
                    <form action="/upload_remarks" method="post" enctype="multipart/form-data" id="remarks-form">
                        <div class="two-column">
                            <div>
                                <h4>Appointments Excel File</h4>
                                <div class="form-group">
                                    <label for="appointments_file">Select Appointments Excel File (must have "Pat ID" column):</label>
                                    <input type="file" id="appointments_file" name="appointments_file" accept=".xlsx,.xls" required>
                                </div>
                                {% if remarks_appointments_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ remarks_appointments_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                            
                            <div>
                                <h4>Remarks Excel File</h4>
                                <div class="form-group">
                                    <label for="remarks_file">Select Remarks Excel File (must have "Patient ID", "Remark", and "Agent Name" columns):</label>
                                    <input type="file" id="remarks_file" name="remarks_file" accept=".xlsx,.xls" required>
                                </div>
                                {% if remarks_remarks_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ remarks_remarks_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                        </div>
                        <button type="submit" id="remarks-btn">📤 Upload & Process Files</button>
                    </form>
                    <div class="loading" id="remarks-loading">
                        <div class="spinner"></div>
                        <p>Processing files and matching Patient IDs...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if remarks_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ remarks_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if remarks_appointments_data %}
                <div class="section">
                    <h3>📊 Processing Results</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ Total appointments processed: {{ remarks_appointments_data | length }}<br>
                            ✅ Appointments updated with remarks: {{ remarks_updated_count }}<br>
                            {% if remarks_appointments_filename %}
                            📄 Appointments file: {{ remarks_appointments_filename }}<br>
                            {% endif %}
                            {% if remarks_remarks_filename %}
                            📄 Remarks file: {{ remarks_remarks_filename }}
                            {% endif %}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if remarks_appointments_data and remarks_result and 'successfully' in remarks_result.lower() %}
                <div class="section">
                    <h3>💾 Download Updated File</h3>
                    <form action="/download_remarks" method="post">
                        <div class="form-group">
                            <label for="remarks_output_filename">Output filename (optional):</label>
                            <input type="text" id="remarks_output_filename" name="filename" 
                                   placeholder="appointments_with_remarks.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Updated File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Remarks Tool</h3>
                    <p>Clear all uploaded files and reset the remarks tool to start fresh.</p>
                    <form action="/reset_remarks" method="post" onsubmit="return confirm('Are you sure you want to reset the remarks tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Remarks Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 5: Appointment Report Formatting -->
            <div id="appointment-tab" class="tab-content {% if active_tab == 'appointment' %}active{% endif %}">
                <div class="section">
                    <h3>📅 Appointment Report Formatting</h3>
                    <p>Upload an Appointment Report Excel file to automatically format insurance names in "Dental Primary Ins Carr" and "Dental Secondary Ins Carr" columns.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel File</h3>
                    <form action="/upload_appointment_report" method="post" enctype="multipart/form-data" id="appointment-form">
                        <div class="form-group">
                            <label for="appointment_file">Select Appointment Report Excel File (.xlsx, .xls):</label>
                            <input type="file" id="appointment_file" name="file" accept=".xlsx,.xls" required>
                        </div>
                        <button type="submit" id="appointment-btn">📤 Upload & Format File</button>
                    </form>
                    <div class="loading" id="appointment-loading">
                        <div class="spinner"></div>
                        <p>Processing and formatting insurance columns...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if appointment_report_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ appointment_report_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if appointment_report_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ File loaded: {{ appointment_report_filename }}<br>
                            📋 Sheets processed: {{ appointment_report_data.keys() | list | length if appointment_report_data else 0 }}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if appointment_report_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ appointment_report_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if appointment_report_data and appointment_report_result and 'error' not in appointment_report_result.lower() %}
                <div class="section">
                    <h3>💾 Download Formatted File</h3>
                    <form action="/download_appointment_report" method="post">
                        <div class="form-group">
                            <label for="appointment_output_filename">Output filename (optional):</label>
                            <input type="text" id="appointment_output_filename" name="filename" 
                                   placeholder="formatted_appointment_report.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Formatted File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Appointment Report Tool</h3>
                    <p>Clear all uploaded files and reset the appointment report formatting tool to start fresh.</p>
                    <form action="/reset_appointment_report" method="post" onsubmit="return confirm('Are you sure you want to reset the appointment report formatting tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Appointment Report Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 6: Smart Assist Report Formatting -->
            <div id="smartassist-tab" class="tab-content {% if active_tab == 'smartassist' %}active{% endif %}">
                <div class="section">
                    <h3>🤖 Smart Assist Report Formatting</h3>
                    <p>Upload a Smart Assist Report Excel file to automatically format columns to create allocation file </p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel File</h3>
                    <form action="/upload_smart_assist" method="post" enctype="multipart/form-data" id="smartassist-form">
                        <div class="form-group">
                            <label for="smartassist_file">Select Smart Assist Report Excel File (.xlsx, .xls):</label>
                            <input type="file" id="smartassist_file" name="file" accept=".xlsx,.xls" required>
                        </div>
                        <button type="submit" id="smartassist-btn">📤 Upload & Format File</button>
                    </form>
                    <div class="loading" id="smartassist-loading">
                        <div class="spinner"></div>
                        <p>Processing and formatting insurance columns...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if smart_assist_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ smart_assist_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if smart_assist_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ File loaded: {{ smart_assist_filename }}<br>
                            📋 Sheets processed: {{ smart_assist_data.keys() | list | length if smart_assist_data else 0 }}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if smart_assist_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ smart_assist_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if smart_assist_data and smart_assist_result and 'processing complete' in smart_assist_result.lower() %}
                <div class="section">
                    <h3>💾 Download Formatted File</h3>
                    <form action="/download_smart_assist" method="post">
                        <div class="form-group">
                            <label for="smartassist_output_filename">Output filename (optional):</label>
                            <input type="text" id="smartassist_output_filename" name="filename" 
                                   placeholder="formatted_smart_assist_report.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Formatted File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Smart Assist Report Tool</h3>
                    <p>Clear all uploaded files and reset the smart assist report formatting tool to start fresh.</p>
                    <form action="/reset_smart_assist" method="post" onsubmit="return confirm('Are you sure you want to reset the smart assist report formatting tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Smart Assist Report Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 7: Consolidate Report -->
            <div id="consolidate-tab" class="tab-content {% if active_tab == 'consolidate' %}active{% endif %}">
                <div class="section">
                    <h3>📊 Consolidate Report</h3>
                    <p>Upload two Excel files: Master Consolidate file and Daily Consolidated file to merge and consolidate the data.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel Files</h3>
                    <form action="/upload_consolidate" method="post" enctype="multipart/form-data" id="consolidate-form">
                        <div class="two-column">
                            <div>
                                <h4>Master Consolidate File</h4>
                                <div class="form-group">
                                    <label for="consolidate_master_file">Select Master Consolidate Excel File:</label>
                                    <input type="file" id="consolidate_master_file" name="master_file" accept=".xlsx,.xls" required>
                                </div>
                                {% if consolidate_master_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ consolidate_master_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                            
                            <div>
                                <h4>Daily Consolidated File</h4>
                                <div class="form-group">
                                    <label for="consolidate_daily_file">Select Daily Consolidated Excel File:</label>
                                    <input type="file" id="consolidate_daily_file" name="daily_file" accept=".xlsx,.xls" required>
                                </div>
                                {% if consolidate_daily_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ consolidate_daily_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                        </div>
                        <button type="submit" id="consolidate-btn">📤 Upload & Consolidate Files</button>
                    </form>
                    <div class="loading" id="consolidate-loading">
                        <div class="spinner"></div>
                        <p>Processing and consolidating files...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if consolidate_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ consolidate_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if consolidate_master_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ Master File: {{ consolidate_master_filename }}<br>
                            {% if consolidate_daily_filename %}
                            ✅ Daily File: {{ consolidate_daily_filename }}
                            {% endif %}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if consolidate_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ consolidate_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if consolidate_master_data and consolidate_result and 'consolidation complete' in consolidate_result.lower() %}
                <div class="section">
                    <h3>💾 Download Consolidated File</h3>
                    <form action="/download_consolidate" method="post">
                        <div class="form-group">
                            <label for="consolidate_output_filename">Output filename (optional):</label>
                            <input type="text" id="consolidate_output_filename" name="filename" 
                                   placeholder="consolidated_report.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Consolidated File</button>
                    </form>
                </div>
                <div class="section">
                    <h3>💾 Download Merged Sheet Only</h3>
                    <p>Downloads just the merged <strong>consolidated</strong> sheet.</p>
                    <form action="/download_consolidate_consolidated_only" method="post">
                        <div class="form-group">
                            <label for="consolidate_only_output_filename">Output filename (optional):</label>
                            <input type="text" id="consolidate_only_output_filename" name="filename" 
                                   placeholder="consolidated_only.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Consolidated Sheet</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Consolidate Report Tool</h3>
                    <p>Clear all uploaded files and reset the consolidate report tool to start fresh.</p>
                    <form action="/reset_consolidate" method="post" onsubmit="return confirm('Are you sure you want to reset the consolidate report tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Consolidate Report Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 8: Generate Reallocation Data -->
            <div id="reallocation-tab" class="tab-content {% if active_tab == 'reallocation' %}active{% endif %}">
                <div class="section">
                    <h3>♻️ Generate Reallocation Data</h3>
                    <p>Upload a consolidated file and blank allocation file to generate reallocation data.</p>
                </div>

                <!-- File Upload Section -->
                <div class="section">
                    <h3>📁 Upload Excel Files</h3>
                    <form action="/upload_reallocation" method="post" enctype="multipart/form-data" id="reallocation-form">
                        <div class="two-column">
                            <div>
                                <h4>Current Consolidate File</h4>
                                <div class="form-group">
                                    <label for="reallocation_consolidate_file">Select Consolidate Excel File:</label>
                                    <input type="file" id="reallocation_consolidate_file" name="consolidate_file" accept=".xlsx,.xls" {% if not reallocation_consolidate_filename %}required{% endif %}>
                                </div>
                                {% if reallocation_consolidate_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ reallocation_consolidate_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                            
                            <div>
                                <h4>Blank Allocation File</h4>
                                <div class="form-group">
                                    <label for="reallocation_blank_file">Select Blank Allocation Excel File:</label>
                                    <input type="file" id="reallocation_blank_file" name="blank_file" accept=".xlsx,.xls" {% if not reallocation_blank_filename %}required{% endif %}>
                                </div>
                                {% if reallocation_blank_filename %}
                                <div class="file-status">
                                    <div class="status-success">
                                        ✅ {{ reallocation_blank_filename }}
                                    </div>
                                </div>
                                {% endif %}
                            </div>
                        </div>
                        <!-- Remark Filter Dropdown -->
                        <div class="form-group" style="margin-top: 12px;">
                            <label for="reallocation_remarks">Filter by Remark (select one or more):</label>
                            <select id="reallocation_remarks" name="remarks_filter" multiple size="6" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;" {% if not (reallocation_available_remarks and reallocation_available_remarks|length > 0) %}disabled{% endif %}>
                                {% if reallocation_available_remarks and reallocation_available_remarks|length > 0 %}
                                    {% for r in reallocation_available_remarks %}
                                    <option value="{{ r }}" {% if reallocation_selected_remarks and r in reallocation_selected_remarks %}selected{% endif %}>{{ r }}</option>
                                    {% endfor %}
                                {% else %}
                                    <option value="" disabled>(Upload the Consolidate file to load remark options)</option>
                                {% endif %}
                            </select>
                            <small>Select at least one remark or a Workable agent to enable generation.</small>
                        </div>

                        <div class="form-group" style="margin-top: 12px;">
                            <label for="reallocation_agents">Filter Workable rows by Agent (select one or more):</label>
                            <select id="reallocation_agents" name="agent_filter" multiple size="6" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;" {% if not (reallocation_available_agents and reallocation_available_agents|length > 0) %}disabled{% endif %}>
                                {% if reallocation_available_agents and reallocation_available_agents|length > 0 %}
                                    {% for a in reallocation_available_agents %}
                                    <option value="{{ a }}" {% if reallocation_selected_agents and a in reallocation_selected_agents %}selected{% endif %}>{{ a }}</option>
                                    {% endfor %}
                                {% else %}
                                    <option value="" disabled>(Upload the Consolidate file to load agent options)</option>
                                {% endif %}
                            </select>
                            <small>Agents apply only to Workable rows (unioned with the remark filter).</small>
                        </div>

                        <button type="submit" id="reallocation-btn" disabled>📤 Generate Reallocation Data</button>
                    </form>
                    <script>
                    (function(){
                        const btn = document.getElementById('reallocation-btn');
                        const consInput = document.getElementById('reallocation_consolidate_file');
                        const blankInput = document.getElementById('reallocation_blank_file');
                        const remarksSel = document.getElementById('reallocation_remarks');
                        const agentsSel = document.getElementById('reallocation_agents');
                        let hasConsAlready = {{ 'true' if reallocation_consolidate_filename else 'false' }};
                        let hasBlankAlready = {{ 'true' if reallocation_blank_filename else 'false' }};
                        function update(){
                            const hasCons = hasConsAlready || (consInput && consInput.files && consInput.files.length > 0);
                            const hasBlank = hasBlankAlready || (blankInput && blankInput.files && blankInput.files.length > 0);
                            let selectedCount = 0;
                            if (remarksSel) {
                                if (typeof remarksSel.selectedOptions !== 'undefined') {
                                    selectedCount = remarksSel.selectedOptions.length;
                                } else {
                                    selectedCount = remarksSel.querySelectorAll('option:checked').length;
                                }
                            }
                            let agentSelected = 0;
                            if (agentsSel) {
                                if (typeof agentsSel.selectedOptions !== 'undefined') {
                                    agentSelected = agentsSel.selectedOptions.length;
                                } else {
                                    agentSelected = agentsSel.querySelectorAll('option:checked').length;
                                }
                            }
                            const hasSelection = (selectedCount > 0) || (agentSelected > 0);
                            if (btn) btn.disabled = !(hasCons && hasBlank && hasSelection);
                        }
                        async function handleConsolidateChange(){
                            update();
                            if (!consInput || !consInput.files || consInput.files.length === 0) return;
                            const fd = new FormData();
                            fd.append('consolidate_file', consInput.files[0]);
                            // Disable remarks select while loading
                            if (remarksSel) {
                                remarksSel.disabled = true;
                                remarksSel.innerHTML = '<option disabled>(Loading remarks...)</option>';
                            }
                            if (agentsSel) {
                                agentsSel.disabled = true;
                                agentsSel.innerHTML = '<option disabled>(Loading agents...)</option>';
                            }
                            try {
                                const resp = await fetch('/load_reallocation_consolidate', { method: 'POST', body: fd });
                                const data = await resp.json();
                                if (!data.ok) throw new Error(data.error || 'Failed to load remarks');
                                // Populate remarks
                                if (remarksSel) {
                                    remarksSel.innerHTML = '';
                                    (data.remarks || []).forEach(r => {
                                        const opt = document.createElement('option');
                                        opt.value = r; opt.textContent = r;
                                        remarksSel.appendChild(opt);
                                    });
                                    remarksSel.disabled = (data.remarks || []).length === 0;
                                }
                                // Populate agents
                                if (agentsSel) {
                                    agentsSel.innerHTML = '';
                                    (data.agents || []).forEach(a => {
                                        const opt = document.createElement('option');
        opt.value = a; opt.textContent = a;
                                        agentsSel.appendChild(opt);
                                    });
                                    agentsSel.disabled = (data.agents || []).length === 0;
                                }
                                hasConsAlready = true;
                                update();
                            } catch (e) {
                                if (remarksSel) {
                                    remarksSel.innerHTML = '<option disabled>(Failed to load remarks)</option>';
                                    remarksSel.disabled = true;
                                }
                                if (agentsSel) {
                                    agentsSel.innerHTML = '<option disabled>(Failed to load agents)</option>';
                                    agentsSel.disabled = true;
                                }
                                alert('Error loading consolidate file: ' + e.message);
                            }
                        }
                        if (consInput) consInput.addEventListener('change', handleConsolidateChange);
                        if (blankInput) blankInput.addEventListener('change', update);
                        if (remarksSel) remarksSel.addEventListener('change', update);
                        if (agentsSel) agentsSel.addEventListener('change', update);
                        update();
                    })();
                    </script>
                    <div class="loading" id="reallocation-loading">
                        <div class="spinner"></div>
                        <p>Generating reallocation data...</p>
                    </div>
                </div>

                <!-- Status Messages -->
                {% if reallocation_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ reallocation_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- File Status -->
                {% if reallocation_consolidate_filename %}
                <div class="section">
                    <h3>📊 File Status</h3>
                    <div class="file-status">
                        <div class="status-success">
                            ✅ Consolidate File: {{ reallocation_consolidate_filename }}<br>
                            {% if reallocation_blank_filename %}
                            ✅ Blank Allocation File: {{ reallocation_blank_filename }}
                            {% endif %}
                        </div>
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if reallocation_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ reallocation_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if reallocation_merged_data and reallocation_result and 'generation complete' in reallocation_result.lower() %}
                <div class="section">
                    <h3>💾 Download Reallocation File</h3>
                    <form action="/download_reallocation" method="post">
                        <div class="form-group">
                            <label for="reallocation_output_filename">Output filename (optional):</label>
                            <input type="text" id="reallocation_output_filename" name="filename" 
                                   value="reallocation_output.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Reallocation File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset Reallocation Tool</h3>
                    <p>Clear all uploaded files and reset the reallocation tool to start fresh.</p>
                    <form action="/reset_reallocation" method="post" onsubmit="return confirm('Are you sure you want to reset the reallocation tool? This will clear all uploaded files and data.')">
                        <button type="submit" class="reset-btn">🗑️ Reset Reallocation Tool</button>
                    </form>
                </div>
            </div>

            <!-- Tab 9: General Comparison -->
            <div id="general-tab" class="tab-content {% if active_tab == 'general' %}active{% endif %}">
                <div class="section">
                    <h3>🧭 General Comparison</h3>
                    <p>Compare a primary file against a main dataset using selected key columns, and update chosen columns in the primary file when matches are found.</p>
                </div>

                <form action="/run_general_comparison" method="post" id="general-comparison-form">
                    <div class="section">
                        <h3>📄 Primary File</h3>
                        <div class="form-group">
                            <label for="gc_primary_file">Upload Primary File:</label>
                            <input type="file" id="gc_primary_file" name="primary_file" accept=".xlsx,.xls">
                        </div>
                        <div class="form-group">
                            <label for="gc_primary_sheet">Select Primary Sheet:</label>
                            <select id="gc_primary_sheet" name="primary_sheet" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;"></select>
                        </div>
                        <div class="form-group">
                            <label for="gc_key_columns">Key Columns (composite allowed):</label>
                            <select id="gc_key_columns" name="gc_key_columns" multiple size="8" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;"></select>
                            <small>Keys are matched after trimming, collapsing whitespace, and lowercasing.</small>
                        </div>
                    </div>

                    <div class="section">
                        <h3>📚 Main Dataset File</h3>
                        <div class="form-group">
                            <label for="gc_main_file">Upload Main Dataset File:</label>
                            <input type="file" id="gc_main_file" name="main_file" accept=".xlsx,.xls">
                        </div>
                        <div class="form-group">
                            <label for="gc_main_sheet">Select Main Dataset Sheet:</label>
                            <select id="gc_main_sheet" name="main_sheet" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;"></select>
                        </div>
                        <div class="form-group">
                            <label for="gc_update_columns">Columns to Update in Primary (from main dataset):</label>
                            <select id="gc_update_columns" name="gc_update_columns" multiple size="8" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;"></select>
                            <small>If a column does not exist in the primary file, it will be added.</small>
                        </div>
                    </div>

                    <div class="section">
                        <button type="submit" id="gc-run-btn" disabled>🚀 Run General Comparison</button>
                    </div>
                </form>

                <script>
                (function(){
                    const runBtn = document.getElementById('gc-run-btn');
                    const primaryInput = document.getElementById('gc_primary_file');
                    const mainInput = document.getElementById('gc_main_file');
                    const primarySheetSel = document.getElementById('gc_primary_sheet');
                    const mainSheetSel = document.getElementById('gc_main_sheet');
                    const keyColsSel = document.getElementById('gc_key_columns');
                    const updateColsSel = document.getElementById('gc_update_columns');

                    let primaryColsBySheet = {};
                    let mainColsBySheet = {};
                    let hasPrimary = false;
                    let hasMain = false;

                    function populateSelect(selectEl, options) {
                        if (!selectEl) return;
                        selectEl.innerHTML = '';
                        (options || []).forEach((optVal) => {
                            const opt = document.createElement('option');
                            opt.value = optVal;
                            opt.textContent = optVal;
                            selectEl.appendChild(opt);
                        });
                    }

                    function updateKeyColumns() {
                        const sheet = primarySheetSel ? primarySheetSel.value : '';
                        const cols = primaryColsBySheet[sheet] || [];
                        populateSelect(keyColsSel, cols);
                    }

                    function updateUpdateColumns() {
                        const sheet = mainSheetSel ? mainSheetSel.value : '';
                        const cols = mainColsBySheet[sheet] || [];
                        populateSelect(updateColsSel, cols);
                    }

                    function updateButton() {
                        const hasPrimarySheet = hasPrimary && primarySheetSel && primarySheetSel.value;
                        const hasMainSheet = hasMain && mainSheetSel && mainSheetSel.value;
                        let keySelected = 0;
                        let updateSelected = 0;
                        if (keyColsSel) {
                            if (typeof keyColsSel.selectedOptions !== 'undefined') {
                                keySelected = keyColsSel.selectedOptions.length;
                            } else {
                                keySelected = keyColsSel.querySelectorAll('option:checked').length;
                            }
                        }
                        if (updateColsSel) {
                            if (typeof updateColsSel.selectedOptions !== 'undefined') {
                                updateSelected = updateColsSel.selectedOptions.length;
                            } else {
                                updateSelected = updateColsSel.querySelectorAll('option:checked').length;
                            }
                        }
                        if (runBtn) runBtn.disabled = !(hasPrimarySheet && hasMainSheet && keySelected > 0 && updateSelected > 0);
                    }

                    async function loadPrimary() {
                        if (!primaryInput || !primaryInput.files || primaryInput.files.length === 0) return;
                        const fd = new FormData();
                        fd.append('primary_file', primaryInput.files[0]);
                        try {
                            const resp = await fetch('/load_general_primary', { method: 'POST', body: fd });
                            const data = await resp.json();
                            if (!data.ok) throw new Error(data.error || 'Failed to load primary file');
                            primaryColsBySheet = data.columns || {};
                            populateSelect(primarySheetSel, data.sheets || []);
                            if (primarySheetSel && primarySheetSel.options.length > 0) {
                                primarySheetSel.selectedIndex = 0;
                            }
                            updateKeyColumns();
                            hasPrimary = true;
                            updateButton();
                        } catch (e) {
                            alert('Error loading primary file: ' + e.message);
                            hasPrimary = false;
                            primaryColsBySheet = {};
                            populateSelect(primarySheetSel, []);
                            populateSelect(keyColsSel, []);
                            updateButton();
                        }
                    }

                    async function loadMain() {
                        if (!mainInput || !mainInput.files || mainInput.files.length === 0) return;
                        const fd = new FormData();
                        fd.append('main_file', mainInput.files[0]);
                        try {
                            const resp = await fetch('/load_general_main', { method: 'POST', body: fd });
                            const data = await resp.json();
                            if (!data.ok) throw new Error(data.error || 'Failed to load main dataset');
                            mainColsBySheet = data.columns || {};
                            populateSelect(mainSheetSel, data.sheets || []);
                            if (mainSheetSel && mainSheetSel.options.length > 0) {
                                mainSheetSel.selectedIndex = 0;
                            }
                            updateUpdateColumns();
                            hasMain = true;
                            updateButton();
                        } catch (e) {
                            alert('Error loading main dataset: ' + e.message);
                            hasMain = false;
                            mainColsBySheet = {};
                            populateSelect(mainSheetSel, []);
                            populateSelect(updateColsSel, []);
                            updateButton();
                        }
                    }

                    if (primaryInput) primaryInput.addEventListener('change', loadPrimary);
                    if (mainInput) mainInput.addEventListener('change', loadMain);
                    if (primarySheetSel) primarySheetSel.addEventListener('change', function(){ updateKeyColumns(); updateButton(); });
                    if (mainSheetSel) mainSheetSel.addEventListener('change', function(){ updateUpdateColumns(); updateButton(); });
                    if (keyColsSel) keyColsSel.addEventListener('change', updateButton);
                    if (updateColsSel) updateColsSel.addEventListener('change', updateButton);

                    updateButton();
                })();
                </script>

                <!-- Status Messages -->
                {% if general_comparison_result %}
                <div class="section">
                    <h3>📢 Processing Status</h3>
                    <div class="status-message">
                        {{ general_comparison_result | safe }}
                    </div>
                </div>
                {% endif %}

                <!-- Processing Output -->
                {% if general_comparison_output %}
                <div class="section">
                    <h3>📝 Processing Output</h3>
                    <div class="output" style="background: #1e1e1e; color: #f8f8f2; padding: 20px; border-radius: 8px; white-space: pre-wrap; font-family: 'Courier New', monospace; max-height: 500px; overflow-y: auto; border: 1px solid #333; font-size: 14px;">
                        {{ general_comparison_output }}
                    </div>
                </div>
                {% endif %}

                <!-- Download Section -->
                {% if general_comparison_updated_data and general_comparison_result %}
                <div class="section">
                    <h3>💾 Download Updated Primary File</h3>
                    <form action="/download_general_comparison" method="post">
                        <div class="form-group">
                            <label for="gc_output_filename">Output filename (optional):</label>
                            <input type="text" id="gc_output_filename" name="filename"
                                   value="general_comparison_output.xlsx" style="width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">
                        </div>
                        <button type="submit">💾 Download Updated File</button>
                    </form>
                </div>
                {% endif %}

                <!-- Reset Section -->
                <div class="section">
                    <h3>🔄 Reset General Comparison</h3>
                    <p>Clear uploaded files, selections, and results to start fresh.</p>
                    <form action="/reset_general" method="post" onsubmit="return confirm('Reset General Comparison and clear all uploaded data?');">
                        <button type="submit" class="reset-btn">🗑️ Reset General Comparison</button>
                    </form>
                </div>
            </div>

            </div> <!-- End content -->

            <!-- Footer -->
            <div class="footer">
                <p>© 2025 Excel Automation Tools | Built with ❤️ for efficiency</p>
            </div>
        </div> <!-- End main-content -->
    </div> <!-- End app-container -->

    <script>
        // Toast notification functions
        function showToast(message, isError = false) {
            const toast = document.getElementById('toast');
            const toastMessage = document.getElementById('toast-message');
            toastMessage.textContent = message;
            toast.className = 'toast' + (isError ? ' error' : '') + ' show';
            setTimeout(hideToast, 5000);
        }
        
        function hideToast() {
            const toast = document.getElementById('toast');
            toast.classList.remove('show');
        }
        
        // Processing modal functions
        function showProcessingModal(title = 'Processing', message = 'Please wait while we process your request') {
            const overlay = document.getElementById('processing-modal-overlay');
            const titleEl = document.getElementById('processing-title');
            const messageEl = document.getElementById('processing-message');
            
            titleEl.textContent = title;
            messageEl.innerHTML = message + '<span class="processing-dots">...</span>';
            overlay.classList.add('show');
            
            // Prevent body scroll when modal is open
            document.body.style.overflow = 'hidden';
        }
        
        function hideProcessingModal() {
            const overlay = document.getElementById('processing-modal-overlay');
            overlay.classList.remove('show');
            
            // Restore body scroll
            document.body.style.overflow = '';
        }
        
        // Toggle sidebar on mobile
        function toggleSidebar() {
            const sidebar = document.getElementById('sidebar');
            sidebar.classList.toggle('active');
        }
        
        // Tab switching function
        function switchTab(tabName, skipUrlUpdate = false) {
            if (!tabName) return;
            
            // Hide all tab contents
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.remove('active');
            });
            
            // Remove active class from all menu items
            document.querySelectorAll('.menu-item').forEach(item => {
                item.classList.remove('active');
            });
            
            // Show selected tab content
            const tabContent = document.getElementById(tabName + '-tab');
            if (tabContent) {
                tabContent.classList.add('active');
            }
            
            // Add active class to menu item
            const menuItems = document.querySelectorAll('.menu-item');
            menuItems.forEach(item => {
                const itemText = item.textContent.toLowerCase();
                if ((tabName === 'comparison' && itemText.includes('comparison')) ||
                    (tabName === 'conversion' && itemText.includes('conversion')) ||
                    (tabName === 'insurance' && itemText.includes('insurance')) ||
                    (tabName === 'remarks' && itemText.includes('remarks')) ||
                    (tabName === 'appointment' && itemText.includes('appointment')) ||
                    (tabName === 'smartassist' && itemText.includes('smart assist')) ||
                    (tabName === 'consolidate' && itemText.includes('consolidate')) ||
                    (tabName === 'reallocation' && itemText.includes('reallocation')) ||
                    (tabName === 'general' && itemText.includes('general comparison'))) {
                    item.classList.add('active');
                }
            });
            
            // Update page title and description
            const pageTitles = {
                'comparison': '🔄 Comparison Tool',
                'conversion': '📋 Conversion Report Formatting',
                'insurance': '🦷 Insurance Name Formatting',
                'remarks': '📝 Update Remarks',
                'appointment': '📅 Appointment Report Formatting',
                'smartassist': '🤖 Smart Assist Report Formatting',
                'consolidate': '📊 Consolidate Report',
                'reallocation': '♻️ Generate Reallocation Data',
                'general': '🧭 General Comparison'
            };
            
            const pageDescriptions = {
                'comparison': 'Compare Patient IDs and add insurance columns',
                'conversion': 'Format and process conversion reports',
                'insurance': 'Standardize insurance carrier names',
                'remarks': 'Add remarks to appointments',
                'appointment': 'Format appointment report insurance columns',
                'smartassist': 'Format smart assist report insurance columns',
                'consolidate': 'Consolidate master and daily report files',
                'reallocation': 'Generate reallocation data from consolidate file',
                'general': 'Compare two files and update primary rows on match'
            };
            
            document.getElementById('page-title').textContent = pageTitles[tabName] || pageTitles['comparison'];
            document.getElementById('page-description').textContent = pageDescriptions[tabName] || pageDescriptions['comparison'];
            
            // Close sidebar on mobile
            if (window.innerWidth <= 968) {
                document.getElementById('sidebar').classList.remove('active');
            }
            
            // Update URL
            if (!skipUrlUpdate) {
                const newUrl = window.location.pathname + '?tab=' + tabName;
                window.history.pushState({path: newUrl}, '', newUrl);
            }
        }
        // Form submission with processing modal
        const rawForm = document.getElementById('raw-form');
        if (rawForm) {
            rawForm.addEventListener('submit', function() {
                showProcessingModal('Uploading Appointment Report', 'Processing your Appointment Report file');
                const btn = document.getElementById('raw-btn');
                if (btn) btn.disabled = true;
            });
        }

        const previousForm = document.getElementById('previous-form');
        if (previousForm) {
            previousForm.addEventListener('submit', function() {
                showProcessingModal('Uploading Smart Assist File', 'Processing your Smart Assist file');
                const btn = document.getElementById('previous-btn');
                if (btn) btn.disabled = true;
            });
        }

        const compareForm = document.getElementById('compare-form');
        if (compareForm) {
            compareForm.addEventListener('submit', function() {
                showProcessingModal('Comparing Files', 'Matching Patient IDs and updating insurance columns');
                const btn = document.getElementById('compare-btn');
                if (btn) btn.disabled = true;
            });
        }

        // Conversion form submission
        const conversionForm = document.getElementById('conversion-form');
        if (conversionForm) {
            conversionForm.addEventListener('submit', function() {
                showProcessingModal('Processing Conversion Report', 'Validating and formatting conversion report');
                document.getElementById('conversion-btn').disabled = true;
            });
        }

        // Insurance formatting form submission
        const insuranceForm = document.getElementById('insurance-form');
        if (insuranceForm) {
            insuranceForm.addEventListener('submit', function() {
                showProcessingModal('Formatting Insurance Names', 'Processing file and formatting insurance columns');
                document.getElementById('insurance-btn').disabled = true;
            });
        }

        // Remarks form submission
        const remarksForm = document.getElementById('remarks-form');
        if (remarksForm) {
            remarksForm.addEventListener('submit', function() {
                showProcessingModal('Processing Files', 'Matching Patient IDs and updating appointments with remarks');
                document.getElementById('remarks-btn').disabled = true;
            });
        }

        // Appointment report form submission
        const appointmentForm = document.getElementById('appointment-form');
        if (appointmentForm) {
            appointmentForm.addEventListener('submit', function() {
                showProcessingModal('Formatting Insurance Columns', 'Processing file and formatting insurance names');
                document.getElementById('appointment-btn').disabled = true;
            });
        }

        // Smart assist form submission
        const smartAssistForm = document.getElementById('smartassist-form');
        if (smartAssistForm) {
            smartAssistForm.addEventListener('submit', function() {
                showProcessingModal('Formatting Insurance Columns', 'Processing file and formatting insurance names');
                document.getElementById('smartassist-btn').disabled = true;
            });
        }

        // Consolidate report form submission
        const consolidateForm = document.getElementById('consolidate-form');
        if (consolidateForm) {
            consolidateForm.addEventListener('submit', function() {
                showProcessingModal('Consolidating Files', 'Merging master and daily consolidated files');
                document.getElementById('consolidate-btn').disabled = true;
            });
        }

        // Reallocation form submission
        const reallocationForm = document.getElementById('reallocation-form');
        if (reallocationForm) {
            reallocationForm.addEventListener('submit', function() {
                showProcessingModal('Generating Reallocation Data', 'Processing consolidate and blank allocation files');
                document.getElementById('reallocation-btn').disabled = true;
            });
        }

        // General comparison form submission
        const generalComparisonForm = document.getElementById('general-comparison-form');
        if (generalComparisonForm) {
            generalComparisonForm.addEventListener('submit', function() {
                showProcessingModal('Running General Comparison', 'Matching keys and updating primary rows');
                const btn = document.getElementById('gc-run-btn');
                if (btn) btn.disabled = true;
            });
        }

        // Reset forms - show modal when form is submitted (HTML confirm already handled)
        const resetForms = document.querySelectorAll('form[action*="reset"]');
        resetForms.forEach(form => {
            form.addEventListener('submit', function() {
                // Show modal - HTML onsubmit confirm already handled the confirmation
                showProcessingModal('Resetting', 'Clearing all data and files');
            });
        });

        // Auto-scroll results to bottom
        function scrollResults() {
            const results = document.getElementById('results');
            results.scrollTop = results.scrollHeight;
        }

        // Scroll results on page load
        window.onload = function() {
            scrollResults();
            
            // Hide processing modal on page load (in case it was left open)
            hideProcessingModal();
            
            // Check URL parameter for active tab
            const urlParams = new URLSearchParams(window.location.search);
            const activeTab = urlParams.get('tab');
            if (activeTab === 'conversion') {
                switchTab('conversion', true);
            } else if (activeTab === 'insurance') {
                switchTab('insurance', true);
            } else if (activeTab === 'remarks') {
                switchTab('remarks', true);
            } else if (activeTab === 'appointment') {
                switchTab('appointment', true);
            } else if (activeTab === 'smartassist') {
                switchTab('smartassist', true);
            } else if (activeTab === 'consolidate') {
                switchTab('consolidate', true);
            } else if (activeTab === 'reallocation') {
                switchTab('reallocation', true);
            } else if (activeTab === 'comparison') {
                switchTab('comparison', true);
            }
            
            // Show toast notification for conversion validation and processing
            {% if conversion_result %}
            var result = `{{ conversion_result | safe }}`;
            if (result.includes('processing completed successfully')) {
                showToast('✅ Processing completed! Formatted Insurance column added to all sheets.', false);
            } else if (result.includes('Validation successful')) {
                showToast('✅ Validation successful! Insurance Note column found in all sheets.', false);
            } else if (result.includes('Validation Error') || result.includes('❌ Error')) {
                var errorMsg = result.split('\\n')[0];
                showToast(errorMsg, true);
            }
            {% endif %}
        }
        
        // Helper function to switch tab programmatically
        function switchTabProgrammatically(tabName) {
            // Hide all tab contents
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.remove('active');
            });
            
            // Remove active class from all tabs
            document.querySelectorAll('.tab').forEach(tab => {
                tab.classList.remove('active');
            });
            
            // Show selected tab content
            document.getElementById(tabName + '-tab').classList.add('active');
            
            // Add active class to corresponding tab button
            const tabs = document.querySelectorAll('.tab');
            tabs.forEach(tab => {
                if (tab.textContent.includes(tabName === 'comparison' ? 'Comparison' : 'Conversion')) {
                    tab.classList.add('active');
                }
            });
        }
    </script>
</body>
</html>
"""

# State abbreviations mapping
STATE_ABBREVIATIONS = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AR": "Arkansas",
    "AZ": "Arizona",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "DC": "District of Columbia",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
}


def expand_state_abbreviations(text):
    """Expand state abbreviations to full state names - only if the entire text is an abbreviation"""
    if pd.isna(text):
        return text

    text_str = str(text).strip()

    # Check if the entire text (case-insensitive) is a state abbreviation
    # Only expand if it's exactly an abbreviation, not if abbreviation appears within a longer word
    text_upper = text_str.upper()
    if text_upper in STATE_ABBREVIATIONS:
        return STATE_ABBREVIATIONS[text_upper]

    # If not a direct match, return the original text unchanged
    # This preserves full state names like "Arizona", "Minnesota", etc.
    return text_str


def format_insurance_name(insurance_text):
    """Format insurance name to match expected format"""
    if pd.isna(insurance_text):
        return insurance_text

    insurance_str = str(insurance_text).strip()

    # Handle special cases first
    if insurance_str.upper() == "NO INSURANCE":
        return "No Insurance"
    elif insurance_str.upper() == "PATIENT NOT FOUND":
        return "PATIENT NOT FOUND"
    elif insurance_str.upper() == "DUPLICATE":
        return "DUPLICATE"
    elif re.search(r"no\s+patient\s+chart", insurance_str, re.IGNORECASE):
        return "No Patient chart"

    # Extract company name before "Ph#"
    if "Ph#" in insurance_str:
        company_name = insurance_str.split("Ph#")[0].strip()
    else:
        company_name = insurance_str

    # Remove "Primary" and "Secondary" text
    company_name = re.sub(r"\s*\(Primary\)", "", company_name, flags=re.IGNORECASE)
    company_name = re.sub(r"\s*\(Secondary\)", "", company_name, flags=re.IGNORECASE)
    company_name = re.sub(r"\s*Primary", "", company_name, flags=re.IGNORECASE)
    company_name = re.sub(r"\s*Secondary", "", company_name, flags=re.IGNORECASE)

    # If already formatted as "DD [State]", preserve it (don't reformat)
    if re.match(r"^dd\s+", company_name, re.IGNORECASE):
        # Extract the state part
        state_part = re.sub(r"^dd\s+", "", company_name, flags=re.IGNORECASE).strip()
        # Remove any trailing text (like "Ph#")
        state_part = re.split(r"\s*[,\|;]\s*|\s+Ph#", state_part, flags=re.IGNORECASE)[
            0
        ].strip()
        # Capitalize properly (Title Case) if not already uppercase
        if state_part and not state_part.isupper():
            state_part = state_part.title()
        return f"DD {state_part}"

    # Handle Delta Dental variations
    if re.search(r"delta\s+dental", company_name, re.IGNORECASE):
        # Extract state from Delta Dental - handles "Delta Dental Arizona", "Delta Dental of Arizona", etc.
        delta_match = re.search(
            r"delta\s+dental\s+(?:of\s+)?(.+)", company_name, re.IGNORECASE
        )
        if delta_match:
            state = delta_match.group(1).strip()
            # Remove any trailing text after state (like "Ph#" or other info)
            # Split on common separators and take first part
            state = re.split(r"\s*[,\|;]\s*|\s+Ph#", state, flags=re.IGNORECASE)[
                0
            ].strip()
            # Expand state abbreviations (e.g., "AZ" -> "Arizona")
            state = expand_state_abbreviations(state)
            # Capitalize properly (Title Case)
            if state and not state.isupper():
                state = state.title()
            return f"DD {state}"
        else:
            return "DD"

    # Handle Anthem variations FIRST (before BCBS to avoid conflicts)
    if re.search(
        r"anthem|blue\s+cross.*anthem|anthem.*blue\s+cross", company_name, re.IGNORECASE
    ):
        return "Anthem"

    # Handle BCBS variations
    elif re.search(
        r"bcbs|bc/bs|bc\s+of|blue\s+cross|blue\s+shield|bcbbs",
        company_name,
        re.IGNORECASE,
    ):
        # Check for full "Blue Cross Blue Shield" pattern first
        if re.search(r"blue\s+cross\s+blue\s+shield", company_name, re.IGNORECASE):
            # Extract state from "Blue Cross Blue Shield of [State]"
            bcbs_match = re.search(
                r"blue\s+cross\s+blue\s+shield\s+(?:of\s+)?(.+)",
                company_name,
                re.IGNORECASE,
            )
            if bcbs_match:
                state = bcbs_match.group(1).strip()
                # Expand state abbreviations
                state = expand_state_abbreviations(state)
                return f"BCBS {state}"
            else:
                return "BCBS"
        # Handle BC/BS patterns
        elif re.search(r"bc/bs", company_name, re.IGNORECASE):
            bcbs_match = re.search(
                r"bc/bs\s+(?:of\s+)?(.+)", company_name, re.IGNORECASE
            )
            if bcbs_match:
                state = bcbs_match.group(1).strip()
                # Expand state abbreviations
                state = expand_state_abbreviations(state)
                return f"BCBS {state}"
            else:
                return "BCBS"
        # Handle BC Of patterns
        elif re.search(r"bc\s+of", company_name, re.IGNORECASE):
            bcbs_match = re.search(r"bc\s+of\s+(.+)", company_name, re.IGNORECASE)
            if bcbs_match:
                state = bcbs_match.group(1).strip()
                # Expand state abbreviations
                state = expand_state_abbreviations(state)
                return f"BCBS {state}"
            else:
                return "BCBS"
        # Handle BCBBS typo
        elif re.search(r"bcbbs", company_name, re.IGNORECASE):
            return "BCBS"
        # Handle other BCBS patterns
        else:
            bcbs_match = re.search(
                r"(?:bcbs|blue\s+cross|blue\s+shield)\s+(?:of\s+)?(.+)",
                company_name,
                re.IGNORECASE,
            )
            if bcbs_match:
                state = bcbs_match.group(1).strip()
                # Expand state abbreviations
                state = expand_state_abbreviations(state)
                return f"BCBS {state}"
            else:
                return "BCBS"

    # Handle other specific companies
    elif re.search(r"metlife|met\s+life", company_name, re.IGNORECASE):
        return "Metlife"
    elif re.search(r"cigna", company_name, re.IGNORECASE):
        return "Cigna"
    elif re.search(r"aarp", company_name, re.IGNORECASE):
        return "AARP"
    elif re.search(r"adn\s+administrators", company_name, re.IGNORECASE):
        return "ADN Administrators"
    elif re.search(r"beam", company_name, re.IGNORECASE):
        return "Beam"
    elif re.search(
        r"uhc|united.*health|united.*heal|unitedhelathcare", company_name, re.IGNORECASE
    ):
        return "UHC"
    elif re.search(r"teamcare", company_name, re.IGNORECASE):
        return "Teamcare"
    elif re.search(r"humana", company_name, re.IGNORECASE):
        return "Humana"
    elif re.search(r"aetna", company_name, re.IGNORECASE):
        return "Aetna"
    elif re.search(r"guardian", company_name, re.IGNORECASE):
        return "Guardian"
    elif re.search(r"g\s*e\s*h\s*a", company_name, re.IGNORECASE):
        return "GEHA"
    elif re.search(r"principal", company_name, re.IGNORECASE):
        return "Principal"
    elif re.search(r"ameritas", company_name, re.IGNORECASE):
        return "Ameritas"
    elif re.search(r"physicians\s+mutual", company_name, re.IGNORECASE):
        return "Physicians Mutual"
    elif re.search(r"mutual\s+of\s+omaha", company_name, re.IGNORECASE):
        return "Mutual Omaha"
    elif re.search(r"sunlife|sun\s+life", company_name, re.IGNORECASE):
        return "Sunlife"
    elif re.search(r"liberty(?:\s+dental)?", company_name, re.IGNORECASE):
        return "Liberty Dental Plan"
    elif re.search(r"careington", company_name, re.IGNORECASE):
        return "Careington Benefit Solutions"
    elif re.search(r"automated\s+benefit", company_name, re.IGNORECASE):
        return "Automated Benefit Services Inc"
    elif re.search(r"network\s+health", company_name, re.IGNORECASE):
        # Check if it has "Wisconsin" in the name
        if re.search(r"wisconsin", company_name, re.IGNORECASE):
            return "Network Health Wisconsin"
        else:
            return "Network Health Go"
    elif re.search(r"regence", company_name, re.IGNORECASE):
        return "REGENCE BCBS"
    elif re.search(r"united\s+concordia", company_name, re.IGNORECASE):
        return "United Concordia"
    elif re.search(r"medical\s+mutual", company_name, re.IGNORECASE):
        return "Medical Mutual"
    elif re.search(r"blue\s+care\s+dental", company_name, re.IGNORECASE):
        return "Blue Care Dental"
    elif re.search(r"dominion\s+dental", company_name, re.IGNORECASE):
        return "Dominion Dental"
    elif re.search(r"carefirst", company_name, re.IGNORECASE):
        return "CareFirst BCBS"
    elif re.search(r"health\s*partners", company_name, re.IGNORECASE):
        # Check if it has "of [State]" pattern
        if re.search(r"health\s*partners\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"health\s*partners\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            return f"Health Partners {state}"
        else:
            return "Health Partners"
    elif re.search(r"keenan", company_name, re.IGNORECASE):
        return "Keenan"
    elif re.search(r"wilson\s+mcshane", company_name, re.IGNORECASE):
        return "Wilson McShane- Delta Dental"
    elif re.search(r"standard\s+(?:life\s+)?insurance", company_name, re.IGNORECASE):
        return "Standard Life Insurance"
    elif re.search(r"plan\s+for\s+health", company_name, re.IGNORECASE):
        return "Plan for Health"
    elif re.search(r"kansas\s+city", company_name, re.IGNORECASE):
        return "Kansas City"
    elif re.search(r"the\s+guardian", company_name, re.IGNORECASE):
        return "The Guardian"
    elif re.search(r"community\s+dental", company_name, re.IGNORECASE):
        return "Community Dental Associates"
    elif re.search(r"northeast\s+delta\s+dental", company_name, re.IGNORECASE):
        return "Northeast Delta Dental"
    elif re.search(r"say\s+cheese\s+dental", company_name, re.IGNORECASE):
        return "Say Cheese Dental Network"
    elif re.search(r"dentaquest", company_name, re.IGNORECASE):
        return "Dentaquest"
    elif re.search(r"umr", company_name, re.IGNORECASE):
        return "UMR"
    elif re.search(r"mhbp", company_name, re.IGNORECASE):
        return "MHBP"
    elif re.search(r"united\s+states\s+army", company_name, re.IGNORECASE):
        return "United States Army"
    elif re.search(r"conversion\s+default", company_name, re.IGNORECASE):
        return "CONVERSION DEFAULT - Do NOT Delete! Change Pt Ins!"
    elif re.search(r"equitable", company_name, re.IGNORECASE):
        return "Equitable"
    elif re.search(r"manhattan\s+life", company_name, re.IGNORECASE):
        return "Manhattan Life"
    elif re.search(r"ucci", company_name, re.IGNORECASE):
        return "UCCI"
    elif re.search(r"ccpoa|cc\s*poa|c\s+c\s+p\s+o\s+a", company_name, re.IGNORECASE):
        return "CCPOA"
    elif re.search(
        r"dd\s+of|dd\s+[a-z]{2}|delta\s+dental|dental\s+dental|denta\s+dental|dleta\s+dental|dektal?\s+dental",
        company_name,
        re.IGNORECASE,
    ):
        # Extract state from various Delta Dental patterns
        # Handle DD OF [State] pattern
        if re.search(r"dd\s+of\s+([a-z]{2})", company_name, re.IGNORECASE):
            state_match = re.search(
                r"dd\s+of\s+([a-z]{2})", company_name, re.IGNORECASE
            )
            state = state_match.group(1).upper()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle DD [State] pattern - only match if it's exactly 2 letters (abbreviation) followed by word boundary
        elif re.search(r"dd\s+([a-z]{2})\b", company_name, re.IGNORECASE):
            state_match = re.search(r"dd\s+([a-z]{2})\b", company_name, re.IGNORECASE)
            state = state_match.group(1).upper()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle DD [Full State Name] pattern - for cases like "DD Pennsylvania"
        elif re.search(r"^dd\s+([a-z]{3,})", company_name, re.IGNORECASE):
            state_match = re.search(r"^dd\s+([a-z]{3,})", company_name, re.IGNORECASE)
            state = state_match.group(1).strip()
            # Remove any trailing text after state (like "Ph#" or other info)
            state = re.split(r"\s*[,\|;]\s*|\s+Ph#", state, flags=re.IGNORECASE)[
                0
            ].strip()
            # Don't expand abbreviations - preserve full state name
            # Capitalize properly (Title Case)
            if state and not state.isupper():
                state = state.title()
            return f"DD {state}"
        # Handle Delta Dental of [State] pattern
        elif re.search(r"delta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"delta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Dental Dental Of [State] pattern
        elif re.search(r"dental\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"dental\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Denta Dental Of [State] pattern
        elif re.search(r"denta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"denta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Dleta Dental of [State] pattern
        elif re.search(r"dleta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"dleta\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Dekta Dental of [State] pattern
        elif re.search(r"dektal?\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(
                r"dektal?\s+dental\s+of\s+(.+)", company_name, re.IGNORECASE
            )
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Dental of [State] pattern
        elif re.search(r"dental\s+of\s+(.+)", company_name, re.IGNORECASE):
            state_match = re.search(r"dental\s+of\s+(.+)", company_name, re.IGNORECASE)
            state = state_match.group(1).strip()
            state = expand_state_abbreviations(state)
            return f"DD {state}"
        # Handle Dental Network of America
        elif re.search(r"dental\s+network\s+of\s+america", company_name, re.IGNORECASE):
            return "DD Network of America"
        # Default DD
        else:
            return "DD"

    # If no specific pattern matches, return the cleaned company name
    return company_name.strip()


def normalize_patient_id(patient_id_val):
    """Normalize patient ID for consistent matching - handles numeric, string, whitespace, leading zeros"""
    if pd.isna(patient_id_val):
        return None

    # Convert to string first
    patient_id_str = str(patient_id_val).strip()

    if not patient_id_str:
        return None

    # Try to convert to number to remove leading zeros and scientific notation
    try:
        # If it's a number, convert to int (removes leading zeros and decimals)
        if "." in patient_id_str:
            num_val = float(patient_id_str)
            # If it's a whole number, convert to int
            if num_val.is_integer():
                return str(int(num_val))
            else:
                return str(num_val)
        else:
            num_val = int(patient_id_str)
            return str(num_val)
    except (ValueError, OverflowError):
        # If it's not a number, return the cleaned string
        return patient_id_str.strip()


def compare_patient_names(raw_df, previous_df):
    """Compare Patient ID from Appointment report with PATID from Smart Assist and add insurance columns"""
    try:
        # Debug: Show available columns
        raw_columns = list(raw_df.columns)
        previous_columns = list(previous_df.columns)

        # Function to find specific column names
        def find_column(columns, target_names):
            """Find column matching any of the target names (case-insensitive)"""
            for col in columns:
                col_lower = col.lower().strip()
                for target in target_names:
                    if col_lower == target.lower().strip():
                        return col
            return None

        # Find Patient ID in raw file (Appointment report)
        raw_patient_col = find_column(
            raw_columns, ["Patient ID", "PatientID", "patient id"]
        )
        if not raw_patient_col:
            # Try flexible search as fallback
            for col in raw_columns:
                if "patient" in col.lower() and "id" in col.lower():
                    raw_patient_col = col
                    break

        # Find PATID in previous file (Smart Assist) - now also checking for "Patient ID"
        previous_patient_col = find_column(
            previous_columns,
            [
                "PATID",
                "PatID",
                "patid",
                "PAT ID",
                "Patient ID",
                "PatientID",
                "patient id",
            ],
        )
        if not previous_patient_col:
            # Try flexible search as fallback
            for col in previous_columns:
                col_lower = col.lower().strip()
                if (
                    "patid" in col_lower
                    or "pat id" in col_lower
                    or ("patient" in col_lower and "id" in col_lower)
                ):
                    previous_patient_col = col
                    break

        if not previous_patient_col:
            return (
                f"❌ Error: 'PATID' or 'Patient ID' column not found in Smart Assist file.\nAvailable columns: {previous_columns}",
                None,
            )

        # Find insurance columns in Smart Assist file, or create them if missing
        primary_ins_col = find_column(
            previous_columns,
            [
                "Dental Primary Ins Carr",
                "DentalPrimaryInsCarr",
                "Dental Primary Insurance Carrier",
            ],
        )
        secondary_ins_col = find_column(
            previous_columns,
            [
                "Dental Secondary Ins Carr",
                "DentalSecondaryInsCarr",
                "Dental Secondary Insurance Carrier",
            ],
        )

        # Create a working copy of the Smart Assist DataFrame
        smart_assist_df = previous_df.copy()

        # Track if columns were added
        columns_added = []

        # Add missing insurance columns if they don't exist
        if not primary_ins_col:
            primary_ins_col = "Dental Primary Ins Carr"
            smart_assist_df[primary_ins_col] = ""
            columns_added.append(primary_ins_col)

        if not secondary_ins_col:
            secondary_ins_col = "Dental Secondary Ins Carr"
            smart_assist_df[secondary_ins_col] = ""
            columns_added.append(secondary_ins_col)

        # Check if Patient ID column exists in raw file
        if not raw_patient_col:
            return (
                f"❌ Error: 'Patient ID' column not found in Appointment report file.\nAvailable columns: {raw_columns}",
                None,
            )

        # Find insurance columns in Appointment Report file (more flexible search)
        appointment_primary_col = None
        appointment_secondary_col = None

        # Try exact matches first
        appointment_primary_col = find_column(
            raw_columns,
            [
                "Dental Primary Ins Carr",
                "DentalPrimaryInsCarr",
                "Dental Primary Insurance Carrier",
            ],
        )
        appointment_secondary_col = find_column(
            raw_columns,
            [
                "Dental Secondary Ins Carr",
                "DentalSecondaryInsCarr",
                "Dental Secondary Insurance Carrier",
            ],
        )

        # If not found, try partial matches
        if not appointment_primary_col:
            for col in raw_columns:
                col_lower = col.lower()
                if ("primary" in col_lower and "ins" in col_lower) or (
                    "primary" in col_lower and "insurance" in col_lower
                ):
                    appointment_primary_col = col
                    break

        if not appointment_secondary_col:
            for col in raw_columns:
                col_lower = col.lower()
                if ("secondary" in col_lower and "ins" in col_lower) or (
                    "secondary" in col_lower and "insurance" in col_lower
                ):
                    appointment_secondary_col = col
                    break

        # Create a mapping from Patient ID to insurance data in Appointment Report
        appointment_insurance_map = {}
        for idx, row in raw_df.iterrows():
            patient_id_val = row[raw_patient_col]
            # Normalize patient ID for consistent matching
            patient_id = normalize_patient_id(patient_id_val)
            if patient_id:
                primary_value = ""
                secondary_value = ""

                if appointment_primary_col:
                    primary_val = row[appointment_primary_col]
                    if pd.notna(primary_val):
                        primary_value = str(primary_val).strip()

                if appointment_secondary_col:
                    secondary_val = row[appointment_secondary_col]
                    if pd.notna(secondary_val):
                        secondary_value = str(secondary_val).strip()

                appointment_insurance_map[patient_id] = {
                    "primary": primary_value,
                    "secondary": secondary_value,
                }

        # Use Smart Assist file as the result file (base)
        result_df = smart_assist_df.copy()

        # Find the position of the PATID column in Smart Assist
        patid_col_index = result_df.columns.get_loc(previous_patient_col)

        # Check if insurance columns already exist, if not create them
        primary_col_name = "Dental Primary Ins Carr"
        secondary_col_name = "Dental Secondary Ins Carr"

        if primary_col_name not in result_df.columns:
            result_df.insert(patid_col_index + 1, primary_col_name, "")
        else:
            # Column already exists, initialize empty values if needed
            result_df[primary_col_name] = result_df[primary_col_name].fillna("")

        if secondary_col_name not in result_df.columns:
            # Insert after primary column (or after PATID if primary doesn't exist)
            if primary_col_name in result_df.columns:
                # Find the position of primary column and insert after it
                primary_col_index = result_df.columns.get_loc(primary_col_name)
                result_df.insert(primary_col_index + 1, secondary_col_name, "")
            else:
                # Primary column doesn't exist, insert after PATID
                result_df.insert(patid_col_index + 1, secondary_col_name, "")
        else:
            # Column already exists, initialize empty values if needed
            result_df[secondary_col_name] = result_df[secondary_col_name].fillna("")

        # Compare and populate insurance columns in Smart Assist file from Appointment Report
        matched_count = 0
        unmatched_patids = []  # Track unmatched PATIDs for debugging
        for idx, row in result_df.iterrows():
            patid_val = row[previous_patient_col]
            # Normalize patient ID for consistent matching
            patid = normalize_patient_id(patid_val)
            if patid and patid in appointment_insurance_map:
                # Match found - Patient ID from Appointment Report matches PATID in Smart Assist
                # Copy insurance data FROM Appointment Report TO Smart Assist file
                # Format the insurance names before adding
                insurance_data = appointment_insurance_map[patid]
                result_df.at[idx, primary_col_name] = format_insurance_name(
                    insurance_data["primary"]
                )
                result_df.at[idx, secondary_col_name] = format_insurance_name(
                    insurance_data["secondary"]
                )
                matched_count += 1
            elif patid:
                # Track unmatched PATIDs (limit to first 10 for display)
                if len(unmatched_patids) < 10:
                    unmatched_patids.append(str(patid_val))

        # Count statistics
        total_patients = len(result_df)
        not_matched_patients = total_patients - matched_count

        # Generate result message
        columns_info = ""
        if columns_added:
            columns_info = f"\n⚠️ Note: The following columns were added to Smart Assist file (with empty values): {', '.join(columns_added)}"

        # Add debugging info for unmatched records
        unmatched_info = ""
        if unmatched_patids and not_matched_patients > 0:
            unmatched_info = f"\n\n🔍 Sample unmatched PATIDs from Smart Assist (first 10): {', '.join(unmatched_patids)}"
            unmatched_info += f"\n💡 Tip: Check if these Patient IDs exist in the Appointment Report file with the same values."

        result_message = f"""✅ Comparison completed successfully!

📊 Statistics:
- Total records in Smart Assist file: {total_patients}
- Records matched with Appointment report: {matched_count}
- Records not matched: {not_matched_patients}
- Match rate: {(matched_count/total_patients*100):.1f}%

📋 Columns used:
- Appointment report (Raw file): "{raw_patient_col}"
  - Primary Insurance: {appointment_primary_col if appointment_primary_col else "❌ Not found - will use empty values"}
  - Secondary Insurance: {appointment_secondary_col if appointment_secondary_col else "❌ Not found - will use empty values"}
- Smart Assist file (Previous file): "{previous_patient_col}"
  - Insurance columns: "{primary_ins_col}", "{secondary_ins_col}"{columns_info}
  
📊 Matching details:
- Total Patient IDs in Appointment Report: {len(appointment_insurance_map)}
- Records in Smart Assist: {total_patients}
- Successful matches: {matched_count}{unmatched_info}

📋 Result file:
- Based on Smart Assist file
- "Dental Primary Ins Carr" - populated from Appointment Report for matched records
- "Dental Secondary Ins Carr" - populated from Appointment Report for matched records
- Empty cells for records not found in Appointment report

💾 Ready to download the result file!"""

        return result_message, result_df

    except Exception as e:
        return f"❌ Error during comparison: {str(e)}", None


@app.route("/")
def root():
    return redirect("/comparison")


@app.route("/comparison")
def comparison_index():
    global raw_data, previous_data, raw_filename, previous_filename, comparison_result
    global conversion_data, conversion_filename, conversion_result
    global insurance_formatting_data, insurance_formatting_filename, insurance_formatting_result, insurance_formatting_output
    global remarks_appointments_data, remarks_excel_data, remarks_appointments_filename, remarks_remarks_filename, remarks_result, remarks_updated_count
    global appointment_report_data, appointment_report_filename, appointment_report_result, appointment_report_output
    global smart_assist_data, smart_assist_filename, smart_assist_result, smart_assist_output
    global consolidate_master_data, consolidate_daily_data, consolidate_master_filename, consolidate_daily_filename, consolidate_result, consolidate_output
    global reallocation_consolidate_data, reallocation_blank_data, reallocation_consolidate_filename, reallocation_blank_filename, reallocation_result, reallocation_output, reallocation_merged_data, reallocation_available_remarks, reallocation_selected_remarks, reallocation_available_agents, reallocation_selected_agents
    global general_primary_data, general_main_data, general_primary_filename, general_main_filename, general_comparison_result, general_comparison_output, general_comparison_updated_data

    # Get the active tab from URL parameter
    active_tab = request.args.get("tab", "comparison")

    return render_template_string(
        HTML_TEMPLATE,
        raw_data=raw_data,
        previous_data=previous_data,
        raw_filename=raw_filename,
        previous_filename=previous_filename,
        comparison_result=comparison_result,
        conversion_data=conversion_data,
        conversion_filename=conversion_filename,
        conversion_result=conversion_result,
        insurance_formatting_data=insurance_formatting_data,
        insurance_formatting_filename=insurance_formatting_filename,
        insurance_formatting_result=insurance_formatting_result,
        insurance_formatting_output=insurance_formatting_output,
        remarks_appointments_data=remarks_appointments_data,
        remarks_excel_data=remarks_excel_data,
        remarks_appointments_filename=remarks_appointments_filename,
        remarks_remarks_filename=remarks_remarks_filename,
        remarks_result=remarks_result,
        remarks_updated_count=remarks_updated_count,
        appointment_report_data=appointment_report_data,
        appointment_report_filename=appointment_report_filename,
        appointment_report_result=appointment_report_result,
        appointment_report_output=appointment_report_output,
        smart_assist_data=smart_assist_data,
        smart_assist_filename=smart_assist_filename,
        smart_assist_result=smart_assist_result,
        smart_assist_output=smart_assist_output,
        consolidate_master_data=consolidate_master_data,
        consolidate_daily_data=consolidate_daily_data,
        consolidate_master_filename=consolidate_master_filename,
        consolidate_daily_filename=consolidate_daily_filename,
        consolidate_result=consolidate_result,
        consolidate_output=consolidate_output,
        reallocation_consolidate_data=reallocation_consolidate_data,
        reallocation_blank_data=reallocation_blank_data,
        reallocation_consolidate_filename=reallocation_consolidate_filename,
        reallocation_blank_filename=reallocation_blank_filename,
        reallocation_result=reallocation_result,
        reallocation_output=reallocation_output,
        reallocation_merged_data=reallocation_merged_data,
        reallocation_available_remarks=reallocation_available_remarks,
        reallocation_selected_remarks=reallocation_selected_remarks,
        reallocation_available_agents=reallocation_available_agents,
        reallocation_selected_agents=reallocation_selected_agents,
        general_primary_data=general_primary_data,
        general_main_data=general_main_data,
        general_primary_filename=general_primary_filename,
        general_main_filename=general_main_filename,
        general_comparison_result=general_comparison_result,
        general_comparison_output=general_comparison_output,
        general_comparison_updated_data=general_comparison_updated_data,
        active_tab=active_tab,
    )


@app.route("/upload_raw", methods=["POST"])
def upload_raw_file():
    global raw_data, raw_filename, comparison_result

    if "file" not in request.files:
        comparison_result = "❌ Error: No file provided"
        return redirect("/comparison")

    file = request.files["file"]
    if file.filename == "":
        comparison_result = "❌ Error: No file selected"
        return redirect("/comparison")

    try:
        # Get filename without saving to disk
        filename = secure_filename(file.filename)

        # Read Excel file directly from memory (no disk storage)
        file.seek(0)  # Reset file pointer to beginning
        raw_data = pd.read_excel(file, sheet_name=None, engine="openpyxl")

        # Remove "Unnamed:" columns from all sheets
        cleaned_data = {}
        for sheet_name, df in raw_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned
        raw_data = cleaned_data

        raw_filename = filename

        comparison_result = f"✅ Appointment Report uploaded successfully! Loaded {len(raw_data)} sheets: {', '.join(list(raw_data.keys()))}"
        return redirect("/comparison?tab=comparison")

    except Exception as e:
        comparison_result = f"❌ Error uploading Appointment Report: {str(e)}"
        return redirect("/comparison?tab=comparison")


@app.route("/upload_previous", methods=["POST"])
def upload_previous_file():
    global previous_data, previous_filename, comparison_result

    if "file" not in request.files:
        comparison_result = "❌ Error: No file provided"
        return redirect("/comparison?tab=comparison")

    file = request.files["file"]
    if file.filename == "":
        comparison_result = "❌ Error: No file selected"
        return redirect("/comparison?tab=comparison")

    try:
        # Get filename without saving to disk
        filename = secure_filename(file.filename)

        # Read Excel file directly from memory (no disk storage)
        file.seek(0)  # Reset file pointer to beginning
        previous_data = pd.read_excel(file, sheet_name=None, engine="openpyxl")

        # Remove "Unnamed:" columns from all sheets
        cleaned_data = {}
        for sheet_name, df in previous_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned
        previous_data = cleaned_data

        previous_filename = filename

        comparison_result = f"✅ Smart Assist file uploaded successfully! Loaded {len(previous_data)} sheets: {', '.join(list(previous_data.keys()))}"
        return redirect("/comparison?tab=comparison")

    except Exception as e:
        comparison_result = f"❌ Error uploading Smart Assist file: {str(e)}"
        return redirect("/comparison?tab=comparison")


@app.route("/compare", methods=["POST"])
def compare_files():
    global raw_data, previous_data, comparison_result

    if not raw_data or not previous_data:
        comparison_result = "❌ Error: Please upload both files first"
        return redirect("/comparison?tab=comparison")

    raw_sheet = request.form.get("raw_sheet")
    previous_sheet = request.form.get("previous_sheet")

    if not raw_sheet or not previous_sheet:
        comparison_result = "❌ Error: Please select sheets for both files"
        return redirect("/comparison?tab=comparison")

    try:
        # Get the selected sheets
        raw_df = raw_data[raw_sheet]
        previous_df = previous_data[previous_sheet]

        # Perform comparison
        result_message, result_df = compare_patient_names(raw_df, previous_df)

        if result_df is not None:
            # Store the result for download
            comparison_result = result_message
            # Update the previous_data with the result (result is based on Smart Assist file)
            previous_data[previous_sheet] = result_df
        else:
            comparison_result = result_message

        return redirect("/comparison?tab=comparison")

    except Exception as e:
        comparison_result = f"❌ Error comparing files: {str(e)}"
        return redirect("/comparison?tab=comparison")


@app.route("/download_result", methods=["POST"])
def download_result():
    global previous_data, previous_filename

    if not previous_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"smart_assist_result_{timestamp}.xlsx"

    try:
        # Create a temporary file
        import tempfile

        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in previous_data.items():
                    df_clean = df.copy()

                    # Format "Appt Date" column to MM/DD/YYYY format (flexible column name search)
                    appt_date_col = None
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        # Check for variations: "appt date", "appointment date", "apptdate", etc.
                        if "appt" in col_lower and "date" in col_lower:
                            appt_date_col = col
                            break

                    if appt_date_col:
                        # Convert dates to MM/DD/YYYY format
                        def format_date(date_val):
                            if pd.isna(date_val) or date_val == "":
                                return ""
                            try:
                                # Convert to datetime if not already
                                if isinstance(date_val, pd.Timestamp):
                                    date_obj = date_val
                                elif isinstance(date_val, str):
                                    # Try to parse string date
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse
                                else:
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse

                                # Format as MM/DD/YYYY
                                return date_obj.strftime("%m/%d/%Y")
                            except (ValueError, TypeError, AttributeError):
                                # If parsing fails, return as-is
                                return str(date_val)

                        # Format dates and convert column to string type to prevent Excel auto-formatting
                        df_clean[appt_date_col] = df_clean[appt_date_col].apply(
                            format_date
                        )
                        df_clean[appt_date_col] = df_clean[appt_date_col].astype(str)

                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Set date column format to text in Excel to preserve MM/DD/YYYY format
                    if appt_date_col:
                        ws = writer.sheets[sheet_name]

                        # Find the column index
                        col_idx = None
                        for idx, col_name in enumerate(df_clean.columns, 1):
                            if col_name == appt_date_col:
                                col_idx = idx
                                break

                        if col_idx:
                            # Set all cells in this column to text format
                            for row in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row, column=col_idx)
                                if cell.value:
                                    cell.number_format = "@"  # Text format

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/upload_conversion", methods=["POST"])
def upload_conversion_file():
    global conversion_data, conversion_filename, conversion_result

    if "file" not in request.files:
        conversion_result = "❌ Error: No file provided"
        return redirect("/comparison?tab=conversion")

    file = request.files["file"]
    if file.filename == "":
        conversion_result = "❌ Error: No file selected"
        return redirect("/comparison?tab=conversion")

    try:
        # Get filename without saving to disk
        filename = secure_filename(file.filename)

        # Read Excel file directly from memory (no disk storage)
        file.seek(0)  # Reset file pointer to beginning
        conversion_data = pd.read_excel(file, sheet_name=None, engine="openpyxl")
        conversion_filename = filename

        # Remove "Unnamed:" columns from all sheets
        cleaned_data = {}
        for sheet_name, df in conversion_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned
        conversion_data = cleaned_data

        # Validate "Insurance Note" column exists in all sheets
        missing_sheets = []
        for sheet_name, df in conversion_data.items():
            columns = [col.lower().strip() for col in df.columns]
            if "insurance note" not in columns:
                missing_sheets.append(sheet_name)

        if missing_sheets:
            conversion_result = f"❌ Validation Error: 'Insurance Note' column not found in the following sheets: {', '.join(missing_sheets)}\n\nAvailable columns in first sheet: {list(conversion_data[list(conversion_data.keys())[0]].columns) if conversion_data else 'N/A'}"
            conversion_data = None
            conversion_filename = None
        else:
            # Validation successful - now process the Insurance Note column
            processed_sheets = {}
            total_rows_processed = 0
            patient_name_created = False
            total_duplicates_removed = 0

            for sheet_name, df in conversion_data.items():
                # Find Insurance Note column (case-insensitive)
                insurance_note_col = None
                for col in df.columns:
                    if col.lower().strip() == "insurance note":
                        insurance_note_col = col
                        break

                if insurance_note_col:
                    # Create a copy of the dataframe
                    processed_df = df.copy()

                    # Remove "Conversion" column if it exists (we don't want this column)
                    if "Conversion" in processed_df.columns:
                        processed_df = processed_df.drop(columns=["Conversion"])

                    # Extract insurance names and format them
                    def extract_and_format_insurance(note_text):
                        """Extract insurance name from Insurance Note text and format it"""
                        if pd.isna(note_text) or note_text == "":
                            return ""

                        note_str = str(note_text).strip()
                        import re

                        # Pattern: "From Conversion Carrier: <insurance_name> | ..." or "from conversion carrier: <insurance_name>|..."
                        # Use regex to find "From Conversion Carrier:" (case-insensitive) and extract everything after it until "|" or end of string
                        pattern = (
                            r"from\s+conversion\s+carrier\s*:\s*([^|]+?)(?:\s*\||$)"
                        )
                        match = re.search(pattern, note_str, re.IGNORECASE)

                        if match:
                            insurance_name = match.group(1).strip()
                            # Format the insurance name using existing function
                            return format_insurance_name(insurance_name)

                        # If pattern doesn't match, try to format the whole text
                        return format_insurance_name(note_str)

                    # Extract status from Insurance Note
                    def extract_status(note_text):
                        """Extract status from Insurance Note text - captures only the status value and discards all text after it.
                        Handles multi-word statuses (e.g., "Not Eligible") and stops at the first "-" character after the status.
                        Returns 'Conversion' if no status value is found."""
                        if pd.isna(note_text) or note_text == "":
                            return "Conversion"

                        note_str = str(note_text).strip()

                        # Pattern: "Status - <status>" - capture the status value, stop at first "-" after status or other delimiters
                        # Look for "Status -" or "Status-" followed by the status value
                        # Status value can be multiple words (e.g., "Not Eligible", "Eligible")
                        # Stop at the first "-" character after the status, or at ], |, comma, or end of string
                        import re

                        # Match "Status -" or "Status-" followed by the status value
                        # Capture everything (including spaces for multi-word statuses) until we hit:
                        # - A "-" character (not the one after "Status")
                        # - A "]" bracket
                        # - A "|" pipe
                        # - A comma
                        # - End of string
                        # Use non-greedy matching to stop at the first delimiter
                        status_pattern = r"Status\s*-\s*([^-]+?)(?:\s*-|]|\||,|$)"
                        match = re.search(status_pattern, note_str, re.IGNORECASE)
                        if match:
                            status = match.group(1).strip()
                            # Remove any trailing whitespace or delimiters
                            status = status.rstrip(" ]|,")
                            return status

                        # If no status found, return 'Conversion' instead of empty string
                        return "Conversion"

                    # Apply extraction and formatting
                    processed_df["Formatted Insurance"] = processed_df[
                        insurance_note_col
                    ].apply(extract_and_format_insurance)
                    processed_df["Status"] = processed_df[insurance_note_col].apply(
                        extract_status
                    )

                    # Update Status based on Eligibility column
                    # Find Eligibility column (case-insensitive, flexible matching)
                    eligibility_col = None
                    for col in processed_df.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        if "eligibility" in col_lower:
                            eligibility_col = col
                            break

                    if eligibility_col:
                        # Apply logic: ✗ -> Workable, ✓ -> Completed
                        def set_status_from_eligibility(row):
                            eligibility_val = row[eligibility_col]
                            if pd.notna(eligibility_val):
                                eligibility_str = str(eligibility_val).strip()
                                if (
                                    "✗" in eligibility_str
                                    or "x" in eligibility_str.lower()
                                ):
                                    return "Workable"
                                elif (
                                    "✓" in eligibility_str
                                    or "√" in eligibility_str
                                    or "check" in eligibility_str.lower()
                                ):
                                    return "Completed"
                            # If no eligibility match, keep existing status
                            return row["Status"]

                        processed_df["Status"] = processed_df.apply(
                            set_status_from_eligibility, axis=1
                        )

                    # Create "Patient Name" column from "Patient Last Name" and "Patient First Name"
                    def find_column(columns, target_names):
                        """Find column matching any of the target names (case-insensitive, flexible matching)"""
                        for col in columns:
                            col_lower = (
                                col.lower().strip().replace(" ", "").replace("_", "")
                            )
                            for target in target_names:
                                target_lower = (
                                    target.lower()
                                    .strip()
                                    .replace(" ", "")
                                    .replace("_", "")
                                )
                                if col_lower == target_lower:
                                    return col
                        # Try partial matching as fallback
                        for col in columns:
                            col_lower = col.lower().strip()
                            for target in target_names:
                                target_lower = target.lower().strip()
                                if (
                                    target_lower in col_lower
                                    or col_lower in target_lower
                                ):
                                    return col
                        return None

                    patient_last_name_col = find_column(
                        processed_df.columns,
                        [
                            "Patient Last Name",
                            "PatientLastName",
                            "patient last name",
                            "Last Name",
                            "LastName",
                        ],
                    )
                    patient_first_name_col = find_column(
                        processed_df.columns,
                        [
                            "Patient First Name",
                            "PatientFirstName",
                            "patient first name",
                            "First Name",
                            "FirstName",
                        ],
                    )

                    if patient_last_name_col and patient_first_name_col:
                        # Create Patient Name column with format "<last_name>, <first_name>"
                        def combine_names(row):
                            last_name = (
                                str(row[patient_last_name_col]).strip()
                                if pd.notna(row[patient_last_name_col])
                                else ""
                            )
                            first_name = (
                                str(row[patient_first_name_col]).strip()
                                if pd.notna(row[patient_first_name_col])
                                else ""
                            )

                            # Format as "<last_name>, <first_name>"
                            if last_name and first_name:
                                return f"{last_name}, {first_name}"
                            elif last_name:
                                return last_name
                            elif first_name:
                                return first_name
                            else:
                                return ""

                        # Create the Patient Name column
                        processed_df["Patient Name"] = processed_df.apply(
                            combine_names, axis=1
                        )
                        patient_name_created = True

                        # Remove Patient Last Name and Patient First Name columns
                        processed_df = processed_df.drop(
                            columns=[patient_last_name_col, patient_first_name_col]
                        )

                    # Find position of Insurance Note column and insert new columns after it
                    insurance_note_index = processed_df.columns.get_loc(
                        insurance_note_col
                    )
                    # Move the new columns to right after Insurance Note
                    cols = list(processed_df.columns)
                    # Remove new columns from their current position (if they exist)
                    cols_to_remove = ["Formatted Insurance", "Status"]
                    if "Patient Name" in cols:
                        cols_to_remove.append("Patient Name")
                    for col_name in cols_to_remove:
                        if col_name in cols:
                            cols.remove(col_name)
                    # Insert new columns after Insurance Note
                    cols.insert(insurance_note_index + 1, "Formatted Insurance")
                    cols.insert(insurance_note_index + 2, "Status")
                    # Insert Patient Name column if it was created (after Status)
                    if "Patient Name" in processed_df.columns:
                        status_index = cols.index("Status")
                        cols.insert(status_index + 1, "Patient Name")

                    # Ensure Patient Name is in cols if it was created
                    if patient_name_created:
                        if "Patient Name" not in cols:
                            # Add it after Status
                            if "Status" in cols:
                                status_index = cols.index("Status")
                                cols.insert(status_index + 1, "Patient Name")
                            else:
                                # If Status not found, add at end
                                cols.append("Patient Name")

                    # Reorder dataframe with all columns
                    processed_df = processed_df[cols]

                    # Rename "Formatted Insurance" to "Dental Primary Ins Carr" and create "Dental Secondary Ins Carr"
                    if "Formatted Insurance" in processed_df.columns:
                        # Find Pat ID column (flexible matching)
                        pat_id_col = None
                        pat_id_variations = [
                            "Pat ID",
                            "PATID",
                            "PatID",
                            "pat id",
                            "Patient ID",
                            "PatientID",
                            "patient id",
                        ]
                        for col in processed_df.columns:
                            col_lower = (
                                col.lower().strip().replace(" ", "").replace("_", "")
                            )
                            for variation in pat_id_variations:
                                if col_lower == variation.lower().strip().replace(
                                    " ", ""
                                ).replace("_", ""):
                                    pat_id_col = col
                                    break
                            if pat_id_col:
                                break

                        if pat_id_col:
                            # Count rows before consolidation
                            rows_before = len(processed_df)

                            # Group by Pat ID and consolidate insurance information
                            def consolidate_by_pat_id(group):
                                # Get unique insurance names for this Pat ID (preserve order of first occurrence)
                                insurance_names = []
                                seen = set()
                                for ins in group["Formatted Insurance"]:
                                    if pd.notna(ins):
                                        ins_str = str(ins).strip()
                                        if ins_str and ins_str not in seen:
                                            insurance_names.append(ins_str)
                                            seen.add(ins_str)

                                # Take the first row as base
                                first_row = group.iloc[0].copy()

                                # Set Primary Insurance (first unique insurance)
                                if len(insurance_names) > 0:
                                    first_row["Dental Primary Ins Carr"] = (
                                        insurance_names[0]
                                    )
                                else:
                                    first_row["Dental Primary Ins Carr"] = ""

                                # Set Secondary Insurance (second unique insurance if exists)
                                if len(insurance_names) > 1:
                                    first_row["Dental Secondary Ins Carr"] = (
                                        insurance_names[1]
                                    )
                                else:
                                    first_row["Dental Secondary Ins Carr"] = ""

                                return first_row.to_dict()

                            # Group by Pat ID and consolidate
                            consolidated_rows = []
                            for pat_id, group in processed_df.groupby(pat_id_col):
                                consolidated_row = consolidate_by_pat_id(group)
                                consolidated_rows.append(consolidated_row)

                            # Create new DataFrame from consolidated rows
                            processed_df = pd.DataFrame(consolidated_rows)

                            # Remove the old "Formatted Insurance" column
                            if "Formatted Insurance" in processed_df.columns:
                                processed_df = processed_df.drop(
                                    columns=["Formatted Insurance"]
                                )

                            # Ensure "Dental Primary Ins Carr" and "Dental Secondary Ins Carr" columns exist
                            if "Dental Primary Ins Carr" not in processed_df.columns:
                                processed_df["Dental Primary Ins Carr"] = ""
                            if "Dental Secondary Ins Carr" not in processed_df.columns:
                                processed_df["Dental Secondary Ins Carr"] = ""

                            # Reorder columns to place insurance columns after Pat ID
                            cols_list = list(processed_df.columns)
                            # Remove insurance columns from their current position
                            for col_name in [
                                "Dental Primary Ins Carr",
                                "Dental Secondary Ins Carr",
                            ]:
                                if col_name in cols_list:
                                    cols_list.remove(col_name)

                            # Find Pat ID column position
                            if pat_id_col in cols_list:
                                pat_id_index = cols_list.index(pat_id_col)
                                # Insert insurance columns after Pat ID
                                cols_list.insert(
                                    pat_id_index + 1, "Dental Primary Ins Carr"
                                )
                                cols_list.insert(
                                    pat_id_index + 2, "Dental Secondary Ins Carr"
                                )
                            else:
                                # If Pat ID not found, add at the end
                                cols_list.append("Dental Primary Ins Carr")
                                cols_list.append("Dental Secondary Ins Carr")

                            # Ensure Status column is preserved in the column list
                            if (
                                "Status" in processed_df.columns
                                and "Status" not in cols_list
                            ):
                                cols_list.append("Status")

                            processed_df = processed_df[cols_list]

                            # Count rows after consolidation
                            rows_after = len(processed_df)
                            rows_consolidated = rows_before - rows_after
                            total_duplicates_removed += rows_consolidated
                        else:
                            # If Pat ID column not found, just rename the column
                            processed_df = processed_df.rename(
                                columns={
                                    "Formatted Insurance": "Dental Primary Ins Carr"
                                }
                            )
                            processed_df["Dental Secondary Ins Carr"] = ""
                            # Ensure Status column exists
                            if "Status" not in processed_df.columns:
                                processed_df["Status"] = "Conversion"
                    else:
                        # If Formatted Insurance column doesn't exist, create empty insurance columns
                        processed_df["Dental Primary Ins Carr"] = ""
                        processed_df["Dental Secondary Ins Carr"] = ""
                        # Ensure Status column exists
                        if "Status" not in processed_df.columns:
                            processed_df["Status"] = "Conversion"

                    processed_sheets[sheet_name] = processed_df
                    total_rows_processed += len(processed_df)

                    # Debug: Print columns to verify Status is included
                    print(
                        f"DEBUG - Sheet '{sheet_name}' columns: {list(processed_df.columns)}"
                    )
                    if "Status" in processed_df.columns:
                        print(
                            f"DEBUG - Status column exists with values: {processed_df['Status'].unique()[:5]}"
                        )
                    else:
                        print("DEBUG - WARNING: Status column NOT FOUND!")
                else:
                    processed_sheets[sheet_name] = df

            # Update conversion_data with processed data
            conversion_data = processed_sheets

            # Build columns added message
            columns_added_msg = "- 'Dental Primary Ins Carr' - Extracted and formatted insurance names (first insurance for each Pat ID)\n- 'Dental Secondary Ins Carr' - Second insurance for Pat IDs with multiple insurances\n- 'Status' - Extracted status values (shows 'Conversion' when no status is found)"
            if patient_name_created:
                columns_added_msg += "\n- 'Patient Name' - Created from 'Patient Last Name' and 'Patient First Name' (format: <last_name>, <first_name>)"

            # Build consolidation message
            dedup_msg = ""
            if total_duplicates_removed > 0:
                dedup_msg = f"\n\n🔄 Row Consolidation:\n- Consolidated {total_duplicates_removed} row(s) with same Pat ID\n- Multiple insurances for same Pat ID: first in 'Dental Primary Ins Carr', second in 'Dental Secondary Ins Carr'"

            conversion_result = f"✅ Validation and processing completed successfully!\n\n📊 File loaded: {filename}\n📋 Sheets processed: {len(conversion_data)}\n📋 Sheet names: {', '.join(list(conversion_data.keys()))}\n📊 Total rows processed: {total_rows_processed}{dedup_msg}\n\n✅ New columns added:\n{columns_added_msg}\n💾 Ready to download the processed file!"

        return redirect("/comparison?tab=conversion")

    except Exception as e:
        conversion_result = f"❌ Error uploading Conversion Report: {str(e)}"
        conversion_data = None
        conversion_filename = None
        return redirect("/comparison?tab=conversion")


@app.route("/download_conversion", methods=["POST"])
def download_conversion_result():
    global conversion_data, conversion_filename

    if not conversion_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"conversion_report_{timestamp}.xlsx"

    try:
        # Create a temporary file
        import tempfile

        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in conversion_data.items():
                    # Remove "Conversion" column if it exists (safety check)
                    df_clean = df.copy()
                    if "Conversion" in df_clean.columns:
                        df_clean = df_clean.drop(columns=["Conversion"])

                    # Format "Appt Date" column to MM/DD/YYYY format
                    appt_date_col = None
                    for col in df_clean.columns:
                        if col.lower().strip() == "appt date":
                            appt_date_col = col
                            break

                    if appt_date_col:
                        # Convert dates to MM/DD/YYYY format
                        def format_date(date_val):
                            if pd.isna(date_val) or date_val == "":
                                return ""
                            try:
                                # Convert to datetime if not already
                                if isinstance(date_val, pd.Timestamp):
                                    date_obj = date_val
                                elif isinstance(date_val, str):
                                    # Try to parse string date
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse
                                else:
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse

                                # Format as MM/DD/YYYY
                                return date_obj.strftime("%m/%d/%Y")
                            except (ValueError, TypeError, AttributeError):
                                # If parsing fails, return as-is
                                return str(date_val)

                        df_clean[appt_date_col] = df_clean[appt_date_col].apply(
                            format_date
                        )

                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

            # Clear data after successful download
            global conversion_result
            conversion_data = None
            conversion_filename = None
            conversion_result = None

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reset_comparison", methods=["POST"])
def reset_comparison():
    global raw_data, previous_data, raw_filename, previous_filename, comparison_result
    # Explicitly do NOT touch conversion_data, conversion_filename, or conversion_result

    try:
        # Reset ONLY comparison tool variables - do not affect conversion tool
        raw_data = {}
        previous_data = {}
        raw_filename = None
        previous_filename = None
        comparison_result = "🔄 Comparison tool reset successfully! All files and data have been cleared."

        return redirect("/comparison?tab=comparison")

    except Exception as e:
        comparison_result = f"❌ Error resetting comparison tool: {str(e)}"
        return redirect("/comparison?tab=comparison")


@app.route("/reset_conversion", methods=["POST"])
def reset_conversion():
    global conversion_data, conversion_filename, conversion_result
    # Explicitly do NOT touch raw_data, previous_data, raw_filename, previous_filename, or comparison_result

    try:
        # Reset ONLY conversion tool variables - do not affect comparison tool
        conversion_data = {}
        conversion_filename = None
        conversion_result = "🔄 Conversion tool reset successfully! All files and data have been cleared."

        return redirect("/comparison?tab=conversion")

    except Exception as e:
        conversion_result = f"❌ Error resetting conversion tool: {str(e)}"
        return redirect("/comparison?tab=conversion")


def reformat_insurance_column(df, source_column, new_column_name):
    """Reformat the source column and create a new column next to it"""
    if source_column not in df.columns:
        return None, f"Column '{source_column}' not found in sheet."

    # Apply the reformatting
    df[new_column_name] = df[source_column].apply(format_insurance_name)

    # Get the position of the source column
    source_col_idx = df.columns.get_loc(source_column)

    # Reorder columns to place new column right after source column
    cols = df.columns.tolist()
    # Remove new column from its current position
    cols.remove(new_column_name)
    # Insert it right after source column
    cols.insert(source_col_idx + 1, new_column_name)
    df = df[cols]

    return df, None


def process_insurance_formatting(
    data_dict,
    source_column="Dental Primary Ins Carr",
    new_column_name="formated insurance names",
):
    """Process all sheets in the Excel file for insurance formatting"""
    output_lines = []
    output_lines.append("=" * 70)
    output_lines.append("PROCESSING EXCEL FILE - REFORMATTING COLUMNS")
    output_lines.append("=" * 70)
    output_lines.append("")

    processed_sheets = {}

    for sheet_name, df in data_dict.items():
        output_lines.append(f"📋 Processing sheet: {sheet_name}")
        output_lines.append(
            f"   Original shape: {df.shape[0]} rows × {df.shape[1]} columns"
        )

        if source_column not in df.columns:
            output_lines.append(
                f"   ⚠️  Column '{source_column}' not found. Skipping this sheet."
            )
            processed_sheets[sheet_name] = df
            output_lines.append("")
            continue

        # Reformat the column
        df_processed, error = reformat_insurance_column(
            df.copy(), source_column, new_column_name
        )

        if error:
            output_lines.append(f"   ❌ Error: {error}")
            processed_sheets[sheet_name] = df
        else:
            processed_sheets[sheet_name] = df_processed
            output_lines.append(f"   ✅ Column reformatted successfully!")
            output_lines.append(
                f"   New shape: {df_processed.shape[0]} rows × {df_processed.shape[1]} columns"
            )

            # Show sample
            sample_df = df_processed[[source_column, new_column_name]].head(10)
            output_lines.append(f"   Sample of original vs formatted (first 10 rows):")
            output_lines.append(sample_df.to_string(index=False))
            output_lines.append(
                f"   Total formatted entries: {df_processed[new_column_name].notna().sum()}"
            )
            output_lines.append(
                f"   Unique formatted values: {df_processed[new_column_name].value_counts().head(10).to_string()}"
            )

        output_lines.append("")

    output_lines.append("=" * 70)
    output_lines.append("PROCESSING COMPLETE!")
    output_lines.append("=" * 70)

    return processed_sheets, "\n".join(output_lines)


@app.route("/upload_insurance_formatting", methods=["POST"])
def upload_insurance_formatting():
    global insurance_formatting_data, insurance_formatting_filename, insurance_formatting_result, insurance_formatting_output

    if "file" not in request.files:
        insurance_formatting_result = "❌ Error: No file provided"
        return redirect("/comparison?tab=insurance")

    file = request.files["file"]
    if file.filename == "":
        insurance_formatting_result = "❌ Error: No file selected"
        return redirect("/comparison?tab=insurance")

    try:
        # Get filename without saving to disk
        filename = secure_filename(file.filename)

        # Read Excel file directly from memory (no disk storage)
        file.seek(0)  # Reset file pointer to beginning
        excel_data = pd.read_excel(file, sheet_name=None, engine="openpyxl")

        # Remove "Unnamed:" columns from all sheets
        cleaned_data = {}
        for sheet_name, df in excel_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned

        # Process all sheets automatically
        insurance_formatting_data, insurance_formatting_output = (
            process_insurance_formatting(cleaned_data)
        )
        insurance_formatting_filename = filename

        # Count sheets processed
        sheets_count = len(insurance_formatting_data)
        insurance_formatting_result = f"✅ Processing complete! Processed {sheets_count} sheet(s). Formatted insurance names column added to all sheets."

        return redirect("/comparison?tab=insurance")

    except Exception as e:
        insurance_formatting_result = f"❌ Error processing file: {str(e)}"
        insurance_formatting_output = f"Error: {str(e)}"
        return redirect("/comparison?tab=insurance")


@app.route("/download_insurance_formatting", methods=["POST"])
def download_insurance_formatting():
    global insurance_formatting_data, insurance_formatting_filename, insurance_formatting_result, insurance_formatting_output

    if not insurance_formatting_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"formatted_insurance_names_{timestamp}.xlsx"

    try:
        # Create a temporary file
        import tempfile

        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in insurance_formatting_data.items():
                    df_clean = df.copy()

                    # Format "Appointment Date" column to MM/DD/YYYY format (flexible column name search)
                    appt_date_col = None
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        # Check for variations: "appointment date", "appt date", "apptdate", etc.
                        if (
                            "appointment" in col_lower or "appt" in col_lower
                        ) and "date" in col_lower:
                            appt_date_col = col
                            break

                    if appt_date_col:
                        # Convert dates to MM/DD/YYYY format
                        def format_date(date_val):
                            if pd.isna(date_val) or date_val == "":
                                return ""
                            try:
                                # Convert to datetime if not already
                                if isinstance(date_val, pd.Timestamp):
                                    date_obj = date_val
                                elif isinstance(date_val, str):
                                    # Try to parse string date
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse
                                else:
                                    date_obj = pd.to_datetime(date_val, errors="coerce")
                                    if pd.isna(date_obj):
                                        return str(
                                            date_val
                                        )  # Return original if can't parse

                                # Format as MM/DD/YYYY
                                return date_obj.strftime("%m/%d/%Y")
                            except (ValueError, TypeError, AttributeError):
                                # If parsing fails, return as-is
                                return str(date_val)

                        # Format dates and convert column to string type to prevent Excel auto-formatting
                        df_clean[appt_date_col] = df_clean[appt_date_col].apply(
                            format_date
                        )
                        df_clean[appt_date_col] = df_clean[appt_date_col].astype(str)

                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Set date column format to text in Excel to preserve MM/DD/YYYY format
                    if appt_date_col:
                        ws = writer.sheets[sheet_name]

                        # Find the column index
                        col_idx = None
                        for idx, col_name in enumerate(df_clean.columns, 1):
                            if col_name == appt_date_col:
                                col_idx = idx
                                break

                        if col_idx:
                            # Set all cells in this column to text format
                            for row in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row, column=col_idx)
                                if cell.value:
                                    cell.number_format = "@"  # Text format

            # Clear data after successful download
            insurance_formatting_data = None
            insurance_formatting_filename = None
            insurance_formatting_result = None
            insurance_formatting_output = ""

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reset_insurance_formatting", methods=["POST"])
def reset_insurance_formatting():
    global insurance_formatting_data, insurance_formatting_filename, insurance_formatting_result, insurance_formatting_output
    # Explicitly do NOT touch other tool variables

    try:
        # Reset ONLY insurance formatting tool variables
        insurance_formatting_data = {}
        insurance_formatting_filename = None
        insurance_formatting_result = "🔄 Insurance formatting tool reset successfully! All files and data have been cleared."
        insurance_formatting_output = ""

        return redirect("/comparison?tab=insurance")

    except Exception as e:
        insurance_formatting_result = (
            f"❌ Error resetting insurance formatting tool: {str(e)}"
        )
        return redirect("/comparison?tab=insurance")


# Remarks update functions (from excel-remark)
def process_remarks_excel_file(file_stream):
    """Process uploaded Excel file and extract Patient ID, Remark, and Agent Name data.
    Checks all sheets to find the one with required columns."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_stream)
        ws = None
        patient_id_col = None
        remark_col = None
        agent_name_col = None
        header_row = 1
        selected_sheet_name = None  # Track which sheet was selected

        # Check if "Today" sheet exists
        if "Today" not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise Exception(
                f'Sheet "Today" not found in Excel file. Available sheets: {available_sheets}'
            )

        # Use only the "Today" sheet
        ws = wb["Today"]
        selected_sheet_name = "Today"

        # Look for headers in the first few rows
        for row_num in range(1, min(6, ws.max_row + 1)):
            temp_patient_id_col = None
            temp_remark_col = None
            temp_agent_name_col = None

            for col in range(1, ws.max_column + 1):
                cell_value = (
                    str(ws.cell(row=row_num, column=col).value or "").strip().lower()
                )
                cell_value_clean = (
                    cell_value.replace(" ", "")
                    .replace("_", "")
                    .replace("-", "")
                    .replace(".", "")
                )

                if (
                    "patient" in cell_value_clean and "id" in cell_value_clean
                ) or cell_value_clean == "pid":
                    temp_patient_id_col = col
                elif (
                    "remark" in cell_value_clean
                    or cell_value_clean == "remarks"
                    or "note" in cell_value_clean
                ):
                    # More flexible remark detection - also check for 'remarks' or 'note'
                    if not temp_remark_col:  # Only set if not already found
                        temp_remark_col = col
                elif "agent" in cell_value_clean and "name" in cell_value_clean:
                    temp_agent_name_col = col

            # If we found Patient ID and Remark (Agent Name is optional), use this header row
            if temp_patient_id_col and temp_remark_col:
                patient_id_col = temp_patient_id_col
                remark_col = temp_remark_col
                agent_name_col = temp_agent_name_col
                header_row = row_num
                break
            for col in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=1, column=col).value or "").strip().lower()
                cell_value_clean = (
                    cell_value.replace(" ", "")
                    .replace("_", "")
                    .replace("-", "")
                    .replace(".", "")
                )

                if (
                    "patient" in cell_value_clean and "id" in cell_value_clean
                ) or cell_value_clean == "pid":
                    patient_id_col = col
                elif "remark" in cell_value_clean:
                    remark_col = col
                elif "agent" in cell_value_clean and "name" in cell_value_clean:
                    agent_name_col = col

        if not patient_id_col:
            raise Exception(
                f'Patient ID column not found in "Today" sheet. Please ensure the sheet contains a Patient ID column.'
            )

        if not remark_col:
            raise Exception(
                f'Remark column not found in "Today" sheet. Please ensure the sheet contains a Remark column.'
            )

        # Extract data - now returns list of records for each patient ID
        excel_data = {}
        data_start_row = header_row + 1
        for row in range(data_start_row, ws.max_row + 1):  # Skip header row
            patient_id_raw = ws.cell(row=row, column=patient_id_col).value
            remark_raw = ws.cell(row=row, column=remark_col).value
            agent_name_raw = (
                ws.cell(row=row, column=agent_name_col).value
                if agent_name_col
                else None
            )

            # Normalize Patient ID for consistent matching
            patient_id = normalize_patient_id(patient_id_raw) if patient_id_raw else ""

            # Ensure remark and agent_name are strings, handling None, empty strings, and whitespace
            # Convert None to empty string, then strip whitespace
            if remark_raw is None:
                remark = ""
            else:
                remark = str(remark_raw).strip() if str(remark_raw).strip() else ""

            if agent_name_raw is None:
                agent_name = ""
            else:
                agent_name = (
                    str(agent_name_raw).strip() if str(agent_name_raw).strip() else ""
                )

            if patient_id:  # Only add non-empty patient IDs
                if patient_id not in excel_data:
                    excel_data[patient_id] = []

                excel_data[patient_id].append(
                    {"remark": remark, "agent_name": agent_name}
                )

        # Store sheet name as metadata (will be removed in update_appointments_with_remarks)
        excel_data["_debug_sheet_name"] = selected_sheet_name

        return excel_data

    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")


def process_remarks_appointments_excel(file_stream):
    """Read an appointments Excel and return list of appointment dicts with all columns.

    Only requires 'Pat ID' column. All other columns are preserved as-is.
    Checks all sheets to find the one with Pat ID column.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_stream)
    ws = None
    headers = []
    pat_id_col = None
    header_row = 1

    # Try all sheets to find the one with Pat ID column
    for sheet_name in wb.sheetnames:
        current_ws = wb[sheet_name]

        # Try to find header row (check first 5 rows)
        for row_num in range(1, min(6, current_ws.max_row + 1)):
            temp_headers = []
            for col in range(1, current_ws.max_column + 1):
                raw = current_ws.cell(row=row_num, column=col).value
                name = (str(raw or "")).strip()
                temp_headers.append(name)

            # Check if this row looks like headers (has a Pat ID column)
            for i, header in enumerate(temp_headers):
                header_lower = (
                    header.lower()
                    .replace(" ", "")
                    .replace("_", "")
                    .replace("-", "")
                    .replace(".", "")
                )
                # Check for various patterns: pat id, patient id, patientid, patid, etc.
                if (
                    "pat" in header_lower and "id" in header_lower
                ) or header_lower == "pid":
                    headers = temp_headers
                    header_row = row_num
                    pat_id_col = i + 1  # 1-based column index
                    ws = current_ws
                    break

            if pat_id_col:
                break

        if pat_id_col:
            break

    # If still not found, use active sheet and check again
    if pat_id_col is None:
        ws = wb.active
        headers = []
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=1, column=col).value
            name = (str(raw or "")).strip()
            headers.append(name)

        for i, header in enumerate(headers):
            header_lower = (
                header.lower()
                .replace(" ", "")
                .replace("_", "")
                .replace("-", "")
                .replace(".", "")
            )
            if (
                "pat" in header_lower and "id" in header_lower
            ) or header_lower == "pid":
                pat_id_col = i + 1
                header_row = 1
                break

    if pat_id_col is None:
        # Provide helpful error message with found columns
        sheet_names = ", ".join(wb.sheetnames)
        found_columns = ", ".join([f"'{h}'" for h in headers if h]) or "none"
        raise Exception(
            f"Pat ID column not found in appointments Excel. Checked sheets: {sheet_names}. Found columns: {found_columns}. Please ensure there's a column containing 'Pat ID', 'Patient ID', or similar."
        )

    # Read all rows starting after the header row
    appointments = []
    data_start_row = header_row + 1
    for row in range(data_start_row, ws.max_row + 1):
        record = {}

        # Read all columns
        for col, header in enumerate(headers, 1):
            value = ws.cell(row=row, column=col).value
            record[header] = "" if value is None else str(value)

        # Normalize Patient ID using the same normalization function as remarks
        pat_id_value = record.get(headers[pat_id_col - 1], "")
        # Use normalize_patient_id for consistent matching with remarks file
        pid_normalized = normalize_patient_id(pat_id_value) if pat_id_value else ""
        # Store both normalized and original for maximum compatibility
        record["Pat ID"] = (
            pid_normalized if pid_normalized else str(pat_id_value).strip()
        )
        # Also store original for reference
        record["Pat ID Original"] = str(pat_id_value).strip() if pat_id_value else ""

        # Ensure Remark and Agent Name exist
        if "Remark" not in record:
            record["Remark"] = ""
        if "Agent Name" not in record:
            record["Agent Name"] = ""

        # Skip empty rows (no Pat ID)
        if pid_normalized or pat_id_value:
            appointments.append(record)

    return appointments


def update_appointments_with_remarks(appointments, excel_data):
    """Update appointments with remarks and agent names from Excel data based on Patient ID matching.
    Creates separate rows for each match when Patient ID appears multiple times."""
    updated_appointments = []
    updated_count = 0

    # Remove debug metadata if present
    excel_data.pop("_debug_sheet_name", None)

    # Create a comprehensive lookup dictionary with all possible Patient ID formats
    # Note: excel_data keys are already normalized from process_remarks_excel_file
    # But we'll create variations to handle any edge cases
    lookup_dict = {}

    # First, build lookup dictionary with all possible variations
    # Convert all keys to strings for consistent matching
    for pid, data_list in excel_data.items():
        if not pid:
            continue

        # Convert to string first for consistency
        pid_str_base = str(pid).strip()
        if not pid_str_base:
            continue

        # Store with string version (primary key)
        if pid_str_base not in lookup_dict:
            lookup_dict[pid_str_base] = []
        lookup_dict[pid_str_base].extend(data_list)

        # Also store with original key if it's different (for type matching)
        if pid != pid_str_base and pid not in lookup_dict:
            lookup_dict[pid] = []
            lookup_dict[pid].extend(data_list)

        # Store with .0 suffix removed (if it's a float string)
        if pid_str_base.endswith(".0"):
            pid_no_suffix = pid_str_base[:-2]
            if pid_no_suffix and pid_no_suffix not in lookup_dict:
                lookup_dict[pid_no_suffix] = []
            lookup_dict[pid_no_suffix].extend(data_list)

        # Store with .0 suffix added (in case appointments have it)
        if not pid_str_base.endswith(".0"):
            pid_with_suffix = f"{pid_str_base}.0"
            if pid_with_suffix not in lookup_dict:
                lookup_dict[pid_with_suffix] = []
            lookup_dict[pid_with_suffix].extend(data_list)

        # Re-normalize (in case there are any edge cases)
        # But only if it's different from what we already have
        renormalized = normalize_patient_id(pid) if pid else None
        if (
            renormalized
            and renormalized != pid_str_base
            and renormalized not in lookup_dict
        ):
            lookup_dict[renormalized] = []
            lookup_dict[renormalized].extend(data_list)

        # Also try normalizing the string version
        renormalized_str = normalize_patient_id(pid_str_base) if pid_str_base else None
        if (
            renormalized_str
            and renormalized_str != pid_str_base
            and renormalized_str not in lookup_dict
        ):
            lookup_dict[renormalized_str] = []
            lookup_dict[renormalized_str].extend(data_list)

    for appointment in appointments:
        # Get both normalized and original Patient ID
        patient_id_normalized = appointment.get("Pat ID", "")
        patient_id_original = appointment.get("Pat ID Original", "")
        # Fallback to normalized if original not available
        patient_id_raw = (
            patient_id_original if patient_id_original else patient_id_normalized
        )

        matches_found = False

        # Try multiple matching strategies
        match_keys_to_try = []

        if patient_id_normalized:
            # Strategy 1: Use the normalized version (primary)
            match_keys_to_try.append(patient_id_normalized)

        if patient_id_raw:
            # Strategy 2: Normalize the original (in case it wasn't normalized during reading)
            normalized_from_original = normalize_patient_id(patient_id_raw)
            if (
                normalized_from_original
                and normalized_from_original not in match_keys_to_try
            ):
                match_keys_to_try.append(normalized_from_original)

            # Strategy 3: Original string version
            pid_str = str(patient_id_raw).strip()
            if pid_str and pid_str not in match_keys_to_try:
                match_keys_to_try.append(pid_str)

            # Strategy 4: Remove .0 suffix if present
            if pid_str.endswith(".0"):
                pid_no_suffix = pid_str[:-2]
                if pid_no_suffix and pid_no_suffix not in match_keys_to_try:
                    match_keys_to_try.append(pid_no_suffix)

            # Strategy 5: Add .0 suffix (in case original doesn't have it)
            if not pid_str.endswith(".0"):
                pid_with_suffix = f"{pid_str}.0"
                if pid_with_suffix not in match_keys_to_try:
                    match_keys_to_try.append(pid_with_suffix)

        # Convert all match keys to strings for consistent comparison
        match_keys_to_try_str = [
            str(k).strip() if k else "" for k in match_keys_to_try if k
        ]
        # Also try the original keys (in case of type matching)
        match_keys_to_try_all = list(set(match_keys_to_try + match_keys_to_try_str))

        # Try all matching strategies
        for match_key in match_keys_to_try_all:
            if not match_key:
                continue
            # Try both the key as-is and as string
            match_found = False
            if match_key in lookup_dict:
                match_found = True
                lookup_key = match_key
            elif str(match_key).strip() in lookup_dict:
                match_found = True
                lookup_key = str(match_key).strip()

            if match_found:
                # Match found!
                for match_data in lookup_dict[lookup_key]:
                    new_appointment = appointment.copy()

                    # Remove any existing remark/Remark keys to avoid conflicts
                    keys_to_remove = [
                        k for k in new_appointment.keys() if k.lower() == "remark"
                    ]
                    for key in keys_to_remove:
                        del new_appointment[key]

                    # Ensure Remark and Agent Name are set as strings, handling None and empty values
                    remark_raw = match_data.get("remark", "")
                    if remark_raw is None:
                        remark_value = ""
                    else:
                        remark_value = str(remark_raw).strip()

                    agent_raw = match_data.get("agent_name", "")
                    if agent_raw is None:
                        agent_value = ""
                    else:
                        agent_value = str(agent_raw).strip()

                    # Explicitly set with uppercase keys
                    new_appointment["Remark"] = remark_value
                    new_appointment["Agent Name"] = agent_value
                    updated_appointments.append(new_appointment)
                    updated_count += 1
                    matches_found = True
                break  # Found a match, no need to try other keys

        # If no match was found, still include the appointment with empty remarks
        # This ensures all appointments are included in the output, even if they don't have matching remarks
        if not matches_found:
            new_appointment = appointment.copy()

            # Remove any existing remark/Remark keys to avoid conflicts
            keys_to_remove = [
                k for k in new_appointment.keys() if k.lower() == "remark"
            ]
            for key in keys_to_remove:
                del new_appointment[key]

            # Ensure Remark and Agent Name are set as empty strings for unmatched appointments
            new_appointment["Remark"] = ""
            new_appointment["Agent Name"] = ""
            updated_appointments.append(new_appointment)

    return updated_appointments, updated_count


def create_excel_from_appointments(appointments, filename):
    """Create Excel file from processed appointment data with all columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Appointment Data"

    if not appointments:
        return wb

    # Get all unique headers from all appointments
    all_headers = set()
    for appointment in appointments:
        all_headers.update(appointment.keys())

    # Convert to list and ensure Pat ID, Insurance Name, Remark, and Agent Name are at the end for visibility
    headers = list(all_headers)
    if "Pat ID" in headers:
        headers.remove("Pat ID")
    if "Insurance Name" in headers:
        headers.remove("Insurance Name")
    if "Remark" in headers:
        headers.remove("Remark")
    if "Agent Name" in headers:
        headers.remove("Agent Name")
    headers.extend(
        ["Pat ID", "Insurance Name", "Remark", "Agent Name"]
    )  # Put these at the end

    # Set headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Find "Time" column index for formatting
    time_col_idx = None
    for col, header in enumerate(headers, 1):
        if header.lower().strip() == "time":
            time_col_idx = col
            break

    # Format date function for Time column
    def format_time_value(time_val):
        """Format Time column value to MM/DD/YYYY format"""
        if not time_val or time_val == "":
            return ""
        try:
            # Convert to datetime if not already
            if isinstance(time_val, pd.Timestamp):
                date_obj = time_val
            elif isinstance(time_val, str):
                # Try to parse string date
                date_obj = pd.to_datetime(time_val, errors="coerce")
                if pd.isna(date_obj):
                    return str(time_val)  # Return original if can't parse
            else:
                date_obj = pd.to_datetime(time_val, errors="coerce")
                if pd.isna(date_obj):
                    return str(time_val)  # Return original if can't parse

            # Format as MM/DD/YYYY
            return date_obj.strftime("%m/%d/%Y")
        except (ValueError, TypeError, AttributeError):
            # If parsing fails, return as-is
            return str(time_val)

    # Add appointment data
    for row, appointment in enumerate(appointments, 2):
        for col, header in enumerate(headers, 1):
            value = appointment.get(header, "")

            # Ensure value is not None - convert to string if needed
            if value is None:
                value = ""
            elif not isinstance(value, str):
                value = str(value)

            # Format Time column if this is the Time column
            if time_col_idx and col == time_col_idx:
                value = format_time_value(value)

            # Write the value to the cell
            cell = ws.cell(row=row, column=col, value=value)

            # Set Time column format to text to preserve MM/DD/YYYY format
            if time_col_idx and col == time_col_idx and value:
                cell.number_format = "@"  # Text format

            # Ensure Remark and Agent Name columns are written as text
            if header in ["Remark", "Agent Name"] and value:
                cell.number_format = "@"  # Text format to preserve content

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


@app.route("/upload_remarks", methods=["POST"])
def upload_remarks():
    global remarks_appointments_data, remarks_excel_data, remarks_appointments_filename, remarks_remarks_filename, remarks_result, remarks_updated_count

    appointments_file = request.files.get("appointments_file")
    remarks_file = request.files.get("remarks_file")

    # Require both Excel files
    if not appointments_file or appointments_file.filename == "":
        remarks_result = "❌ Error: Please upload the Appointments Excel file."
        return redirect("/comparison?tab=remarks")

    if not remarks_file or remarks_file.filename == "":
        remarks_result = "❌ Error: Please upload the Remarks Excel file."
        return redirect("/comparison?tab=remarks")

    try:
        # Process appointments Excel directly from memory
        appointments_filename_raw = secure_filename(appointments_file.filename)
        appointments_file.seek(0)  # Reset file pointer
        remarks_appointments_data = process_remarks_appointments_excel(
            appointments_file
        )
        remarks_appointments_filename = appointments_filename_raw

        # Process remarks Excel
        remarks_file.seek(0)  # Reset file pointer
        remarks_excel_data = process_remarks_excel_file(remarks_file)
        remarks_remarks_filename = secure_filename(remarks_file.filename)

        # Debug: Check what remarks data we have
        sample_remarks_count = 0
        sample_remarks_with_data = 0
        sample_patient_ids = []
        for pid, data_list in list(remarks_excel_data.items())[
            :10
        ]:  # Check first 10 patient IDs
            sample_patient_ids.append(str(pid))
            sample_remarks_count += len(data_list)
            for data in data_list:
                if data.get("remark") and str(data.get("remark")).strip():
                    sample_remarks_with_data += 1

        # Debug: Check appointment Patient IDs (both normalized and original)
        sample_appointment_pids = []
        sample_appointment_pids_orig = []
        for appt in remarks_appointments_data[:10]:  # Check first 10 appointments
            pid_norm = appt.get("Pat ID", "")
            pid_orig = appt.get("Pat ID Original", "")
            if pid_norm:
                sample_appointment_pids.append(str(pid_norm))
            if pid_orig:
                sample_appointment_pids_orig.append(str(pid_orig))

        # Update appointments with remarks
        updated_appointments, updated_count = update_appointments_with_remarks(
            remarks_appointments_data, remarks_excel_data
        )

        # Debug: Verify remarks were set
        remarks_set_count = 0
        for appt in updated_appointments[:10]:  # Check first 10 appointments
            if appt.get("Remark") and str(appt.get("Remark")).strip():
                remarks_set_count += 1

        # Format Insurance Note column to create Insurance Name column
        for appointment in updated_appointments:
            # Find Insurance Note column (case-insensitive search)
            insurance_note_value = None
            insurance_note_key = None

            for key, value in appointment.items():
                if key.lower().strip() == "insurance note":
                    insurance_note_key = key
                    insurance_note_value = value
                    break

            # Format the insurance note value using existing format_insurance_name function
            if insurance_note_value:
                # Handle None, empty string, or NaN values
                insurance_str = (
                    str(insurance_note_value).strip() if insurance_note_value else ""
                )
                if insurance_str and insurance_str.lower() not in ["nan", "none", ""]:
                    try:
                        # Extract insurance name from "from conversion carrier:" pattern (case-insensitive)
                        # Example 1: "from conversion carrier: <insurance_name> | Plan #684 | ..."
                        # Example 2: "2025-10-29 18:34:01 PT, By - 5240HSHINDE, Status - Eligible ] From conversion carrier: <insurance_name>"
                        # We want to extract only the insurance name after "from conversion carrier:" and before the first "|" (if present)
                        extracted_insurance_name = None

                        # Check for "from conversion carrier:" pattern (case-insensitive)
                        if "from conversion carrier:" in insurance_str.lower():
                            # Find the position of "from conversion carrier:" (case-insensitive)
                            import re

                            pattern_match = re.search(
                                r"from conversion carrier:\s*(.+?)(?:\s*\||$)",
                                insurance_str,
                                re.IGNORECASE,
                            )

                            if pattern_match:
                                # Extract the insurance name (everything after "from conversion carrier:" until "|" or end of string)
                                insurance_name = pattern_match.group(1).strip()
                                # Remove any trailing characters like "]" if present
                                insurance_name = insurance_name.rstrip(" ]|")
                                extracted_insurance_name = insurance_name
                            else:
                                # Fallback: Split by "|" to get only the part before the first pipe
                                parts = insurance_str.split(
                                    "|", 1
                                )  # Split only on first "|"
                                if len(parts) > 0:
                                    carrier_part = parts[0].strip()
                                    # Find "from conversion carrier:" in this part (case-insensitive)
                                    carrier_match = re.search(
                                        r"from conversion carrier:\s*(.+?)$",
                                        carrier_part,
                                        re.IGNORECASE,
                                    )
                                    if carrier_match:
                                        insurance_name = carrier_match.group(1).strip()
                                        insurance_name = insurance_name.rstrip(" ]|")
                                        extracted_insurance_name = insurance_name
                                    elif (
                                        "from conversion carrier:"
                                        in carrier_part.lower()
                                    ):
                                        # Extract everything after "from conversion carrier:"
                                        insurance_name = (
                                            carrier_part.split(":", 1)[1].strip()
                                            if ":" in carrier_part
                                            else carrier_part
                                        )
                                        insurance_name = insurance_name.rstrip(" ]|")
                                        extracted_insurance_name = insurance_name

                        # If we extracted a name, use it; otherwise use the full text
                        insurance_to_format = (
                            extracted_insurance_name
                            if extracted_insurance_name
                            else insurance_str
                        )

                        # Format the insurance name using existing function
                        formatted_insurance = format_insurance_name(insurance_to_format)
                        appointment["Insurance Name"] = (
                            formatted_insurance
                            if formatted_insurance
                            and not (
                                isinstance(formatted_insurance, float)
                                and pd.isna(formatted_insurance)
                            )
                            else ""
                        )
                    except Exception as e:
                        # If formatting fails, use empty string
                        appointment["Insurance Name"] = ""
                else:
                    appointment["Insurance Name"] = ""
            else:
                appointment["Insurance Name"] = ""

        # Update the global processed_appointments with the new data
        remarks_appointments_data = updated_appointments
        remarks_updated_count = updated_count

        # Build result message with debug info
        result_msg = f"✅ Successfully processed {len(remarks_appointments_data)} appointment(s) and updated {updated_count} appointment(s) with remarks and agent names. Insurance Name column added based on Insurance Note formatting."

        # Add debug information if matching rate is low
        if (
            updated_count < len(remarks_appointments_data) * 0.5
        ):  # Less than 50% matched
            unmatched_count = len(remarks_appointments_data) - updated_count
            result_msg += f"\n\n⚠️ Warning: Only {updated_count} out of {len(remarks_appointments_data)} appointments were matched with remarks."
            result_msg += f"\n🔍 Sample Patient IDs from Remarks file (normalized): {', '.join(sample_patient_ids[:5]) if sample_patient_ids else 'None'}"
            result_msg += f"\n🔍 Sample Patient IDs from Appointments file (normalized): {', '.join(sample_appointment_pids[:5]) if sample_appointment_pids else 'None'}"
            if sample_appointment_pids_orig:
                result_msg += f"\n🔍 Sample Patient IDs from Appointments file (original): {', '.join(sample_appointment_pids_orig[:5])}"
            result_msg += f"\n📊 Total unique Patient IDs in Remarks file: {len(remarks_excel_data)}"
            result_msg += f"\n💡 Tip: Check if Patient IDs in both files match exactly. Check the console/terminal for detailed debug output."

        remarks_result = result_msg

        return redirect("/comparison?tab=remarks")

    except Exception as e:
        remarks_result = f"❌ Error processing files: {str(e)}"
        return redirect("/comparison?tab=remarks")


@app.route("/download_remarks", methods=["POST"])
def download_remarks():
    global remarks_appointments_data, remarks_appointments_filename, remarks_remarks_filename, remarks_result, remarks_updated_count

    if not remarks_appointments_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        if remarks_appointments_filename:
            base_name = os.path.splitext(remarks_appointments_filename)[0]
            filename = f"{base_name}_appointments.xlsx"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"appointments_with_remarks_{timestamp}.xlsx"

    try:
        # Create Excel file
        excel_buffer = create_excel_from_appointments(
            remarks_appointments_data, filename
        )

        # Clear data after successful download
        remarks_appointments_data = None
        remarks_excel_data = None
        remarks_appointments_filename = None
        remarks_remarks_filename = None
        remarks_result = None
        remarks_updated_count = 0

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reset_remarks", methods=["POST"])
def reset_remarks():
    global remarks_appointments_data, remarks_excel_data, remarks_appointments_filename, remarks_remarks_filename, remarks_result, remarks_updated_count
    # Explicitly do NOT touch other tool variables

    try:
        # Reset ONLY remarks tool variables
        remarks_appointments_data = None
        remarks_excel_data = None
        remarks_appointments_filename = None
        remarks_remarks_filename = None
        remarks_result = (
            "🔄 Remarks tool reset successfully! All files and data have been cleared."
        )
        remarks_updated_count = 0

        return redirect("/comparison?tab=remarks")

    except Exception as e:
        remarks_result = f"❌ Error resetting remarks tool: {str(e)}"
        return redirect("/comparison?tab=remarks")


@app.route("/upload_appointment_report", methods=["POST"])
def upload_appointment_report():
    global appointment_report_data, appointment_report_filename, appointment_report_result, appointment_report_output

    if "file" not in request.files:
        appointment_report_result = "❌ Error: No file provided"
        return redirect("/comparison?tab=appointment")

    file = request.files["file"]
    if file.filename == "":
        appointment_report_result = "❌ Error: No file selected"
        return redirect("/comparison?tab=appointment")

    try:
        # Get filename without saving to disk
        filename = secure_filename(file.filename)

        # Read Excel file directly from memory (no disk storage) WITHOUT headers
        # We'll find and set the header row in Step 1
        file.seek(0)  # Reset file pointer to beginning
        excel_data = pd.read_excel(file, sheet_name=None, engine="openpyxl", header=None)

        # Store raw data - we'll process headers in Step 1
        cleaned_data = {}
        for sheet_name, df in excel_data.items():
            # Convert all column names to strings (they'll be 0, 1, 2, etc. since header=None)
            df.columns = [str(i) for i in range(len(df.columns))]
            cleaned_data[sheet_name] = df

        # Process all sheets - format both Primary and Secondary insurance columns
        processed_sheets = {}
        zero_id_rows_by_sheet = {}  # Store zero ID rows per sheet (before duplicate removal)
        total_rows_processed = 0
        output_lines = []
        output_lines.append("=" * 70)
        output_lines.append(
            "PROCESSING APPOINTMENT REPORT - FORMATTING INSURANCE COLUMNS"
        )
        output_lines.append("=" * 70)
        output_lines.append("")

        for sheet_name, df in cleaned_data.items():
            output_lines.append(f"📋 Processing sheet: {sheet_name}")
            output_lines.append(
                f"   Original shape: {df.shape[0]} rows × {df.shape[1]} columns"
            )

            # Step 1: Find header row, set it as column names, then delete rows before header and last 7 rows
            output_lines.append("")
            output_lines.append(f"   🔧 Step 1: Row cleanup and header detection...")
            df_processed = df.copy()
            rows_before_cleanup = len(df_processed)
            
            # First, try to find the actual header row by looking for common column names
            # Common column names to look for: Office Name, Patient ID, Dental Primary Ins Carr, etc.
            header_row_index = None
            common_header_keywords = [
                "office name", "patient id", "pat id", "dental primary", 
                "dental secondary", "appt date", "appointment date", 
                "remark", "agent name", "time"
            ]
            
            # Check first 10 rows to find the header row
            for idx in range(min(10, len(df_processed))):
                row_values = []
                for col_idx in df_processed.columns:
                    val = df_processed.iloc[idx][col_idx]
                    row_values.append(str(val).lower().strip() if pd.notna(val) else "")
                # Count how many common keywords are found in this row
                matches = sum(1 for val in row_values if any(keyword in val for keyword in common_header_keywords))
                # If we find at least 3 matches, this is likely the header row
                if matches >= 3:
                    header_row_index = idx
                    output_lines.append(f"   🔍 Found header row at index {idx}")
                    break
            
            if header_row_index is not None and header_row_index >= 0:
                # Set the header row as column names
                new_columns = []
                for col_idx in df_processed.columns:
                    val = df_processed.iloc[header_row_index][col_idx]
                    col_name = str(val).strip() if pd.notna(val) and str(val).strip() != "" else f"Unnamed_{col_idx}"
                    new_columns.append(col_name)
                df_processed.columns = new_columns
                # Delete all rows up to and including the header row
                df_processed = df_processed.iloc[header_row_index + 1:].reset_index(drop=True)
                # Remove "Unnamed:" columns if any (but keep columns that are actually named "Unnamed_X")
                df_processed = df_processed.loc[:, ~df_processed.columns.str.contains("^Unnamed:", na=False, regex=True)]
                output_lines.append(f"   ✅ Set row {header_row_index + 1} as headers and deleted {header_row_index + 1} rows from top")
            elif header_row_index == 0:
                # Header is in the first row, set it as column names
                new_columns = []
                for col_idx in df_processed.columns:
                    val = df_processed.iloc[0][col_idx]
                    col_name = str(val).strip() if pd.notna(val) and str(val).strip() != "" else f"Unnamed_{col_idx}"
                    new_columns.append(col_name)
                df_processed.columns = new_columns
                # Delete the header row (first row)
                df_processed = df_processed.iloc[1:].reset_index(drop=True)
                # Remove "Unnamed:" columns if any
                df_processed = df_processed.loc[:, ~df_processed.columns.str.contains("^Unnamed:", na=False, regex=True)]
                output_lines.append(f"   ✅ Set first row as headers and deleted header row")
            else:
                # Couldn't find header row, try to delete first 3 rows but keep current structure
                if len(df_processed) > 3:
                    df_processed = df_processed.iloc[3:].reset_index(drop=True)
                    output_lines.append(f"   ⚠️  Could not identify header row, deleted first 3 rows by default")
                elif len(df_processed) > 0:
                    rows_deleted = len(df_processed)
                    df_processed = df_processed.iloc[rows_deleted:].reset_index(drop=True)
                    output_lines.append(f"   ⚠️  Deleted {rows_deleted} row(s) from top (less than 3 rows available)")
                else:
                    output_lines.append(f"   ⚠️  Sheet is empty, skipping row deletion")
            
            # Delete specified columns from output file
            columns_to_delete = [
                "Appointment ID",
                "Appointment Time",
                "Appointment Length (min)",
                "Appointment Status",
                "Operatory ID",
                "Operatory Name",
                "Chart#",
                "Patient Balance",
                "Medical Primary Ins Carr",
                "Medical Secondary Ins Carr",
                "Pref Provider",
                "Provider Name",
                "Procedure (Code  Th  Surf  Descr)",
                "Appointment Notes",
                "Created On",
                "Created By"
            ]
            
            columns_deleted = []
            columns_not_found = []
            
            for col_to_delete in columns_to_delete:
                found = False
                col_to_delete_normalized = re.sub(r'\s+', ' ', str(col_to_delete).strip().lower())
                
                # Try exact match first
                if col_to_delete in df_processed.columns:
                    df_processed = df_processed.drop(columns=[col_to_delete])
                    columns_deleted.append(col_to_delete)
                    found = True
                else:
                    # Try case-insensitive and flexible whitespace matching
                    for col in df_processed.columns:
                        col_normalized = re.sub(r'\s+', ' ', str(col).strip().lower())
                        if col_normalized == col_to_delete_normalized:
                            df_processed = df_processed.drop(columns=[col])
                            columns_deleted.append(col)
                            found = True
                            break
                
                if not found:
                    columns_not_found.append(col_to_delete)
            
            if columns_deleted:
                output_lines.append(f"   ✅ Deleted {len(columns_deleted)} column(s): {', '.join(columns_deleted[:5])}")
                if len(columns_deleted) > 5:
                    output_lines.append(f"      ... and {len(columns_deleted) - 5} more columns")
            if columns_not_found:
                output_lines.append(f"   ⚠️  Could not find {len(columns_not_found)} column(s) to delete: {', '.join(columns_not_found[:3])}")
                if len(columns_not_found) > 3:
                    output_lines.append(f"      ... and {len(columns_not_found) - 3} more columns")
            
            # Delete last 7 rows from bottom
            if len(df_processed) > 7:
                df_processed = df_processed.iloc[:-7].reset_index(drop=True)
                output_lines.append(f"   ✅ Deleted last 7 rows from bottom")
            elif len(df_processed) > 0:
                rows_deleted_from_bottom = len(df_processed)
                df_processed = df_processed.iloc[:0].reset_index(drop=True)
                output_lines.append(f"   ⚠️  Deleted {rows_deleted_from_bottom} row(s) from bottom (less than 7 rows available)")
            else:
                output_lines.append(f"   ℹ️  No rows available to delete from bottom")
            
            rows_after_cleanup = len(df_processed)
            output_lines.append(f"   Shape after cleanup: {rows_after_cleanup} rows × {df_processed.shape[1]} columns")
            
            # Show column names for verification
            if len(df_processed.columns) > 0:
                output_lines.append(f"   Column names: {', '.join(list(df_processed.columns[:10]))}")
                if len(df_processed.columns) > 10:
                    output_lines.append(f"   ... and {len(df_processed.columns) - 10} more columns")
            output_lines.append("")

            # Find the Patient ID column (case-insensitive search)
            patient_id_col = None
            for col in df_processed.columns:
                col_lower = col.lower().strip().replace(" ", "").replace("_", "")
                if ("patient" in col_lower or "pat" in col_lower) and "id" in col_lower:
                    patient_id_col = col
                    break

            # Find the insurance columns (case-insensitive search)
            primary_col = None
            secondary_col = None

            for col in df_processed.columns:
                col_lower = col.lower().strip()
                if (
                    "dental" in col_lower
                    and "primary" in col_lower
                    and "ins" in col_lower
                ):
                    primary_col = col
                elif (
                    "dental" in col_lower
                    and "secondary" in col_lower
                    and "ins" in col_lower
                ):
                    secondary_col = col

            # Step 2: Format insurance names
            output_lines.append("")
            output_lines.append(f"   📝 Step 2: Formatting insurance names...")
            formatted_primary = 0
            formatted_secondary = 0

            if primary_col:
                # Format primary insurance column
                df_processed[primary_col] = df_processed[primary_col].apply(
                    format_insurance_name
                )
                formatted_primary = df_processed[primary_col].notna().sum()
                output_lines.append(
                    f"   ✅ Formatted '{primary_col}' column ({formatted_primary} entries)"
                )
            else:
                output_lines.append(f"   ⚠️  'Dental Primary Ins Carr' column not found")

            if secondary_col:
                # Format secondary insurance column
                df_processed[secondary_col] = df_processed[secondary_col].apply(
                    format_insurance_name
                )
                formatted_secondary = df_processed[secondary_col].notna().sum()
                output_lines.append(
                    f"   ✅ Formatted '{secondary_col}' column ({formatted_secondary} entries)"
                )
            else:
                output_lines.append(
                    f"   ⚠️  'Dental Secondary Ins Carr' column not found"
                )

            # Extract zero ID rows BEFORE duplicate removal (to preserve duplicates in Zero ID sheet)
            if patient_id_col:
                def is_zero_id(val):
                    if pd.isna(val):
                        return False
                    try:
                        float_val = float(val)
                        return float_val == 0.0
                    except (ValueError, TypeError):
                        str_val = str(val).strip()
                        return str_val == "0" or str_val == "0.0" or str_val.lower() == "zero"
                
                zero_id_mask = df_processed[patient_id_col].apply(is_zero_id)
                zero_id_df = df_processed[zero_id_mask].copy()
                
                if len(zero_id_df) > 0:
                    # Store zero ID rows for this sheet (before duplicate removal)
                    zero_id_rows_by_sheet[sheet_name] = zero_id_df
                    # Remove zero ID rows from df_processed before duplicate removal
                    df_processed = df_processed[~zero_id_mask].copy()
                    output_lines.append(f"   📋 Extracted {len(zero_id_df)} row(s) with Patient ID = 0 (will be added to 'Zero ID' sheet without duplicate removal)")

            # Step 3: Remove duplicates based on Patient ID + Primary Insurance
            duplicates_removed = 0
            if patient_id_col and primary_col:
                output_lines.append("")
                output_lines.append(f"   🔍 Step 3: Checking for duplicate Patient IDs...")

                # Normalize Patient IDs for comparison
                df_processed["_normalized_patient_id"] = df_processed[
                    patient_id_col
                ].apply(normalize_patient_id)

                # Find duplicates
                rows_before = len(df_processed)

                # Group by normalized Patient ID and Primary Insurance
                # Keep only first occurrence when Patient ID + Primary Insurance are the same
                df_processed = df_processed.drop_duplicates(
                    subset=["_normalized_patient_id", primary_col], keep="first"
                )

                # Remove the temporary normalized column
                df_processed = df_processed.drop(columns=["_normalized_patient_id"])

                rows_after = len(df_processed)
                duplicates_removed = rows_before - rows_after

                if duplicates_removed > 0:
                    output_lines.append(
                        f"   ✅ Removed {duplicates_removed} duplicate record(s) with same Patient ID and same insurance"
                    )
                    output_lines.append(
                        f"   ℹ️  Kept records with same Patient ID but different insurance names"
                    )
                else:
                    output_lines.append(
                        f"   ℹ️  No duplicates found (or all duplicates have different insurance)"
                    )
            elif patient_id_col:
                output_lines.append(
                    f"   ⚠️  Cannot check duplicates: Primary insurance column not found"
                )
            else:
                output_lines.append(
                    f"   ⚠️  Cannot check duplicates: Patient ID column not found"
                )

            # Show sample data (first 5 rows)
            output_lines.append("")
            output_lines.append(f"   Sample of formatted data (first 5 rows):")
            sample_df = df_processed.head(5)
            output_lines.append(sample_df.to_string(index=False))

            processed_sheets[sheet_name] = df_processed
            total_rows_processed += len(df_processed)
            output_lines.append("")
            output_lines.append(
                f"   Final shape: {df_processed.shape[0]} rows × {df_processed.shape[1]} columns"
            )
            output_lines.append("")

        # Create "Zero ID" sheet with rows where Patient ID is 0 (using pre-extracted rows, duplicates preserved)
        output_lines.append("")
        output_lines.append("=" * 70)
        output_lines.append("CREATING 'ZERO ID' SHEET")
        output_lines.append("=" * 70)
        
        # Use pre-extracted zero ID rows (before duplicate removal, so duplicates are preserved)
        if zero_id_rows_by_sheet:
            zero_id_rows_list = []
            for sheet_name, zero_id_df in zero_id_rows_by_sheet.items():
                zero_id_rows_list.append(zero_id_df)
                output_lines.append(f"   Found {len(zero_id_df)} row(s) with Patient ID = 0 in sheet '{sheet_name}' (duplicates preserved)")
            
            # Combine all zero ID rows from all sheets
            zero_id_combined = pd.concat(zero_id_rows_list, ignore_index=True)
            processed_sheets["Zero ID"] = zero_id_combined
            output_lines.append(f"   ✅ Created 'Zero ID' sheet with {len(zero_id_combined)} total row(s) (duplicates NOT removed)")
        else:
            output_lines.append(f"   ℹ️  No rows with Patient ID = 0 found, skipping 'Zero ID' sheet creation")
        
        # Create "No Ins" sheet with rows where "Dental Primary Ins Carr" is blank, then remove from main sheets
        output_lines.append("")
        output_lines.append("=" * 70)
        output_lines.append("CREATING 'NO INS' SHEET")
        output_lines.append("=" * 70)
        
        no_ins_rows = []
        total_no_ins_rows = 0
        
        for sheet_name, df_sheet in processed_sheets.items():
            # Skip special sheets
            if sheet_name == "Zero ID" or sheet_name == "No Ins":
                continue
                
            # Find "Dental Primary Ins Carr" column in this sheet (case-insensitive)
            primary_ins_col = None
            for col in df_sheet.columns:
                col_lower = str(col).lower().strip()
                if (
                    "dental" in col_lower
                    and "primary" in col_lower
                    and "ins" in col_lower
                ):
                    primary_ins_col = col
                    break
            
            if primary_ins_col:
                # Filter rows where "Dental Primary Ins Carr" is blank/empty
                # Check for: NaN, empty string, whitespace-only strings
                def is_blank_insurance(val):
                    if pd.isna(val):
                        return True
                    str_val = str(val).strip()
                    return str_val == "" or str_val.lower() in ["none", "n/a", "na", "null"]
                
                no_ins_mask = df_sheet[primary_ins_col].apply(is_blank_insurance)
                no_ins_df = df_sheet[no_ins_mask].copy()
                
                if len(no_ins_df) > 0:
                    no_ins_rows.append(no_ins_df)
                    total_no_ins_rows += len(no_ins_df)
                    output_lines.append(f"   Found {len(no_ins_df)} row(s) with blank 'Dental Primary Ins Carr' in sheet '{sheet_name}'")
                    
                    # Remove no insurance rows from the main sheet
                    df_sheet_cleaned = df_sheet[~no_ins_mask].copy()
                    processed_sheets[sheet_name] = df_sheet_cleaned
                    rows_removed = len(df_sheet) - len(df_sheet_cleaned)
                    output_lines.append(f"   ✅ Removed {rows_removed} row(s) from sheet '{sheet_name}'")
            else:
                output_lines.append(f"   ⚠️  'Dental Primary Ins Carr' column not found in sheet '{sheet_name}', skipping")
        
        # Create "No Ins" sheet if we found any rows
        if no_ins_rows:
            # Combine all no insurance rows from all sheets
            no_ins_combined = pd.concat(no_ins_rows, ignore_index=True)
            processed_sheets["No Ins"] = no_ins_combined
            output_lines.append(f"   ✅ Created 'No Ins' sheet with {len(no_ins_combined)} total row(s)")
        else:
            output_lines.append(f"   ℹ️  No rows with blank 'Dental Primary Ins Carr' found, skipping 'No Ins' sheet creation")
        
        output_lines.append("")
        output_lines.append("=" * 70)
        output_lines.append("PROCESSING COMPLETE!")
        output_lines.append("=" * 70)

        # Update global variables
        appointment_report_data = processed_sheets
        appointment_report_filename = filename
        appointment_report_output = "\n".join(output_lines)

        # Count sheets processed
        sheets_count = len(processed_sheets)
        appointment_report_result = f"✅ Processing complete! Formatted insurance columns in {sheets_count} sheet(s). Total rows processed: {total_rows_processed}"

        return redirect("/comparison?tab=appointment")

    except Exception as e:
        appointment_report_result = f"❌ Error processing file: {str(e)}"
        appointment_report_output = f"Error: {str(e)}"
        return redirect("/comparison?tab=appointment")


@app.route("/download_appointment_report", methods=["POST"])
def download_appointment_report():
    global appointment_report_data, appointment_report_filename, appointment_report_result, appointment_report_output

    if not appointment_report_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"formatted_appointment_report_{timestamp}.xlsx"

    try:
        # Create a temporary file
        import tempfile

        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in appointment_report_data.items():
                    df_clean = df.copy()

                    # Format date columns to MM/DD/YYYY format (flexible column name search)
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        # Check for date columns
                        if "date" in col_lower or "time" in col_lower:
                            # Convert dates to MM/DD/YYYY format
                            def format_date(date_val):
                                if pd.isna(date_val) or date_val == "":
                                    return ""
                                try:
                                    # Convert to datetime if not already
                                    if isinstance(date_val, pd.Timestamp):
                                        date_obj = date_val
                                    elif isinstance(date_val, str):
                                        # Try to parse string date
                                        date_obj = pd.to_datetime(
                                            date_val, errors="coerce"
                                        )
                                        if pd.isna(date_obj):
                                            return str(
                                                date_val
                                            )  # Return original if can't parse
                                    else:
                                        date_obj = pd.to_datetime(
                                            date_val, errors="coerce"
                                        )
                                        if pd.isna(date_obj):
                                            return str(
                                                date_val
                                            )  # Return original if can't parse

                                    # Format as MM/DD/YYYY
                                    return date_obj.strftime("%m/%d/%Y")
                                except (ValueError, TypeError, AttributeError):
                                    # If parsing fails, return as-is
                                    return str(date_val)

                            # Format dates and convert column to string type to prevent Excel auto-formatting
                            df_clean[col] = df_clean[col].apply(format_date)
                            df_clean[col] = df_clean[col].astype(str)

                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Set date column format to text in Excel to preserve MM/DD/YYYY format
                    ws = writer.sheets[sheet_name]
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        if "date" in col_lower or "time" in col_lower:
                            # Find the column index
                            col_idx = None
                            for idx, col_name in enumerate(df_clean.columns, 1):
                                if col_name == col:
                                    col_idx = idx
                                    break

                            if col_idx:
                                # Set all cells in this column to text format
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=col_idx)
                                    if cell.value:
                                        cell.number_format = "@"  # Text format

            # Clear data after successful download
            appointment_report_data = None
            appointment_report_filename = None
            appointment_report_result = None
            appointment_report_output = ""

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            # Clean up temporary file
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reset_appointment_report", methods=["POST"])
def reset_appointment_report():
    global appointment_report_data, appointment_report_filename, appointment_report_result, appointment_report_output
    # Explicitly do NOT touch other tool variables

    try:
        # Reset ONLY appointment report tool variables
        appointment_report_data = None
        appointment_report_filename = None
        appointment_report_result = "🔄 Appointment report formatting tool reset successfully! All files and data have been cleared."
        appointment_report_output = ""

        return redirect("/comparison?tab=appointment")

    except Exception as e:
        appointment_report_result = (
            f"❌ Error resetting appointment report formatting tool: {str(e)}"
        )
        return redirect("/comparison?tab=appointment")


@app.route("/upload_smart_assist", methods=["POST"])
def upload_smart_assist():
    global smart_assist_data, smart_assist_filename, smart_assist_result, smart_assist_output

    if "file" not in request.files:
        smart_assist_result = "❌ Error: No file provided"
        return redirect("/comparison?tab=smartassist")

    file = request.files["file"]
    if file.filename == "":
        smart_assist_result = "❌ Error: No file selected"
        return redirect("/comparison?tab=smartassist")

    try:
        filename = secure_filename(file.filename)
        file.seek(0)

        # Read raw Excel data without headers first to analyze structure
        from openpyxl import load_workbook

        wb = load_workbook(file)

        processed_sheets = {}
        output_lines = []
        output_lines.append("=" * 70)
        output_lines.append("PROCESSING SMART ASSIST REPORT - ADVANCED FORMATTING")
        output_lines.append("=" * 70)
        output_lines.append("")

        for sheet_name in wb.sheetnames:
            output_lines.append(f"📋 Processing sheet: {sheet_name}")
            ws = wb[sheet_name]

            # Step 1: Read all data as raw values
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(row)

            output_lines.append(f"   Original rows: {len(all_rows)}")

            # Step 2: Find header rows (rows that look like headers)
            header_rows = []
            header_indices = []

            for idx, row in enumerate(all_rows):
                # Check if row looks like a header (has text values, not mostly empty)
                if row and any(row):
                    non_empty_count = sum(
                        1
                        for cell in row
                        if cell is not None and str(cell).strip() != ""
                    )
                    # If more than 30% of cells are filled, could be a header or data row
                    if non_empty_count > 0:
                        # Check if it looks like a header (contains common header keywords)
                        row_text = " ".join(
                            [str(cell).lower() for cell in row if cell is not None]
                        )
                        # Look for header patterns
                        if any(
                            keyword in row_text
                            for keyword in [
                                "patient",
                                "name",
                                "date",
                                "time",
                                "insurance",
                                "carrier",
                                "id",
                                "appt",
                            ]
                        ):
                            # Check if next few rows are data (to confirm this is a header)
                            is_likely_header = False

                            # First occurrence or significant gap from last header
                            if not header_indices or (idx - header_indices[-1]) > 5:
                                is_likely_header = True

                            if is_likely_header:
                                header_rows.append(row)
                                header_indices.append(idx)

            output_lines.append(
                f"   Found {len(header_rows)} potential header row(s) at positions: {header_indices}"
            )

            # Step 3: Use the first valid header row
            if not header_rows:
                output_lines.append(
                    f"   ⚠️  No valid header row found, using first row as header"
                )
                header_row = all_rows[0] if all_rows else []
                data_start_idx = 1
            else:
                header_row = header_rows[0]
                data_start_idx = header_indices[0] + 1

            # Clean header row - remove None and empty strings
            headers = []
            for cell in header_row:
                if cell is not None and str(cell).strip() != "":
                    headers.append(str(cell).strip())
                else:
                    headers.append(f"Column_{len(headers)}")

            output_lines.append(f"   Headers identified: {len(headers)} columns")

            # Step 4: Collect all data rows (skip blank rows and intermediate header rows)
            data_rows = []
            blank_rows_removed = 0
            duplicate_headers_removed = 0
            summary_rows_removed = 0

            for idx in range(data_start_idx, len(all_rows)):
                row = all_rows[idx]

                # Skip completely blank rows
                if not row or all(
                    cell is None or str(cell).strip() == "" for cell in row
                ):
                    blank_rows_removed += 1
                    continue

                # Skip rows that look like duplicate headers (middle section headers)
                row_text = " ".join(
                    [str(cell).lower() for cell in row if cell is not None]
                )
                is_duplicate_header = False

                # Check if this row matches any header pattern
                if idx in header_indices[1:]:  # Skip first header
                    is_duplicate_header = True
                    duplicate_headers_removed += 1
                    continue

                # Additional check: if row closely matches header row, skip it
                if len(row) == len(header_row):
                    matches = sum(
                        1 for i, cell in enumerate(row) if cell == header_row[i]
                    )
                    if matches > len(header_row) * 0.5:  # More than 50% match
                        is_duplicate_header = True
                        duplicate_headers_removed += 1
                        continue

                # Skip summary/total rows
                # Check first few cells for total/summary indicators
                first_cells_text = " ".join(
                    [str(cell).lower().strip() for cell in row[:3] if cell is not None]
                )

                if any(
                    pattern in first_cells_text
                    for pattern in [
                        "total appointments for office",
                        "office :",
                        "office:",
                        "grand total",
                    ]
                ):
                    summary_rows_removed += 1
                    continue

                # Valid data row - add it
                data_rows.append(row)

            output_lines.append(f"   ✅ Removed {blank_rows_removed} blank row(s)")
            output_lines.append(
                f"   ✅ Removed {duplicate_headers_removed} duplicate header row(s)"
            )
            output_lines.append(
                f"   ✅ Removed {summary_rows_removed} summary/total row(s)"
            )
            output_lines.append(f"   Data rows collected: {len(data_rows)}")

            # Step 5: Create DataFrame with consolidated data
            if not data_rows:
                output_lines.append(f"   ⚠️  No data rows found in sheet")
                continue

            # Ensure all rows have same length as headers
            normalized_rows = []
            for row in data_rows:
                if len(row) < len(headers):
                    # Pad with None
                    row = list(row) + [None] * (len(headers) - len(row))
                elif len(row) > len(headers):
                    # Trim
                    row = row[: len(headers)]
                normalized_rows.append(row)

            df = pd.DataFrame(normalized_rows, columns=headers)

            # Step 6: Remove "Unnamed:" columns
            df.columns = df.columns.astype(str)
            df = df.loc[:, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)]
            df = df.loc[:, ~df.columns.str.startswith("Column_")]

            # Step 7: Remove any remaining rows that are all NaN
            df = df.dropna(how="all")

            output_lines.append(
                f"   Cleaned shape: {df.shape[0]} rows × {df.shape[1]} columns"
            )

            # Step 8: Build standardized output with requested columns
            standard_columns = [
                "Office Name",
                "Appointment Date",
                "Patient ID",
                "Patient Name",
                "Chart#",
                "Dental Primary Ins Carr",
                "Dental Secondary Ins Carr",
                "Provider Name",
                "Received date",
                "Source",
                "Type",
                "Status Code",
                "Comment",
                "Group Number",
                "Category",
                "Agent Name",
                "Work Date",
                "Remark",
                "Priority Status",
                "QC Agent",
                "QC Status",
                "QC Comments",
                "QC Date",
            ]
            output_lines.append(
                f"   Standardizing to {len(standard_columns)} predefined column(s)"
            )

            # Helper to fuzzy-find a column by keywords
            def find_col(keywords):
                for col in df.columns:
                    col_norm = col.lower().strip().replace(" ", "").replace("_", "")
                    if all(k in col_norm for k in keywords):
                        return col
                return None

            office_col = find_col(["office"])
            appt_date_col = find_col(["appt", "date"]) or find_col(
                ["appointment", "date"]
            )
            patid_col = find_col(["patid"]) or find_col(["patient", "id"])
            last_name_col = find_col(["patient", "last"]) or find_col(["last", "name"])
            first_name_col = find_col(["patient", "first"]) or find_col(
                ["first", "name"]
            )
            chart_col = find_col(["chart"])
            provider_col = find_col(["provider"]) or find_col(["dr"])
            received_date_col = find_col(["received", "date"])
            source_col = find_col(["source"])
            type_col = find_col(["type"])
            status_code_col = find_col(["status", "code"]) or find_col(["code"])
            comment_col = find_col(["comment"]) or find_col(["notes"])
            group_number_col = find_col(["group", "number"]) or find_col(["group#"])
            category_col = find_col(["category"])
            agent_name_col = find_col(["agent", "name"]) or find_col(["agent"])
            work_date_col = find_col(["work", "date"])
            remark_col = find_col(["remark"])
            priority_status_col = find_col(["priority", "status"]) or find_col(
                ["priority"]
            )
            qc_agent_col = find_col(["qc", "agent"])
            qc_status_col = find_col(["qc", "status"])
            qc_comments_col = find_col(["qc", "comment"]) or find_col(["qc", "notes"])
            qc_date_col = find_col(["qc", "date"])

            # Find insurance columns
            primary_ins_col = find_col(["dental", "primary", "ins"]) or find_col(
                ["primary", "insurance"]
            )
            secondary_ins_col = find_col(["dental", "secondary", "ins"]) or find_col(
                ["secondary", "insurance"]
            )

            standardized = pd.DataFrame(index=df.index, columns=standard_columns)
            standardized["Office Name"] = df[office_col] if office_col else ""
            standardized["Appointment Date"] = (
                df[appt_date_col] if appt_date_col else ""
            )
            standardized["Patient ID"] = df[patid_col] if patid_col else ""
            if last_name_col or first_name_col:
                ln = df[last_name_col] if last_name_col else ""
                fn = df[first_name_col] if first_name_col else ""
                standardized["Patient Name"] = (
                    ln.fillna("").astype(str).str.strip()
                    + ", "
                    + fn.fillna("").astype(str).str.strip()
                )
                standardized["Patient Name"] = (
                    standardized["Patient Name"]
                    .str.strip(", ")
                    .replace(", $", "", regex=True)
                )
            else:
                standardized["Patient Name"] = ""
            standardized["Chart#"] = df[chart_col] if chart_col else ""
            standardized["Dental Primary Ins Carr"] = (
                df[primary_ins_col] if primary_ins_col else ""
            )
            standardized["Dental Secondary Ins Carr"] = (
                df[secondary_ins_col] if secondary_ins_col else ""
            )
            standardized["Provider Name"] = df[provider_col] if provider_col else ""
            standardized["Received date"] = (
                df[received_date_col] if received_date_col else ""
            )
            standardized["Source"] = df[source_col] if source_col else ""
            standardized["Type"] = df[type_col] if type_col else ""
            standardized["Status Code"] = df[status_code_col] if status_code_col else ""

            # Clean Comment column - remove special characters
            if comment_col:

                def clean_comment(text):
                    if pd.isna(text) or text == "":
                        return ""
                    text_str = str(text)
                    # Remove non-printable characters and keep only standard ASCII and common punctuation
                    import re

                    # Keep letters, numbers, spaces, and common punctuation
                    cleaned = re.sub(r"[^\x20-\x7E\n\r\t]", "", text_str)
                    return cleaned.strip()

                standardized["Comment"] = df[comment_col].apply(clean_comment)
            else:
                standardized["Comment"] = ""

            standardized["Group Number"] = (
                df[group_number_col] if group_number_col else ""
            )
            standardized["Category"] = df[category_col] if category_col else ""
            standardized["Agent Name"] = df[agent_name_col] if agent_name_col else ""
            standardized["Work Date"] = df[work_date_col] if work_date_col else ""
            standardized["Remark"] = df[remark_col] if remark_col else ""
            standardized["Priority Status"] = (
                df[priority_status_col] if priority_status_col else ""
            )
            standardized["QC Agent"] = df[qc_agent_col] if qc_agent_col else ""
            standardized["QC Status"] = df[qc_status_col] if qc_status_col else ""
            standardized["QC Comments"] = df[qc_comments_col] if qc_comments_col else ""
            standardized["QC Date"] = df[qc_date_col] if qc_date_col else ""

            # Find Eligibility column and set Remark based on its value
            eligibility_col = find_col(["eligibility"])
            print(f"DEBUG - Looking for Eligibility column. Found: {eligibility_col}")
            print(f"DEBUG - Available columns in df: {list(df.columns)}")

            # Create a temporary status column for filtering
            temp_status = pd.Series("", index=standardized.index)

            if eligibility_col:
                output_lines.append(f"   Found Eligibility column: '{eligibility_col}'")
                workable_count = 0
                completed_count = 0
                for idx in standardized.index:
                    eligibility_val = (
                        df.loc[idx, eligibility_col]
                        if eligibility_col and idx in df.index
                        else None
                    )
                    if pd.notna(eligibility_val):
                        eligibility_str = str(eligibility_val).strip()
                        print(
                            f"DEBUG - Row {idx}: Eligibility value = '{eligibility_str}'"
                        )
                        # Check for X marks (✗, x, û - encoded version)
                        if (
                            "✗" in eligibility_str
                            or "x" in eligibility_str.lower()
                            or "û" in eligibility_str
                        ):
                            standardized.at[idx, "Remark"] = "Workable"
                            temp_status.at[idx] = "Workable"
                            workable_count += 1
                        # Check for check marks (✓, √, ü - encoded version)
                        elif (
                            "✓" in eligibility_str
                            or "√" in eligibility_str
                            or "ü" in eligibility_str
                            or "check" in eligibility_str.lower()
                        ):
                            temp_status.at[idx] = "Completed"
                            completed_count += 1
                print(
                    f"DEBUG - Remark set: {workable_count} Workable, {completed_count} Completed"
                )
                output_lines.append(
                    f"   ✅ Remark column populated: {workable_count} Workable rows"
                )
            else:
                output_lines.append(
                    f"   ⚠️  Eligibility column not found - Remark column will be empty"
                )

            # Filter: Remove rows where temp_status is blank or "Completed" (keep only Workable)
            rows_before_filter = len(standardized)
            keep_mask = temp_status == "Workable"
            standardized = standardized[keep_mask]
            rows_removed = rows_before_filter - len(standardized)

            if rows_removed > 0:
                output_lines.append(
                    f"   🗑️  Removed {rows_removed} row(s) with blank or 'Completed' eligibility"
                )
                output_lines.append(
                    f"   ✅ Kept {len(standardized)} row(s) with 'Workable' in Remark"
                )

            df = standardized

            # Report which source columns were mapped
            mapped_sources = []
            for label, src in [
                ("Office Name", office_col),
                ("Appointment Date", appt_date_col),
                ("Patient ID", patid_col),
                ("Patient Name", (last_name_col, first_name_col)),
                ("Chart#", chart_col),
                ("Dental Primary Ins Carr", primary_ins_col),
                ("Dental Secondary Ins Carr", secondary_ins_col),
                ("Provider Name", provider_col),
                ("Received date", received_date_col),
                ("Source", source_col),
                ("Type", type_col),
                ("Status Code", status_code_col),
                ("Comment", comment_col),
                ("Group Number", group_number_col),
                ("Category", category_col),
                ("Agent Name", agent_name_col),
                ("Work Date", work_date_col),
                ("Remark", remark_col),
                ("Priority Status", priority_status_col),
                ("QC Agent", qc_agent_col),
                ("QC Status", qc_status_col),
                ("QC Comments", qc_comments_col),
                ("QC Date", qc_date_col),
            ]:
                if isinstance(src, tuple):
                    if any(s for s in src):
                        mapped_sources.append(label)
                elif src:
                    mapped_sources.append(label)
            output_lines.append(
                f"   ✅ Mapped {len(mapped_sources)}/{len(standard_columns)} column(s) from source data"
            )

            df = standardized

            # Step 9: Show sample of standardized data
            output_lines.append("")
            sample_cols = df.columns.tolist()[:8]
            if sample_cols:
                sample_df = df[sample_cols].head(5)
                output_lines.append("   Sample (first 5 rows):")
                output_lines.append(sample_df.to_string(index=False))

            output_lines.append("")
            output_lines.append(
                f"   ✅ Final standardized shape: {df.shape[0]} rows × {df.shape[1]} columns"
            )
            output_lines.append(
                f"   Summary: Removed {blank_rows_removed + duplicate_headers_removed + summary_rows_removed} total unwanted rows"
            )
            output_lines.append("")

            processed_sheets[sheet_name] = df

        output_lines.append("=" * 70)
        output_lines.append("PROCESSING COMPLETE! FILE READY FOR DOWNLOAD")
        output_lines.append("=" * 70)

        smart_assist_data = processed_sheets
        smart_assist_filename = filename
        smart_assist_output = "\n".join(output_lines)

        total_rows = sum(len(df) for df in processed_sheets.values())
        sheets_count = len(processed_sheets)
        smart_assist_result = f"✅ Processing complete! Formatted {sheets_count} sheet(s) with {total_rows} total rows. All sections combined, blank rows removed, and headers consolidated."

        return redirect("/comparison?tab=smartassist")

    except Exception as e:
        import traceback

        error_details = traceback.format_exc()
        smart_assist_result = f"❌ Error processing file: {str(e)}"
        smart_assist_output = f"Error: {str(e)}\n\nDetails:\n{error_details}"
        return redirect("/comparison?tab=smartassist")


@app.route("/download_smart_assist", methods=["POST"])
def download_smart_assist():
    global smart_assist_data, smart_assist_filename, smart_assist_result, smart_assist_output

    if not smart_assist_data:
        return jsonify({"error": "No data to download"}), 400

    filename = request.form.get("filename", "").strip()
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"formatted_smart_assist_{timestamp}.xlsx"

    try:
        import tempfile

        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx")

        try:
            with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
                for sheet_name, df in smart_assist_data.items():
                    df_clean = df.copy()

                    # Format date columns
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        if "date" in col_lower or "time" in col_lower:

                            def format_date(date_val):
                                if pd.isna(date_val) or date_val == "":
                                    return ""
                                try:
                                    if isinstance(date_val, pd.Timestamp):
                                        date_obj = date_val
                                    elif isinstance(date_val, str):
                                        date_obj = pd.to_datetime(
                                            date_val, errors="coerce"
                                        )
                                        if pd.isna(date_obj):
                                            return str(date_val)
                                    else:
                                        date_obj = pd.to_datetime(
                                            date_val, errors="coerce"
                                        )
                                        if pd.isna(date_obj):
                                            return str(date_val)
                                    return date_obj.strftime("%m/%d/%Y")
                                except (ValueError, TypeError, AttributeError):
                                    return str(date_val)

                            df_clean[col] = df_clean[col].apply(format_date)
                            df_clean[col] = df_clean[col].astype(str)

                    df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Set date column format to text
                    ws = writer.sheets[sheet_name]
                    for col in df_clean.columns:
                        col_lower = (
                            col.lower().strip().replace(" ", "").replace("_", "")
                        )
                        if "date" in col_lower or "time" in col_lower:
                            col_idx = None
                            for idx, col_name in enumerate(df_clean.columns, 1):
                                if col_name == col:
                                    col_idx = idx
                                    break

                            if col_idx:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=col_idx)
                                    if cell.value:
                                        cell.number_format = "@"

            smart_assist_data = None
            smart_assist_filename = None
            smart_assist_result = None
            smart_assist_output = ""

            return send_file(temp_path, as_attachment=True, download_name=filename)

        finally:
            os.close(temp_fd)
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reset_smart_assist", methods=["POST"])
def reset_smart_assist():
    global smart_assist_data, smart_assist_filename, smart_assist_result, smart_assist_output

    try:
        smart_assist_data = None
        smart_assist_filename = None
        smart_assist_result = "🔄 Smart assist report formatting tool reset successfully! All files and data have been cleared."
        smart_assist_output = ""

        return redirect("/comparison?tab=smartassist")

    except Exception as e:
        smart_assist_result = (
            f"❌ Error resetting smart assist report formatting tool: {str(e)}"
        )
        return redirect("/comparison?tab=smartassist")


@app.route("/upload_consolidate", methods=["POST"])
def upload_consolidate():
    global consolidate_master_data, consolidate_daily_data, consolidate_master_filename, consolidate_daily_filename, consolidate_result, consolidate_output

    try:
        consolidate_output = ""
        output_lines = []

        # Check for both files
        if "master_file" not in request.files or "daily_file" not in request.files:
            consolidate_result = "❌ Error: Both Master Consolidate file and Daily Consolidated file are required."
            return redirect("/comparison?tab=consolidate")

        master_file = request.files["master_file"]
        daily_file = request.files["daily_file"]

        if master_file.filename == "" or daily_file.filename == "":
            consolidate_result = "❌ Error: Both files must be selected."
            return redirect("/comparison?tab=consolidate")

        # Save and read master file
        master_filename = secure_filename(master_file.filename)
        master_filepath = os.path.join("/tmp", master_filename)
        master_file.save(master_filepath)

        # Save and read daily file
        daily_filename = secure_filename(daily_file.filename)
        daily_filepath = os.path.join("/tmp", daily_filename)
        daily_file.save(daily_filepath)

        # Read Excel files (all sheets from master to preserve workbook; specific sheet from daily)
        output_lines.append("=" * 80)
        output_lines.append("CONSOLIDATE REPORT - FILE PROCESSING")
        output_lines.append("=" * 80)
        output_lines.append(f"\nMaster File: {master_filename}")
        output_lines.append(f"Daily File: {daily_filename}")
        output_lines.append(
            f"Processing Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        # Load master workbook (all sheets)
        master_xls = pd.ExcelFile(master_filepath)
        consolidate_master_data = {}
        for sheet in master_xls.sheet_names:
            consolidate_master_data[sheet] = pd.read_excel(
                master_filepath, sheet_name=sheet
            )

        # Ensure we have a 'consolidated' sheet in master (create empty if missing)
        master_has_consolidated = "consolidated" in consolidate_master_data
        if not master_has_consolidated:
            consolidate_master_data["consolidated"] = pd.DataFrame()

        # Load only the required sheet from daily
        try:
            daily_all_agent_df = pd.read_excel(
                daily_filepath, sheet_name="All Agent Data"
            )
        except Exception as e:
            consolidate_result = f"❌ Error: Could not find/read sheet 'All Agent Data' in Daily Consolidated File. Details: {str(e)}"
            return redirect("/comparison?tab=consolidate")

        consolidate_master_filename = master_filename
        consolidate_daily_filename = daily_filename

        output_lines.append(f"\n✅ Master file loaded successfully")
        output_lines.append(f"   Sheets: {', '.join(consolidate_master_data.keys())}")
        output_lines.append(
            f"   Consolidated sheet present: {'Yes' if master_has_consolidated else 'No (created)'}"
        )

        output_lines.append(f"\n✅ Daily file loaded successfully")
        output_lines.append(f"   Using sheet: All Agent Data")
        output_lines.append(f"   Rows in daily sheet: {len(daily_all_agent_df)}")

        # Normalize Appointment Date in daily sheet to MM/DD/YYYY if present
        def normalize_appointment_date(df):
            col = None
            # Exact name first, then case-insensitive search fallback
            if "Appointment Date" in df.columns:
                col = "Appointment Date"
            else:
                for c in df.columns:
                    if str(c).strip().lower() == "appointment date".lower():
                        col = c
                        break
            if col is not None:
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime(
                        "%m/%d/%Y"
                    )
                except Exception:
                    # Best-effort; leave values as-is on failure
                    pass
            return df

        daily_all_agent_df = normalize_appointment_date(daily_all_agent_df)

        # Append daily data to master's 'consolidated' sheet
        master_consolidated_df = consolidate_master_data.get(
            "consolidated", pd.DataFrame()
        )

        # Align columns to union of both DataFrames
        combined_columns = list(
            {
                *map(str, master_consolidated_df.columns),
                *map(str, daily_all_agent_df.columns),
            }
        )
        master_consolidated_aligned = master_consolidated_df.reindex(
            columns=combined_columns
        )
        daily_all_agent_aligned = daily_all_agent_df.reindex(columns=combined_columns)

        appended_df = pd.concat(
            [master_consolidated_aligned, daily_all_agent_aligned], ignore_index=True
        )

        # Identify duplicates by 'Patient ID' across the combined data
        pid_col = None
        if "Patient ID" in appended_df.columns:
            pid_col = "Patient ID"
        else:
            # Fallbacks commonly used in this project
            for alt in ["PATID", "Pat ID", "PatientID", "patient id", "patid"]:
                if alt in appended_df.columns:
                    pid_col = alt
                    break

        duplicates_df = pd.DataFrame()
        if pid_col is not None:
            dup_mask = appended_df.duplicated(subset=[pid_col], keep=False)
            duplicates_df = appended_df[dup_mask].copy()
            # Keep only first occurrence in consolidated sheet
            consolidated_unique_df = appended_df.drop_duplicates(
                subset=[pid_col], keep="first"
            ).copy()
        else:
            # If no patient id column found, proceed without duplicate handling
            consolidated_unique_df = appended_df.copy()
            output_lines.append(
                "\n⚠️  'Patient ID' column not found. Skipping duplicate detection."
            )

        # Update in-memory workbook
        consolidate_master_data["consolidated"] = consolidated_unique_df
        if not duplicates_df.empty:
            consolidate_master_data["Duplicate"] = duplicates_df
            output_lines.append(
                f"\n🧬 Duplicates found on '{pid_col}': {len(duplicates_df)} rows (written to 'Duplicate' sheet)"
            )
        else:
            output_lines.append("\n✅ No duplicates detected by 'Patient ID'.")

        output_lines.append("\n" + "=" * 80)
        output_lines.append("✅ CONSOLIDATION COMPLETE")
        output_lines.append("=" * 80)
        output_lines.append(
            f"Rows in master 'consolidated' before: {len(master_consolidated_df)}"
        )
        output_lines.append(f"Rows appended from daily: {len(daily_all_agent_df)}")
        output_lines.append(
            f"Rows in 'consolidated' after: {len(consolidate_master_data['consolidated'])}"
        )
        output_lines.append("\nReady to download consolidated file.")

        consolidate_output = "\n".join(output_lines)
        consolidate_result = (
            f"✅ Consolidation complete successfully!\n\n{consolidate_output}"
        )

        return redirect("/comparison?tab=consolidate")

    except Exception as e:
        consolidate_result = f"❌ Error processing consolidation: {str(e)}"
        consolidate_output = f"Error: {str(e)}"
        return redirect("/comparison?tab=consolidate")


@app.route("/download_consolidate", methods=["POST"])
def download_consolidate():
    global consolidate_master_data, consolidate_master_filename

    try:
        if not consolidate_master_data:
            return "No consolidate data available", 400

        filename = request.form.get("filename", "consolidated_report.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        # Create Excel file with all sheets
        output_path = os.path.join("/tmp", filename)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in consolidate_master_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        return f"Error downloading file: {str(e)}", 500


@app.route("/download_consolidate_consolidated_only", methods=["POST"])
def download_consolidate_consolidated_only():
    """Download only the merged 'consolidated' sheet."""
    global consolidate_master_data

    try:
        if not consolidate_master_data or "consolidated" not in consolidate_master_data:
            return "No consolidated data available", 400

        filename = request.form.get("filename", "consolidated_only.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        output_path = os.path.join("/tmp", filename)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            consolidate_master_data["consolidated"].to_excel(
                writer, sheet_name="consolidated", index=False
            )

        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        return f"Error downloading consolidated sheet: {str(e)}", 500


@app.route("/reset_consolidate", methods=["POST"])
def reset_consolidate():
    global consolidate_master_data, consolidate_daily_data, consolidate_master_filename, consolidate_daily_filename, consolidate_result, consolidate_output

    try:
        consolidate_master_data = None
        consolidate_daily_data = None
        consolidate_master_filename = None
        consolidate_daily_filename = None
        consolidate_result = "🔄 Consolidate report tool reset successfully! All files and data have been cleared."
        consolidate_output = ""

        return redirect("/comparison?tab=consolidate")

    except Exception as e:
        consolidate_result = f"❌ Error resetting consolidate report tool: {str(e)}"
        return redirect("/comparison?tab=consolidate")


@app.route("/upload_reallocation", methods=["POST"])
def upload_reallocation():
    global reallocation_consolidate_data, reallocation_blank_data, reallocation_consolidate_filename, reallocation_blank_filename, reallocation_result, reallocation_output, reallocation_merged_data, reallocation_available_remarks, reallocation_selected_remarks, reallocation_available_agents, reallocation_selected_agents

    try:
        reallocation_output = ""
        reallocation_merged_data = None
        output_lines = []

        # Read selections (if any)
        selected_remarks = request.form.getlist("remarks_filter")
        selected_agents = request.form.getlist("agent_filter")
        reallocation_selected_remarks = selected_remarks
        reallocation_selected_agents = selected_agents

        # Files may be provided in steps; load any provided, keep previous if not
        consolidate_file = request.files.get("consolidate_file")
        blank_file = request.files.get("blank_file")

        # Save and read consolidate file if provided
        if consolidate_file and consolidate_file.filename:
            consolidate_filename = secure_filename(consolidate_file.filename)
            consolidate_filepath = os.path.join("/tmp", consolidate_filename)
            consolidate_file.save(consolidate_filepath)
            # Load consolidate file
            consolidate_xls = pd.ExcelFile(consolidate_filepath)
            reallocation_consolidate_data = {}
            for sheet in consolidate_xls.sheet_names:
                reallocation_consolidate_data[sheet] = pd.read_excel(
                    consolidate_filepath, sheet_name=sheet
                )
            reallocation_consolidate_filename = consolidate_filename

        # Save and read blank allocation file if provided
        if blank_file and blank_file.filename:
            blank_filename = secure_filename(blank_file.filename)
            blank_filepath = os.path.join("/tmp", blank_filename)
            blank_file.save(blank_filepath)
            blank_xls = pd.ExcelFile(blank_filepath)
            reallocation_blank_data = {}
            for sheet in blank_xls.sheet_names:
                reallocation_blank_data[sheet] = pd.read_excel(
                    blank_filepath, sheet_name=sheet
                )
            reallocation_blank_filename = blank_filename
        # Log processing info
        output_lines.append("=" * 80)
        output_lines.append("REALLOCATION DATA GENERATION - FILE PROCESSING")
        output_lines.append("=" * 80)
        output_lines.append(
            f"\nConsolidate File: {reallocation_consolidate_filename or 'Not provided'}"
        )
        output_lines.append(
            f"Blank Allocation File: {reallocation_blank_filename or 'Not provided'}"
        )
        output_lines.append(
            f"Processing Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        if reallocation_consolidate_data:
            output_lines.append(f"\n✅ Consolidate file loaded successfully")
            output_lines.append(
                f"   Sheets: {', '.join(reallocation_consolidate_data.keys())}"
            )
            output_lines.append(
                f"   Total rows: {sum(len(df) for df in reallocation_consolidate_data.values())}"
            )
        if reallocation_blank_data:
            output_lines.append(f"\n✅ Blank allocation file loaded successfully")
            output_lines.append(
                f"   Sheets: {', '.join(reallocation_blank_data.keys())}"
            )
            output_lines.append(
                f"   Total rows: {sum(len(df) for df in reallocation_blank_data.values())}"
            )

        # STEP 1: Align remarks in consolidate file
        output_lines.append("\n" + "=" * 80)
        output_lines.append("STEP 1: ALIGNING REMARKS")
        output_lines.append("=" * 80)

        total_remarks_aligned = 0
        for sheet_name, df in (reallocation_consolidate_data or {}).items():
            if "Remark" in df.columns:
                # Apply remark alignment
                original_remarks = df["Remark"].copy()
                df["Remark"] = df["Remark"].apply(align_remark)

                # Count how many were changed
                changed_count = (original_remarks != df["Remark"]).sum()
                total_remarks_aligned += changed_count

                reallocation_consolidate_data[sheet_name] = df

                if changed_count > 0:
                    output_lines.append(f"\n📄 Sheet '{sheet_name}':")
                    output_lines.append(f"   ✓ Remarks aligned: {changed_count} rows")
            else:
                output_lines.append(
                    f"\n⚠️  Sheet '{sheet_name}': 'Remark' column not found"
                )

        output_lines.append(
            f"\n✅ Total remarks aligned across all sheets: {total_remarks_aligned}"
        )

        # Build available remark/agent lists from All Agent Data (if present)
        reallocation_available_remarks = []
        reallocation_available_agents = []
        if (
            reallocation_consolidate_data
            and "All Agent Data" in reallocation_consolidate_data
        ):
            df_all = reallocation_consolidate_data["All Agent Data"]
            if "Remark" in df_all.columns:
                remarks_series = df_all["Remark"].dropna()
                reallocation_available_remarks = sorted(
                    pd.Series(remarks_series.astype(str).str.strip().unique()).tolist()
                )
                output_lines.append(
                    f"\n📚 Available remarks detected: {len(reallocation_available_remarks)}"
                )
            if "Agent Name" in df_all.columns:
                agents_series = df_all["Agent Name"].dropna()
                reallocation_available_agents = sorted(
                    pd.Series(agents_series.astype(str).str.strip().unique()).tolist()
                )
                output_lines.append(
                    f"📚 Available agents detected: {len(reallocation_available_agents)}"
                )

        # If no selections and blank file not yet loaded, prompt user and stop early
        if (not selected_remarks and not selected_agents) or (
            not reallocation_blank_data
        ):
            reallocation_output = "\n".join(output_lines)
            if not selected_remarks and not selected_agents:
                reallocation_result = "ℹ️ Data loaded. Please select remarks and/or agents, then click Generate."
            elif not reallocation_blank_data:
                reallocation_result = (
                    "ℹ️ Consolidate loaded. Please upload the Blank Allocation file."
                )
            return redirect("/comparison?tab=reallocation")

        # STEP 2: Merge consolidate into blank allocation file
        output_lines.append("\n" + "=" * 80)
        output_lines.append("STEP 2: MERGING INTO BLANK ALLOCATION FILE")
        output_lines.append("=" * 80)

        merged_data = {}

        # Get the "All Agent Data" sheet from consolidate file
        if "All Agent Data" not in reallocation_consolidate_data:
            raise Exception(
                "'All Agent Data' sheet not found in Current Consolidate File"
            )

        cons_df = reallocation_consolidate_data["All Agent Data"]
        remark_filtered = pd.DataFrame(columns=cons_df.columns)
        agent_workable_filtered = pd.DataFrame(columns=cons_df.columns)

        # Apply remark filter (if any)
        if selected_remarks:
            remark_filtered = cons_df[
                cons_df["Remark"].astype(str).isin(selected_remarks)
            ]
        output_lines.append(
            f"\n🔎 Applied remark filter: {len(selected_remarks)} selected; rows after remark filter: {len(remark_filtered)}"
        )

        # Apply agent+Workable filter (if any)
        if selected_agents:
            agent_workable_filtered = cons_df[
                (cons_df.get("Remark", "").astype(str) == "Workable")
                & (cons_df.get("Agent Name", "").astype(str).isin(selected_agents))
            ]
        output_lines.append(
            f"🔎 Applied agent+Workable filter: {len(selected_agents)} selected; rows after agent filter: {len(agent_workable_filtered)}"
        )

        # Union of remark filter and agent-workable filter
        cons_df = pd.concat(
            [remark_filtered, agent_workable_filtered], ignore_index=True
        )
        output_lines.append(f"📦 Rows after union of filters: {len(cons_df)}")
        output_lines.append(
            f"\n📄 Source: 'All Agent Data' from consolidate ({len(cons_df)} rows)"
        )

        # Get the first sheet from blank allocation file (either "Today" or "Sheet1")
        blank_sheet_names = list(reallocation_blank_data.keys())
        if not blank_sheet_names:
            raise Exception("Blank Allocation File has no sheets")

        target_sheet_name = blank_sheet_names[0]
        blank_df = reallocation_blank_data[target_sheet_name]
        output_lines.append(
            f"📄 Target: '{target_sheet_name}' from blank allocation ({len(blank_df)} rows)"
        )

        # Align columns to the union of both
        combined_cols = list({*map(str, blank_df.columns), *map(str, cons_df.columns)})
        blank_aligned = blank_df.reindex(columns=combined_cols)
        cons_aligned = cons_df.reindex(columns=combined_cols)

        # Merge: append consolidate data to blank allocation data
        merged_df = pd.concat([blank_aligned, cons_aligned], ignore_index=True)
        merged_data[target_sheet_name] = merged_df

        output_lines.append(
            f"\n✅ Merged into '{target_sheet_name}': {len(merged_df)} total rows"
        )
        output_lines.append(f"   - From blank allocation: {len(blank_df)} rows")
        output_lines.append(
            f"   - From consolidate (All Agent Data): {len(cons_df)} rows"
        )

        # STEP 3: Remove duplicates based on composite key
        output_lines.append("\n" + "=" * 80)
        output_lines.append("STEP 3: REMOVING DUPLICATES")
        output_lines.append("=" * 80)

        rows_before_dedup = len(merged_df)

        # Check if required columns exist
        if (
            "Patient ID" not in merged_df.columns
            or "Dental Primary Ins Carr" not in merged_df.columns
        ):
            raise Exception(
                "Required columns not found: 'Patient ID' and/or 'Dental Primary Ins Carr'"
            )

        # Create composite key by concatenating Patient ID + Dental Primary Ins Carr
        # Store original index to restore order later
        merged_df["__original_index__"] = range(len(merged_df))
        merged_df["__composite_key__"] = merged_df["Patient ID"].astype(
            str
        ) + merged_df["Dental Primary Ins Carr"].astype(str)

        # Sort by composite key (groups duplicates together)
        merged_df_sorted = merged_df.sort_values(
            by="__composite_key__", na_position="last"
        )

        # Keep last occurrence (remove=first occurrences, keep=last)
        deduplicated_df = merged_df_sorted.drop_duplicates(
            subset=["__composite_key__"], keep="last"
        )

        # Restore original order
        deduplicated_df = deduplicated_df.sort_values(by="__original_index__")

        # Remove temporary columns (composite key and original index)
        deduplicated_df = deduplicated_df.drop(
            columns=["__composite_key__", "__original_index__"]
        )

        rows_after_dedup = len(deduplicated_df)
        rows_removed = rows_before_dedup - rows_after_dedup

        output_lines.append(f"\n📊 Deduplication Statistics:")
        output_lines.append(f"   - Rows before deduplication: {rows_before_dedup}")
        output_lines.append(f"   - Rows after deduplication: {rows_after_dedup}")
        output_lines.append(f"   - Duplicate rows removed: {rows_removed}")

        # STEP 4: Clear Agent Name except for Workable rows with selected agents
        if "Agent Name" in deduplicated_df.columns:
            mask_keep_agent = (deduplicated_df.get("Remark", "") == "Workable") & (
                deduplicated_df.get("Agent Name", "").astype(str).isin(selected_agents)
            )
            deduplicated_df.loc[~mask_keep_agent, "Agent Name"] = ""
            output_lines.append(
                f"\n🧹 Cleared 'Agent Name' except for Workable rows of selected agents (kept: {mask_keep_agent.sum()} rows)"
            )

        merged_data[target_sheet_name] = deduplicated_df

        # Add any other sheets from blank allocation file (keep as-is)
        for sheet_name, df in reallocation_blank_data.items():
            if sheet_name not in merged_data:
                merged_data[sheet_name] = df.copy()
                output_lines.append(
                    f"\n➕ Kept '{sheet_name}' from blank allocation ({len(df)} rows)"
                )

        reallocation_merged_data = merged_data

        output_lines.append("\n" + "=" * 80)
        output_lines.append("✅ GENERATION COMPLETE")
        output_lines.append("=" * 80)
        output_lines.append(f"Total merged sheets: {len(merged_data)}")
        output_lines.append(
            f"Rows in merged workbook: {sum(len(df) for df in merged_data.values())}"
        )
        output_lines.append("\nReady to download reallocation file.")

        reallocation_output = "\n".join(output_lines)
        reallocation_result = f"✅ Reallocation data generation complete successfully!\n\n{reallocation_output}"

        return redirect("/comparison?tab=reallocation")

    except Exception as e:
        reallocation_result = f"❌ Error processing reallocation: {str(e)}"
        reallocation_output = f"Error: {str(e)}"
        return redirect("/comparison?tab=reallocation")


@app.route("/download_reallocation", methods=["POST"])
def download_reallocation():
    global reallocation_merged_data, reallocation_consolidate_data

    try:
        # Prefer merged data; fallback to consolidate data
        data_to_write = reallocation_merged_data or reallocation_consolidate_data
        if not data_to_write:
            return "No reallocation data available", 400

        filename = request.form.get("filename", "reallocation_data.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        # Create Excel file with consolidate sheets
        output_path = os.path.join("/tmp", filename)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in data_to_write.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return send_file(output_path, as_attachment=True, download_name=filename)

    except Exception as e:
        return f"Error downloading file: {str(e)}", 500


@app.route("/reset_reallocation", methods=["POST"])
def reset_reallocation():
    global reallocation_consolidate_data, reallocation_blank_data, reallocation_consolidate_filename, reallocation_blank_filename, reallocation_result, reallocation_output, reallocation_merged_data, reallocation_available_remarks, reallocation_selected_remarks, reallocation_available_agents, reallocation_selected_agents

    try:
        reallocation_consolidate_data = None
        reallocation_blank_data = None
        reallocation_consolidate_filename = None
        reallocation_blank_filename = None
        reallocation_result = "🔄 Reallocation tool reset successfully! All files and data have been cleared."
        reallocation_output = ""
        reallocation_merged_data = None
        reallocation_available_remarks = []
        reallocation_selected_remarks = []
        reallocation_available_agents = []
        reallocation_selected_agents = []

        return redirect("/comparison?tab=reallocation")

    except Exception as e:
        reallocation_result = f"❌ Error resetting reallocation tool: {str(e)}"
        return redirect("/comparison?tab=reallocation")


# =============================
# General Comparison
# =============================


@app.route("/load_general_primary", methods=["POST"])
def load_general_primary():
    """Load the primary file, capture sheet/column metadata for UI selection."""
    global general_primary_data, general_primary_filename, general_comparison_updated_data, general_comparison_result, general_comparison_output

    try:
        if "primary_file" not in request.files:
            return jsonify({"ok": False, "error": "No primary file provided."}), 400

        primary_file = request.files["primary_file"]
        if not primary_file or primary_file.filename == "":
            return jsonify({"ok": False, "error": "Empty primary file."}), 400

        primary_filename = secure_filename(primary_file.filename)

        # Read Excel file directly from memory (same approach as other modules)
        primary_file.seek(0)  # Reset file pointer to beginning
        excel_data = pd.read_excel(primary_file, sheet_name=None, engine="openpyxl")

        # Remove "Unnamed:" columns from all sheets (same as other modules)
        cleaned_data = {}
        columns_by_sheet = {}
        for sheet_name, df in excel_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned
            columns_by_sheet[sheet_name] = [str(col) for col in df_cleaned.columns]

        general_primary_data = cleaned_data
        general_primary_filename = primary_filename
        general_comparison_updated_data = None
        general_comparison_result = None
        general_comparison_output = ""

        return jsonify(
            {
                "ok": True,
                "filename": general_primary_filename,
                "sheets": list(cleaned_data.keys()),
                "columns": columns_by_sheet,
            }
        )
    except Exception as e:
        error_msg = str(e)
        # Provide more helpful error messages for common file format issues
        if "No such file" in error_msg or "cannot open" in error_msg.lower():
            error_msg = f"File format error: {error_msg}. Please ensure the file is a valid Excel file (.xlsx or .xls)."
        elif "Unsupported format" in error_msg or "not supported" in error_msg.lower():
            error_msg = f"Unsupported file format: {error_msg}. Please save your file as .xlsx format."
        return jsonify({"ok": False, "error": error_msg}), 500


@app.route("/load_general_main", methods=["POST"])
def load_general_main():
    """Load the main dataset file, capture sheet/column metadata for UI selection."""
    global general_main_data, general_main_filename, general_comparison_updated_data, general_comparison_result, general_comparison_output

    try:
        if "main_file" not in request.files:
            return (
                jsonify({"ok": False, "error": "No main dataset file provided."}),
                400,
            )

        main_file = request.files["main_file"]
        if not main_file or main_file.filename == "":
            return jsonify({"ok": False, "error": "Empty main dataset file."}), 400

        main_filename = secure_filename(main_file.filename)

        # Read Excel file directly from memory (same approach as other modules)
        main_file.seek(0)  # Reset file pointer to beginning
        excel_data = pd.read_excel(main_file, sheet_name=None, engine="openpyxl")

        # Remove "Unnamed:" columns from all sheets (same as other modules)
        cleaned_data = {}
        columns_by_sheet = {}
        for sheet_name, df in excel_data.items():
            # Convert column names to strings first, then remove columns that start with "Unnamed:"
            df.columns = df.columns.astype(str)
            df_cleaned = df.loc[
                :, ~df.columns.str.contains("^Unnamed:", na=False, regex=True)
            ]
            cleaned_data[sheet_name] = df_cleaned
            columns_by_sheet[sheet_name] = [str(col) for col in df_cleaned.columns]

        general_main_data = cleaned_data
        general_main_filename = main_filename
        general_comparison_updated_data = None
        general_comparison_result = None
        general_comparison_output = ""

        return jsonify(
            {
                "ok": True,
                "filename": general_main_filename,
                "sheets": list(cleaned_data.keys()),
                "columns": columns_by_sheet,
            }
        )
    except Exception as e:
        error_msg = str(e)
        # Provide more helpful error messages for common file format issues
        if "No such file" in error_msg or "cannot open" in error_msg.lower():
            error_msg = f"File format error: {error_msg}. Please ensure the file is a valid Excel file (.xlsx or .xls)."
        elif "Unsupported format" in error_msg or "not supported" in error_msg.lower():
            error_msg = f"Unsupported file format: {error_msg}. Please save your file as .xlsx format."
        return jsonify({"ok": False, "error": error_msg}), 500


@app.route("/run_general_comparison", methods=["POST"])
def run_general_comparison():
    global general_primary_data, general_main_data, general_primary_filename, general_main_filename
    global general_comparison_result, general_comparison_output, general_comparison_updated_data

    try:
        output_lines = []
        general_comparison_output = ""
        general_comparison_updated_data = None

        if not general_primary_data or not general_main_data:
            general_comparison_result = (
                "❌ Please upload both Primary and Main Dataset files first."
            )
            return redirect("/comparison?tab=general")

        primary_sheet = request.form.get("primary_sheet")
        main_sheet = request.form.get("main_sheet")
        key_columns = request.form.getlist("gc_key_columns")
        update_columns = request.form.getlist("gc_update_columns")

        if not primary_sheet or primary_sheet not in general_primary_data:
            general_comparison_result = "❌ Primary sheet is missing or invalid."
            return redirect("/comparison?tab=general")
        if not main_sheet or main_sheet not in general_main_data:
            general_comparison_result = "❌ Main dataset sheet is missing or invalid."
            return redirect("/comparison?tab=general")
        if not key_columns:
            general_comparison_result = "❌ Please select at least one key column."
            return redirect("/comparison?tab=general")
        if not update_columns:
            general_comparison_result = (
                "❌ Please select at least one column to update."
            )
            return redirect("/comparison?tab=general")

        primary_df = general_primary_data[primary_sheet].copy()
        main_df = general_main_data[main_sheet].copy()

        # Check if dataframes are empty
        if primary_df.empty:
            general_comparison_result = (
                "❌ Primary sheet is empty. Please check your file."
            )
            return redirect("/comparison?tab=general")
        if main_df.empty:
            general_comparison_result = (
                "❌ Main dataset sheet is empty. Please check your file."
            )
            return redirect("/comparison?tab=general")

        # Normalize column names (convert to string, strip whitespace)
        primary_df.columns = [str(col).strip() for col in primary_df.columns]
        main_df.columns = [str(col).strip() for col in main_df.columns]

        # Validate columns exist (with helpful error messages showing available columns)
        for col in key_columns:
            if col not in primary_df.columns:
                available_cols = ", ".join(list(primary_df.columns)[:10])
                if len(primary_df.columns) > 10:
                    available_cols += f", ... ({len(primary_df.columns)} total)"
                general_comparison_result = f"❌ Key column '{col}' not found in primary sheet. Available columns: {available_cols}"
                return redirect("/comparison?tab=general")
            if col not in main_df.columns:
                available_cols = ", ".join(list(main_df.columns)[:10])
                if len(main_df.columns) > 10:
                    available_cols += f", ... ({len(main_df.columns)} total)"
                general_comparison_result = f"❌ Key column '{col}' not found in main dataset sheet. Available columns: {available_cols}"
                return redirect("/comparison?tab=general")
        for col in update_columns:
            if col not in main_df.columns:
                available_cols = ", ".join(list(main_df.columns)[:10])
                if len(main_df.columns) > 10:
                    available_cols += f", ... ({len(main_df.columns)} total)"
                general_comparison_result = f"❌ Update column '{col}' not found in main dataset sheet. Available columns: {available_cols}"
                return redirect("/comparison?tab=general")

        def normalize_val(v, is_patient_id=False):
            """Normalize a value for key matching. Use patient ID normalization for Patient ID columns."""
            if pd.isna(v):
                return ""

            # For Patient ID columns, use the specialized normalization function
            # This handles numeric IDs properly (removes .0, leading zeros, etc.)
            if is_patient_id:
                normalized = normalize_patient_id(v)
                if normalized is None:
                    return ""
                return str(normalized).strip()

            # For other columns, use standard normalization
            s = str(v).strip()
            s = " ".join(s.split())  # collapse internal whitespace
            return s.lower()

        def build_key(df):
            # Check if any key column is a Patient ID column (case-insensitive)
            is_pid_col = {}
            for col in key_columns:
                col_lower = str(col).lower().strip()
                is_pid_col[col] = (
                    "patient" in col_lower and "id" in col_lower
                ) or col_lower in ["pid", "pat id", "patientid"]

            return df[key_columns].apply(
                lambda row: "|".join(
                    normalize_val(row[col], is_patient_id=is_pid_col.get(col, False))
                    for col in key_columns
                ),
                axis=1,
            )

        try:
            primary_keys = build_key(primary_df)
            main_keys = build_key(main_df)
        except Exception as e:
            general_comparison_result = f"❌ Error building keys: {str(e)}"
            return redirect("/comparison?tab=general")

        # Debug: Show sample keys for investigation
        output_lines.append("\n" + "=" * 80)
        output_lines.append("DEBUG: KEY GENERATION")
        output_lines.append("=" * 80)
        output_lines.append(f"\nSample keys from PRIMARY file (first 5 rows):")
        for idx in list(primary_keys.index[:5]):
            key_val = primary_keys[idx]
            key_display = key_val if key_val else "(empty)"
            row_data = {col: primary_df.at[idx, col] for col in key_columns}
            output_lines.append(
                f"  Row {idx}: Key='{key_display}' | Raw values: {row_data}"
            )

        output_lines.append(f"\nSample keys from MAIN file (first 5 rows):")
        for idx in list(main_keys.index[:5]):
            key_val = main_keys[idx]
            key_display = key_val if key_val else "(empty)"
            row_data = {col: main_df.at[idx, col] for col in key_columns}
            output_lines.append(
                f"  Row {idx}: Key='{key_display}' | Raw values: {row_data}"
            )

        # Map from key to first occurrence index in main_df
        main_key_map = {}
        for idx, key in main_keys.items():
            if key and key not in main_key_map:  # Only add non-empty keys
                main_key_map[key] = idx

        output_lines.append(
            f"\nUnique keys in PRIMARY: {primary_keys.nunique()} (total rows: {len(primary_keys)})"
        )
        output_lines.append(
            f"Unique keys in MAIN: {main_keys.nunique()} (total rows: {len(main_keys)})"
        )
        output_lines.append(f"Unique keys in MAIN (non-empty): {len(main_key_map)}")

        # Ensure update columns exist in primary (add if missing)
        for col in update_columns:
            if col not in primary_df.columns:
                primary_df[col] = ""

        matched_rows = 0
        updated_cells = 0
        unmatched_keys = []
        matched_examples = []

        output_lines.append("\n" + "=" * 80)
        output_lines.append("DEBUG: MATCHING PROCESS")
        output_lines.append("=" * 80)

        for idx, key in primary_keys.items():
            if key and key in main_key_map:  # Only process non-empty keys
                matched_rows += 1
                main_idx = main_key_map[key]

                # Store first few matches as examples
                if len(matched_examples) < 3:
                    matched_examples.append(
                        {
                            "primary_row": idx,
                            "main_row": main_idx,
                            "key": key,
                            "updates": {},
                        }
                    )

                for col in update_columns:
                    try:
                        old_value = (
                            primary_df.at[idx, col]
                            if col in primary_df.columns
                            else None
                        )
                        value = (
                            main_df.at[main_idx, col]
                            if col in main_df.columns
                            else None
                        )
                        # Handle NaN values
                        if pd.isna(value):
                            value = ""
                        if pd.isna(old_value):
                            old_value = ""

                        # Store update info for examples
                        if (
                            len(matched_examples) > 0
                            and matched_examples[-1]["primary_row"] == idx
                        ):
                            matched_examples[-1]["updates"][col] = {
                                "old": str(old_value)[:50],
                                "new": str(value)[:50],
                            }

                        primary_df.at[idx, col] = value
                        updated_cells += 1
                    except Exception as e:
                        # Log error but continue with other columns
                        output_lines.append(
                            f"Warning: Error updating column '{col}' at row {idx}: {str(e)}"
                        )
            elif key:  # Non-empty key but no match found
                if len(unmatched_keys) < 5:  # Store first 5 unmatched keys
                    unmatched_keys.append(
                        {
                            "row": idx,
                            "key": key,
                            "raw_values": {
                                col: primary_df.at[idx, col] for col in key_columns
                            },
                        }
                    )

        # Show matching examples
        if matched_examples:
            output_lines.append(
                f"\n✅ Example matches (showing first {len(matched_examples)}):"
            )
            for ex in matched_examples:
                output_lines.append(
                    f"  Primary Row {ex['primary_row']} <-> Main Row {ex['main_row']}"
                )
                output_lines.append(f"    Key: '{ex['key']}'")
                for col, vals in ex["updates"].items():
                    output_lines.append(
                        f"    {col}: '{vals['old']}' -> '{vals['new']}'"
                    )

        # Show unmatched examples
        if unmatched_keys:
            output_lines.append(
                f"\n⚠️ Example unmatched keys from PRIMARY (showing first {len(unmatched_keys)}):"
            )
            for ex in unmatched_keys:
                output_lines.append(
                    f"  Row {ex['row']}: Key='{ex['key']}' | Raw values: {ex['raw_values']}"
                )
            if len(unmatched_keys) == 5:
                output_lines.append(f"  ... (showing first 5, there may be more)")

        # Remove duplicate rows based on Patient ID + Dental Primary Ins Carr
        rows_before_duplicate_removal = len(primary_df)
        output_lines.append("\n" + "=" * 80)
        output_lines.append("DUPLICATE REMOVAL PROCESS")
        output_lines.append("=" * 80)
        
        # Find Patient ID and Dental Primary Ins Carr columns (case-insensitive)
        patient_id_col = None
        dental_primary_col = None
        remark_col = None
        
        for col in primary_df.columns:
            col_lower = str(col).lower().strip()
            if not patient_id_col and (("patient" in col_lower and "id" in col_lower) or col_lower in ["pid", "pat id", "patientid"]):
                patient_id_col = col
            if not dental_primary_col and "dental primary ins carr" in col_lower:
                dental_primary_col = col
            if not remark_col and col_lower == "remark":
                remark_col = col
        
        if patient_id_col and dental_primary_col:
            output_lines.append(f"Patient ID column: {patient_id_col}")
            output_lines.append(f"Dental Primary Ins Carr column: {dental_primary_col}")
            if remark_col:
                output_lines.append(f"Remark column: {remark_col}")
            else:
                output_lines.append("Remark column: Not found (will skip remark-based filtering)")
            
            # Create a composite key for duplicate detection
            def create_duplicate_key(row):
                pid_val = str(row[patient_id_col]).strip() if pd.notna(row[patient_id_col]) else ""
                dental_val = str(row[dental_primary_col]).strip() if pd.notna(row[dental_primary_col]) else ""
                return f"{pid_val}|{dental_val}"
            
            primary_df['_duplicate_key'] = primary_df.apply(create_duplicate_key, axis=1)
            
            # Find duplicates (skip empty keys)
            duplicate_groups = primary_df.groupby('_duplicate_key')
            duplicate_keys = [key for key, group in duplicate_groups if len(group) > 1 and key.strip() != ""]
            
            rows_to_keep = []
            rows_to_remove = []
            
            if len(duplicate_keys) > 0:
                output_lines.append(f"\nFound {len(duplicate_keys)} duplicate key(s) (based on Patient ID + Dental Primary Ins Carr)")
                
                for dup_key in duplicate_keys:
                    dup_rows = primary_df[primary_df['_duplicate_key'] == dup_key]
                    dup_indices = dup_rows.index.tolist()
                    
                    # Check if any row has preferred remark values
                    preferred_row_idx = None
                    if remark_col:
                        for idx in dup_indices:
                            remark_val = str(primary_df.at[idx, remark_col]).strip().upper() if pd.notna(primary_df.at[idx, remark_col]) else ""
                            if remark_val in ["UPDATED", "QCP", "ASST"]:
                                preferred_row_idx = idx
                                break
                    
                    # If no preferred row found, keep the first one
                    if preferred_row_idx is None:
                        preferred_row_idx = dup_indices[0]
                    
                    # Mark rows to keep and remove
                    rows_to_keep.append(preferred_row_idx)
                    for idx in dup_indices:
                        if idx != preferred_row_idx:
                            rows_to_remove.append(idx)
                
                # Remove duplicate rows
                primary_df = primary_df.drop(rows_to_remove)
                rows_after_duplicate_removal = len(primary_df)
                
                output_lines.append(f"Rows before duplicate removal: {rows_before_duplicate_removal}")
                output_lines.append(f"Duplicate rows removed: {len(rows_to_remove)}")
                output_lines.append(f"Rows after duplicate removal: {rows_after_duplicate_removal}")
            else:
                output_lines.append("\nNo duplicate rows found based on Patient ID + Dental Primary Ins Carr")
            
            # Remove the temporary duplicate key column
            primary_df = primary_df.drop(columns=['_duplicate_key'], errors='ignore')
        else:
            missing_cols = []
            if not patient_id_col:
                missing_cols.append("Patient ID")
            if not dental_primary_col:
                missing_cols.append("Dental Primary Ins Carr")
            output_lines.append(f"\n⚠️ Warning: Could not find required columns for duplicate removal: {', '.join(missing_cols)}")
            output_lines.append("Skipping duplicate removal process.")

        # Build output workbook: replace only the selected primary sheet
        updated_workbook = {
            name: df.copy() for name, df in general_primary_data.items()
        }
        updated_workbook[primary_sheet] = primary_df
        general_comparison_updated_data = updated_workbook

        output_lines.insert(0, "=" * 80)
        output_lines.insert(1, "GENERAL COMPARISON RESULTS")
        output_lines.insert(2, "=" * 80)
        output_lines.insert(
            3,
            f"Primary file: {general_primary_filename or 'N/A'} | Sheet: {primary_sheet}",
        )
        output_lines.insert(
            4, f"Main dataset: {general_main_filename or 'N/A'} | Sheet: {main_sheet}"
        )
        output_lines.insert(5, f"Key columns: {', '.join(key_columns)}")
        output_lines.insert(6, f"Update columns: {', '.join(update_columns)}")
        output_lines.insert(7, f"Rows in primary sheet (before processing): {rows_before_duplicate_removal}")
        output_lines.insert(8, f"Rows in primary sheet (after duplicate removal): {len(primary_df)}")
        output_lines.insert(9, f"Rows in main sheet: {len(main_df)}")
        output_lines.insert(10, "")

        # Summary at the end
        output_lines.append("\n" + "=" * 80)
        output_lines.append("SUMMARY")
        output_lines.append("=" * 80)
        output_lines.append(f"Unique keys in primary: {primary_keys.nunique()}")
        output_lines.append(f"Unique keys in main: {len(main_key_map)} (non-empty)")
        output_lines.append(
            f"Matched rows (primary): {matched_rows} out of {len(primary_df)}"
        )
        output_lines.append(
            f"Match rate: {(matched_rows/len(primary_df)*100):.1f}%"
            if len(primary_df) > 0
            else "N/A"
        )
        output_lines.append(f"Total cells updated: {updated_cells}")

        if matched_rows == 0:
            output_lines.append("\n⚠️ WARNING: No rows were matched!")
            output_lines.append("\nTroubleshooting steps:")
            output_lines.append(
                "1. Check the DEBUG sections above to see sample keys from both files"
            )
            output_lines.append(
                "2. Verify key column values match exactly (after normalization)"
            )
            output_lines.append("3. Check for:")
            output_lines.append("   - Extra whitespace or hidden characters")
            output_lines.append("   - Case differences (keys are case-insensitive)")
            output_lines.append("   - Data type differences (numbers vs text)")
            output_lines.append("   - Missing or empty values in key columns")
            output_lines.append("4. Compare the 'Raw values' shown in DEBUG sections")
        else:
            output_lines.append(
                "\n✅ General comparison complete. Download the updated primary file."
            )

        general_comparison_output = "\n".join(output_lines)
        general_comparison_result = "✅ General comparison completed successfully."

        return redirect("/comparison?tab=general")

    except Exception as e:
        general_comparison_result = f"❌ Error during general comparison: {str(e)}"
        general_comparison_output = general_comparison_result
        return redirect("/comparison?tab=general")


@app.route("/reset_general", methods=["POST"])
def reset_general():
    """Reset all General Comparison state and return to the tab."""
    global general_primary_data, general_main_data, general_primary_filename, general_main_filename
    global general_comparison_result, general_comparison_output, general_comparison_updated_data

    try:
        general_primary_data = None
        general_main_data = None
        general_primary_filename = None
        general_main_filename = None
        general_comparison_result = (
            "🔄 General Comparison reset. All uploaded files and results cleared."
        )
        general_comparison_output = ""
        general_comparison_updated_data = None

        return redirect("/comparison?tab=general")
    except Exception as e:
        general_comparison_result = f"❌ Error resetting General Comparison: {str(e)}"
        return redirect("/comparison?tab=general")


@app.route("/download_general_comparison", methods=["POST"])
def download_general_comparison():
    global general_comparison_updated_data, general_primary_data

    try:
        data_to_write = general_comparison_updated_data or general_primary_data
        if not data_to_write:
            return "No general comparison data available", 400

        filename = request.form.get("filename", "general_comparison_output.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        output_path = os.path.join("/tmp", filename)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in data_to_write.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return send_file(output_path, as_attachment=True, download_name=filename)
    except Exception as e:
        return f"Error downloading general comparison file: {str(e)}", 500


if __name__ == "__main__":
    import os

    port = int(os.environ.get("PORT", 5002))
    # Enable debug mode for auto-reload during development
    debug = (
        os.environ.get("FLASK_ENV") == "development"
        or os.environ.get("FLASK_DEBUG", "True").lower() == "true"
    )

    print("🚀 Starting Excel Comparison Tool...")
    print(f"📱 Open your browser and go to: http://localhost:{port}")
    print(f"🔄 Debug mode: {debug} (auto-reload enabled)")
    app.run(debug=debug, host="0.0.0.0", port=port, use_reloader=True)
